# -*- coding: utf-8 -*-

# demand variability

import math
import datetime
import io
import numpy as np
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────
# 0.  PAGE CONFIG
# ─────────────────────────────────────────────────────────────
try:
    st.set_page_config(
        page_title="ITC Ecobyte — Supply Chain Planner",
        page_icon="📦",
        layout="wide",
        initial_sidebar_state="expanded",
    )
except:
    pass


# ─────────────────────────────────────────────────────────────
# 1.  EXCEL PARSER
# ─────────────────────────────────────────────────────────────
def _cell(ws, addr):
    v = ws[addr].value
    if isinstance(v, str) and v.startswith("="):
        return None
    return v


def _num(ws, addr, default=0):
    v = _cell(ws, addr)
    try:
        return float(v) if v is not None else default
    except:
        return default


def parse_sku_master(file_bytes: bytes) -> dict:
    """
    Parse the unified SKU Master Excel (3 sheets):
      Sheet 1 - SKU Basic Details
      Sheet 2 - BOM - RM Details
      Sheet 3 - Capacity Details
      Sheet 4 - WO Log
    Returns {mat_id → raw_dict} for all SKUs.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    # ── Sheet 1: Basic Details ──────────────────────────────
    s1 = wb["SKU Basic Details"]
    sku_basics = {}   # mat_id → row dict
    def _si(v, default=0):
        """Safe int parse — handles '-', None, empty."""
        try:
            s = str(v or "").strip().replace(",", "")
            return int(float(s)) if s and s not in ("-", "—", "") else default
        except Exception:
            return default

    def _sf(v, default=0.0):
        """Safe float parse."""
        try:
            s = str(v or "").strip().replace(",", "")
            return float(s) if s and s not in ("-", "—", "") else default
        except Exception:
            return default

    for row in s1.iter_rows(min_row=4, values_only=True):
        mat_id = str(row[0]).strip() if row[0] else None
        if not mat_id or mat_id == "None":
            continue
        _inhouse_flag = str(row[15] or "").strip().lower() if len(row) > 15 else "inhouse"
        if _inhouse_flag == "outsource":
            continue   # outsource SKUs handled separately via Outsource sheet
        sku_basics[mat_id] = {
            "mat_id":              mat_id,
            "mat_desc":            str(row[1] or "").strip(),
            "monthly_demand":      _si(row[2]),
            "working_days":        _si(row[3], 30),
            "safety_stock_fixed":  _si(row[4]),
            "z_score":             _sf(row[5], 1.65),
            "factory":             str(row[6] or "").strip(),
            "lt_variability":      _sf(row[7], 5),
            "target_fg":           _si(row[8]),
            "wastage_pct":         _sf(row[9], 15),
            "prod_moq":            _si(row[10], 30000),
            "fill_rate_hist":      _sf(row[11]) if row[11] else 0.0,
            "sl_hist":             _sf(row[12]) if row[12] else 0.0,
            "pouch_name":          str(row[13] or "-").strip(),
            "so_to_wo_days":       _si(row[14], 1) if len(row) > 14 and row[14] is not None else 1,
            # defaults populated later
            "current_fg":          0,
            "pouch_stock":         0,
            "pouch_per_1m":        20000,
            "pouch_lt":            3,
            "pouch_moq":           20000,
            "pouch_req":           0,
            "reel_lt_var":         None,
            "reel_order_cycle":    60,
            "bom_components":      [],
            "processes":           [],
            "wo_log":              [],
        }

    # ── Sheet 2: BOM - RM Details ───────────────────────────
    s2 = wb["BOM - RM Details"]
    for row in s2.iter_rows(min_row=5, values_only=True):
        mat_id = str(row[0]).strip() if row[0] else None
        if not mat_id or mat_id not in sku_basics:
            continue
        comp_name = str(row[1] or "Single").strip()
        rm_code   = str(row[2] or "").strip()
        reel_nm   = str(row[3] or "").strip()
        if not reel_nm:
            continue
        width   = float(row[4] or 0)
        length  = float(row[5] or 0)
        gsm     = float(row[6] or 0)
        ups     = int(row[7] or 1)
        wastage = float(row[8] or 15)
        moq_kg  = float(row[9] or 3000)
        rm_lt   = int(row[10] or 15)
        inbound = int(row[11] or 2)
        coating = int(row[12] or 5)
        mult    = float(row[13] or 1)   # Multiplier column
        notes   = str(row[14] or "").strip()
        # Col P (15) = Raw Material ID (SAP material number for RM stock matching)
        # Col Q (16) = Material Description
        rm_material_id   = str(row[15] or "").strip() if len(row) > 15 else ""
        rm_material_desc = str(row[16] or "").strip() if len(row) > 16 else ""
        reel_lt_var = int(row[17] or 5) if len(row) > 17 else 5
        sku_basics[mat_id]["bom_components"].append({
            "component":  comp_name,
            "rm_code":    rm_code,
            "reel_name":  reel_nm,
            "width":      width,
            "length":     length,
            "gsm":        gsm,
            "ups":        ups,
            "wastage":    wastage,
            "moq_kg":     moq_kg,
            "lt":         rm_lt,
            "inbound":    inbound,
            "coating":    coating,
            "total_lt":   rm_lt + inbound + coating,
            "multiplier":    mult,
            "notes":         notes,
            "reel_lt_var":   reel_lt_var,
            "rm_material_id":   rm_material_id,    # SAP RM material number (col P)
            "rm_material_desc": rm_material_desc,  # RM description (col Q)
        })

    # ── Sheet 3: Capacity Details ───────────────────────────
    s3 = wb["Capacity Details"]
    for row in s3.iter_rows(min_row=5, values_only=True):
        mat_id = str(row[0]).strip() if row[0] else None
        if not mat_id or mat_id not in sku_basics:
            continue
        proc_name = str(row[1] or "").strip()
        if not proc_name:
            continue
        daily_cap = float(row[2] or 999999)
        cap_unit  = str(row[3] or "Cartons").strip()
        shifts    = int(row[4] or 3)
        days      = int(row[5] or 1)
        sku_basics[mat_id]["processes"].append({
            "name":      proc_name,
            "capacity":  daily_cap,
            "cap_unit":  cap_unit,
            "shifts":    shifts,
            "shift_cap": round(daily_cap / max(shifts, 1)),
            "days":      days,
        })

    # ── Sheet 4: WO Log ─────────────────────────────────────
    if "WO Log" in wb.sheetnames:
        s4 = wb["WO Log"]
        for row in s4.iter_rows(min_row=3, values_only=True):
            mat_id = str(row[0]).strip() if row[0] else None
            if not mat_id or mat_id not in sku_basics:
                continue
            wo_num = row[1]
            if not wo_num:
                continue
            # Detect in-transit: WO Complete Date = "in production"
            complete_raw = str(row[7] or "").strip().lower() if len(row) > 7 else ""
            is_in_transit = "in production" in complete_raw
            sku_basics[mat_id]["wo_log"].append({
                "wo_num":       str(wo_num),
                "so_num":       str(row[2] or ""),
                "release":      row[3],           # keep as datetime object, NOT str
                "planned_lt":   int(row[4] or 0)  if len(row) > 4  else 0,
                "est_complete": row[5],            # keep as datetime object
                "target_qty":   int(row[6] or 0)  if len(row) > 6  else 0,
                "complete":     row[7],            # keep as datetime object (or "in production" str)
                "produced":     int(row[8] or 0)  if (len(row) > 8  and not is_in_transit and str(row[8] or "").strip() not in ("", "nan")) else 0,
                "wastage_pct":  float(row[9] or 0)  if len(row) > 9  else 0,
                "yield_pct":    float(row[10] or 0) if len(row) > 10 else 0,
                "variance_qty": int(row[11] or 0)  if len(row) > 11 else 0,
                "cycle_time":   int(row[12] or 0)  if len(row) > 12 else 0,
                "deviation":    int(row[13] or 0)  if len(row) > 13 else 0,
                "on_time":      str(row[14] or "").strip().upper() == "YES" if len(row) > 14 else False,
                "notes":        str(row[15] or "") if len(row) > 15 else "",
                "in_transit":   is_in_transit,
            })

    # ── Derive legacy fields from first BOM component ───────
    all_raw = {}
    for mat_id, s in sku_basics.items():
        bom = s["bom_components"]
        first = bom[0] if bom else {}
        prod_lt = int(sum(p["days"] for p in s["processes"]))
        # Legacy single-reel fields from first component
        # For SS/ROP we need reel_sheets expressed in terms of the primary reel's KG.
        # Since components sharing the same reel have DIFFERENT dimensions, we:
        #   1. Compute KG per component: sheets_i × width_i × length_i × gsm_i × (1+waste_i%) / 1e9
        #   2. Sum KG across all components sharing the same reel
        #   3. Back-derive an equivalent "sheet count" from the primary reel's dimensions
        #      so the existing SS/ROP formula (which converts sheets→KG) stays correct.
        _target     = s["target_fg"]
        _first_reel = first.get("reel_name", "")
        _first_w    = first.get("width",   1)
        _first_l    = first.get("length",  1)
        _first_gsm  = first.get("gsm",   120)
        _first_wst  = first.get("wastage", 15) / 100

        # Total KG needed from this reel to produce MTS quantity
        _total_kg = sum(
            math.ceil(
                math.ceil(_target / comp["ups"])
                * comp["width"] * comp["length"]
                * comp["gsm"] / ((1 - comp["wastage"] / 100)
                * 1_000_000_000)
            )
            for comp in bom
            if comp.get("reel_name") == _first_reel and comp.get("ups", 0) > 0
        ) if _first_reel else 0

        # Back-derive equivalent sheet count from primary reel's dimensions
        # so reel_sheets_to_kg = _reel_sheets × w × l × gsm × (1+waste) / 1e9 ≈ _total_kg
        _kg_per_sheet_primary = (
            _first_w * _first_l * _first_gsm / ((1 - _first_wst) * 1_000_000_000)
        )
        _reel_sheets = (
            math.ceil(_total_kg / _kg_per_sheet_primary)
            if _kg_per_sheet_primary > 0 else 0
        )

        s.update({
            "prod_lt":         prod_lt + s.get("so_to_wo_days", 1),
            "so_to_wo_days":   s.get("so_to_wo_days", 1),
            "number_of_ups":   first.get("ups", 9),
            "reel_name":       _first_reel,
            "reel_stock":      0,
            "reel_moq":        first.get("moq_kg", 3000),
            "reel_lt":         first.get("lt", 15),
            "reel_inbound":    first.get("inbound", 2),
            "reel_coating":    first.get("coating", 5),
            "reel_total_lt":   first.get("total_lt", 22),
            "reel_sheets":     _reel_sheets,
            "reel_ups":        first.get("ups", 9),
            "reel_gsm":        _first_gsm,
            "reel_waste":      first.get("wastage", 15),
            "reel_width":      _first_w,
            "reel_length":     _first_l,
            # per-reel lt_var map built from BOM components
            "reel_lt_var_map": {
                comp["reel_name"]: comp.get("reel_lt_var", 5)
                for comp in bom
            },
            # legacy scalar — lt_var of primary reel (for pouch SS etc.)
            "reel_lt_var": first.get("reel_lt_var", 5),
        })
        all_raw[mat_id] = s

    # ── Outsource sheet ───────────────────────────────────────
    outsource_skus: list[dict] = []
    if "Outsource" in wb.sheetnames:
        s_os = wb["Outsource"]
        rows_os = list(s_os.values)

        def _safe_int(v, default=0):
            try:
                s = str(v or "").strip().replace(",", "")
                return int(float(s)) if s and s not in ("-", "—", "N/A", "") else default
            except Exception:
                return default

        def _safe_float(v, default=0.0):
            try:
                s = str(v or "").strip().replace(",", "")
                return float(s) if s and s not in ("-", "—", "N/A", "") else default
            except Exception:
                return default

        for row in rows_os[1:]:   # row 0 = headers
            mat_id = str(row[0] or "").strip()
            if not mat_id or mat_id.lower().startswith("material"):
                continue
            outsource_skus.append({
                "mat_id":         mat_id,
                "mat_desc":       str(row[1] or "").strip(),
                "monthly_demand": _safe_int(row[2]),
                "z_score":        _safe_float(row[4], 1.65),
                "factory":        str(row[5] or "").strip(),
                "place":          str(row[9] or "").strip(),
                "lt_var":         _safe_int(row[10], 3),
                "inbound_lt":     _safe_int(row[11], 10),
                "safety_stock":   _safe_int(row[13]),
                "rop":            _safe_int(row[14]),
            })

    # ── Outbound / DC Transit Times sheet ────────────────────
    dc_transit: dict[str, dict] = {}   # {factory → {destination → transit_days}}
    _outbound_sheet_names = [s for s in wb.sheetnames
                             if any(kw in s.lower() for kw in
                                    ("outbound", "dc", "transit time", "delivery time", "dispatch"))]
    for sheet_name in _outbound_sheet_names:
        s_dc = wb[sheet_name]
        rows_dc = list(s_dc.values)
        if not rows_dc:
            continue
        # Row 0 = header (Factory / Origin | Location | Avg Days)
        # Data from row 1 onwards
        for row in rows_dc[1:]:
            cells = [str(c or "").strip() for c in row]
            if len(cells) < 3 or not cells[0]:
                continue
            factory = cells[0]
            dest    = cells[1]
            try:
                days = int(float(cells[2])) if cells[2] else 0
            except Exception:
                days = 0
            if factory and dest:
                if factory not in dc_transit:
                    dc_transit[factory] = {}
                dc_transit[factory][dest] = days

    # Store DC transit on all SKUs
    for mat_id, s in all_raw.items():
        s["dc_transit"] = dc_transit

    # Store outsource list globally accessible via any raw entry
    for mat_id, s in all_raw.items():
        s["outsource_skus"] = outsource_skus

    return all_raw   # {mat_id → raw_dict}


def get_dc_transit_days(factory: str, destination: str, all_raw: dict) -> int:
    """Look up transit days from factory to destination DC."""
    for raw in all_raw.values():
        transit = raw.get("dc_transit", {})
        if factory in transit and destination in transit[factory]:
            return transit[factory][destination]
        # Case-insensitive fallback
        for f, dests in transit.items():
            if f.lower() == factory.lower():
                for d, days in dests.items():
                    if d.lower() == destination.lower():
                        return days
    return 0
    """
    Parse FG stock list (xls/xlsx).
    - Tries all sheets, picks the one where col A contains material IDs (has digits + hyphen pattern)
    - Header row is found by looking for 'Material' in col A
    - Stock quantity is always col H (index 7)
    - Sums quantity across multiple rows for the same material ID (different batches)
    Returns {mat_id → total_fg_stock}.
    """
    # Load all sheets
    sheets = {}
    for engine in ("openpyxl", None):
        try:
            kwargs = {"sheet_name": None, "header": None}
            if engine:
                kwargs["engine"] = engine
            sheets = pd.read_excel(io.BytesIO(file_bytes), **kwargs)
            break
        except Exception:
            continue

    if not sheets:
        return {}

    def _looks_like_mat_id(val: str) -> bool:
        """True if string looks like a material ID — has digits and is not a pure sentence."""
        v = val.strip()
        return len(v) >= 5 and any(c.isdigit() for c in v) and len(v) < 50

    def _parse_sheet(sdf: pd.DataFrame) -> dict:
        """Parse a single sheet dataframe — returns {} if it doesn't look right."""
        if sdf is None or sdf.empty or sdf.shape[1] < 8:
            return {}

        # Find header row: col A (index 0) == 'Material' (case-insensitive)
        hdr_row = None
        for i, row in sdf.iterrows():
            v = str(row.iloc[0]).strip().lower()
            if v == "material":
                hdr_row = i
                break

        # If no explicit header, check if col A row 0 or 1 has a title/header
        if hdr_row is None:
            # Try to detect data directly: look for first row with a mat-ID-like value in col A
            for i, row in sdf.iterrows():
                if _looks_like_mat_id(str(row.iloc[0])):
                    hdr_row = i - 1  # assume header is one row above first data
                    break
            if hdr_row is None or hdr_row < 0:
                hdr_row = 1  # fallback

        # Data starts one row after header
        data_start = hdr_row + 1
        if data_start >= len(sdf):
            return {}

        totals: dict[str, float] = {}
        for idx in range(data_start, len(sdf)):
            row = sdf.iloc[idx]
            mat = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if not mat or mat.lower() in ("nan", "none", "material", "") or not _looks_like_mat_id(mat):
                continue
            try:
                qty = float(row.iloc[7]) if pd.notna(row.iloc[7]) else 0
            except Exception:
                qty = 0
            if qty > 0:
                totals[mat] = totals.get(mat, 0) + qty

        return totals

    # ── Try all sheets, return the first one that produces results ────
    # Prefer later sheets (sheet 2 etc.) since sheet 1 is often a summary/info page.
    sheet_list = list(sheets.items())
    # Try from last sheet to first, so sheet 2 beats sheet 1
    for name, sdf in reversed(sheet_list):
        result = _parse_sheet(sdf)
        if result:
            return result

    # Last resort: try all in order
    for name, sdf in sheet_list:
        result = _parse_sheet(sdf)
        if result:
            return result

    return {}


def parse_fg_inventory(file_bytes: bytes) -> dict:
    """
    Parse FG stock list (xls/xlsx).
    Tries all sheets; looks for header row with 'Material' in col A,
    stock in col H (index 7). Sums across multiple rows per material ID.
    Returns {mat_id → total_fg_stock}.
    """
    sheets = {}
    for engine in ("openpyxl", None):
        try:
            kwargs = {"sheet_name": None, "header": None}
            if engine:
                kwargs["engine"] = engine
            sheets = pd.read_excel(io.BytesIO(file_bytes), **kwargs)
            break
        except Exception:
            continue
    if not sheets:
        return {}

    def _looks_like_mat_id(val: str) -> bool:
        v = val.strip()
        return len(v) >= 5 and any(c.isdigit() for c in v) and len(v) < 50

    def _parse_sheet(sdf):
        if sdf is None or sdf.empty or sdf.shape[1] < 8:
            return {}
        hdr_row = None
        for i, row in sdf.iterrows():
            v = str(row.iloc[0]).strip().lower()
            if v == "material":
                hdr_row = i
                break
        if hdr_row is None:
            for i, row in sdf.iterrows():
                if _looks_like_mat_id(str(row.iloc[0])):
                    hdr_row = max(0, i - 1)
                    break
        if hdr_row is None:
            hdr_row = 1

        # Find stock column: "Total Stock" (not "value")
        hdr = sdf.iloc[hdr_row]
        stock_col = None
        for c, v in enumerate(hdr):
            if "total stock" in str(v).strip().lower() and "value" not in str(v).strip().lower():
                stock_col = c
                break
        if stock_col is None:
            stock_col = 7

        totals: dict[str, float] = {}
        for idx in range(hdr_row + 1, len(sdf)):
            row = sdf.iloc[idx]
            mat = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if not mat or mat.lower() in ("nan", "none", "material", "") or not _looks_like_mat_id(mat):
                continue
            try:
                qty = float(row.iloc[stock_col]) if pd.notna(row.iloc[stock_col]) else 0
            except Exception:
                qty = 0
            if qty > 0:
                totals[mat] = totals.get(mat, 0) + qty
        return totals

    # Try sheets in reverse (detail sheets last)
    for _, sdf in reversed(list(sheets.items())):
        result = _parse_sheet(sdf)
        if result:
            return result
    for _, sdf in sheets.items():
        result = _parse_sheet(sdf)
        if result:
            return result
    return {}


def parse_rm_inventory(file_bytes: bytes) -> dict:
    """
    Parse RM inventory file (Ageing Report / stock list).
    Handles the SAP Ageing Report format with 'Material' and
    'Total Stock(Batch-wise)' columns, or any sheet with Material+Total Stock headers.
    Returns {material_id -> total_stock_kg}.
    """
    sheets = {}
    for engine in ("openpyxl", None):
        try:
            kwargs = {"sheet_name": None, "header": None}
            if engine:
                kwargs["engine"] = engine
            sheets = pd.read_excel(io.BytesIO(file_bytes), **kwargs)
            break
        except Exception:
            continue
    if not sheets:
        return {}

    def _parse_rm_sheet(sdf):
        if sdf is None or sdf.empty:
            return {}
        # Find header row: any row containing 'material' in col values
        hdr_row = None
        for i, row in sdf.iterrows():
            vals_lower = [str(v).strip().lower() for v in row if pd.notna(v)]
            if "material" in vals_lower:
                hdr_row = i
                break
        if hdr_row is None:
            return {}
        hdr = [str(v).strip() if pd.notna(v) else f"col{c}"
               for c, v in enumerate(sdf.iloc[hdr_row])]
        data = sdf.iloc[hdr_row + 1:].copy()
        data.columns = hdr
        data = data.reset_index(drop=True)

        mat_col = next((h for h in hdr if h.lower() == "material"), None)
        # Prefer "Total Stock(Batch-wise)" then any "total stock" without "value"
        stk_col = next((h for h in hdr if "total stock(batch" in h.lower()), None)
        if stk_col is None:
            stk_col = next((h for h in hdr
                            if "total stock" in h.lower() and "value" not in h.lower()), None)
        if mat_col is None or stk_col is None:
            return {}

        totals: dict[str, float] = {}
        for _, row in data.iterrows():
            mat = str(row[mat_col]).strip() if pd.notna(row[mat_col]) else ""
            if not mat or mat.lower() in ("nan", "none", "material", ""):
                continue
            try:
                qty = float(row[stk_col]) if pd.notna(row[stk_col]) else 0
            except Exception:
                qty = 0
            if qty > 0:
                totals[mat] = totals.get(mat, 0) + qty
        return totals

    # Try sheets in reverse (detail sheets are usually last)
    for _, sdf in reversed(list(sheets.items())):
        result = _parse_rm_sheet(sdf)
        if result:
            return result
    for _, sdf in sheets.items():
        result = _parse_rm_sheet(sdf)
        if result:
            return result
    return {}


def parse_po_history(file_bytes: bytes) -> pd.DataFrame:
    """
    Parse PO history Excel.
    Uses Customer PO Date as the demand date (reflects when customer placed the order).
    Columns detected by name with positional fallback:
      Customer PO Date → date, Material Code → mat_id, Billed quantity → qty
    Returns DataFrame with columns: mat_id, date, qty
    """
    df = pd.read_excel(io.BytesIO(file_bytes), header=0)

    # Find columns by name (case-insensitive)
    col_lower = {str(c).strip().lower(): i for i, c in enumerate(df.columns)}

    def _find_col(*candidates):
        for c in candidates:
            if c.lower() in col_lower:
                return col_lower[c.lower()]
        return None

    # Customer PO Date is the primary date — falls back to Bill.Date
    idx_date = _find_col("customer po date", "cust. po date", "po date",
                          "bill.date", "billdate", "bill date")
    idx_mat  = _find_col("material code", "material", "mat_id", "sku")
    idx_qty  = _find_col("billed quantity", "billed qty", "quantity", "qty")

    # Positional fallback
    if idx_date is None: idx_date = 13   # Customer PO Date is col 14 (0-indexed=13)
    if idx_mat  is None: idx_mat  = 6
    if idx_qty  is None: idx_qty  = 8

    df.columns = range(df.shape[1])
    out = []
    for _, row in df.iterrows():
        try:
            mat_id = str(row[idx_mat]).strip()
            date   = pd.to_datetime(row[idx_date], dayfirst=True, errors="coerce")
            qty    = float(row[idx_qty]) if pd.notna(row[idx_qty]) else 0
            if mat_id and mat_id not in ("nan", "None", "") and pd.notna(date) and qty > 0:
                out.append({"mat_id": mat_id, "date": date, "qty": qty})
        except:
            continue
    return pd.DataFrame(out) if out else pd.DataFrame(columns=["mat_id", "date", "qty"])


def parse_demand_events_excel(file_bytes: bytes) -> pd.DataFrame:
    """
    Parse the demand events / billing Excel (same format as PO history export).
    Columns detected by name:
      Material Code → mat_id
      Billed quantity → qty
      Customer PO Date → date  (used as demand date)
      Bill.No → invoice_no
      GST Invoice No. → gst_invoice_no
      Transporter Name → transporter
      LR Number → lr_number
      Ship to Party Name → ship_to_name
      Ship to Destination → destination
      Sold-to-Party Name → sold_to_name
      Customer PO No. → customer_po_no
      Bill.Date → bill_date
      SO Date → so_date
    Returns a DataFrame with all relevant columns retained.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), header=0)
    except Exception:
        return pd.DataFrame()

    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]
    col_map = {c.lower(): c for c in df.columns}

    def _col(*candidates):
        for cand in candidates:
            if cand.lower() in col_map:
                return col_map[cand.lower()]
        return None

    c_mat  = _col("Material Code", "material code", "material", "mat_id")
    c_qty  = _col("Billed quantity", "billed qty", "quantity", "qty")
    c_date = _col("Customer PO Date", "cust. po date", "po date", "customer po date")
    c_bill_no   = _col("Bill.No", "bill no", "billno")
    c_gst       = _col("GST Invoice No.", "gst invoice no", "gst invoice no.")
    c_trans     = _col("Transporter Name", "transporter name", "transporter")
    c_lr        = _col("LR Number", "lr number", "lr no")
    c_ship_name = _col("Ship to Party Name", "ship-to-party name", "ship to party name")
    c_dest      = _col("Ship to Destination", "ship to destination", "destination")
    c_sold_name = _col("Sold-to-Party Name", "sold-to-party name", "sold to party name")
    c_cpo       = _col("Customer PO No.", "customer po no", "customer po no.")
    c_bill_date = _col("Bill.Date", "bill date", "billdate")
    c_so_date   = _col("SO Date", "so date", "sodate")
    c_so_num    = _col("Sales Doc.", "sales doc", "so num", "so#")

    out_rows = []
    for _, row in df.iterrows():
        try:
            mat_id = str(row[c_mat]).strip() if c_mat else ""
            if not mat_id or mat_id in ("nan", "None", ""):
                continue
            qty = float(row[c_qty]) if c_qty and pd.notna(row[c_qty]) else 0
            date = pd.to_datetime(row[c_date], dayfirst=True, errors="coerce") if c_date else pd.NaT
            bill_date = pd.to_datetime(row[c_bill_date], dayfirst=True, errors="coerce") if c_bill_date else pd.NaT
            so_date   = pd.to_datetime(row[c_so_date],   dayfirst=True, errors="coerce") if c_so_date   else pd.NaT
            # Use Customer PO Date as demand date; fall back to bill date
            demand_date = date if pd.notna(date) else bill_date
            out_rows.append({
                "mat_id":      mat_id,
                "qty":         qty,
                "date":        demand_date,
                "bill_date":   bill_date,
                "so_date":     so_date,
                "invoice_no":  str(row[c_bill_no]).strip()  if c_bill_no  and pd.notna(row[c_bill_no])  else "",
                "gst_invoice": str(row[c_gst]).strip()      if c_gst      and pd.notna(row[c_gst])      else "",
                "transporter": str(row[c_trans]).strip()    if c_trans    and pd.notna(row[c_trans])    else "",
                "lr_number":   str(row[c_lr]).strip()       if c_lr       and pd.notna(row[c_lr])       else "",
                "ship_to":     str(row[c_ship_name]).strip()if c_ship_name and pd.notna(row[c_ship_name]) else "",
                "destination": str(row[c_dest]).strip()     if c_dest     and pd.notna(row[c_dest])     else "",
                "sold_to":     str(row[c_sold_name]).strip()if c_sold_name and pd.notna(row[c_sold_name]) else "",
                "customer_po": str(row[c_cpo]).strip()      if c_cpo      and pd.notna(row[c_cpo])      else "",
                "so_num":      str(row[c_so_num]).strip()   if c_so_num   and pd.notna(row[c_so_num])   else "",
            })
        except Exception:
            continue

    if not out_rows:
        return pd.DataFrame()
    return pd.DataFrame(out_rows)


# ─────────────────────────────────────────────────────────────
# OUTPUT EXCEL WRITER
# ─────────────────────────────────────────────────────────────
OUTPUT_EXCEL_PATH = "ecobyte_output.xlsx"

def write_output_excel(all_raw: dict, all_issued_wos: dict,
                       all_demand_events: dict, bom_req: dict,
                       reel_registry: dict, reel_pos_released: dict):
    """
    4-sheet snapshot report:
      1. FG Status       — current FG vs targets for all SKUs
      2. WOs In Transit  — all WOs currently in production (historical + in-transit)
      3. RM Available    — current stock per RM (from registry)
      4. RM In Transit   — all RM POs currently in production
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    today = datetime.date.today()

    # ── Helpers ──────────────────────────────────────────────
    def _hdr(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1E3A5F")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = ws["A2"]

    def _auto_width(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 45)

    def _d(v):
        """Format date as dd-mm-yyyy for Excel output."""
        if v is None: return ""
        if isinstance(v, (datetime.date, datetime.datetime)):
            return v.strftime("%d-%m-%Y")
        s = str(v).strip()
        # Try to parse and reformat any date string
        for _fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d %b %Y", "%d-%b-%Y"):
            try:
                return datetime.datetime.strptime(s[:10], _fmt).strftime("%d-%m-%Y")
            except Exception:
                pass
        return s  # return as-is if unparseable (e.g. "in transit")

    # ── Sheet 1: FG Status ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "FG Status"
    _hdr(ws1, [
        "Material ID", "Description", "Factory",
        "Current FG", "MTS Target", "Safety Stock", "Max FG",
        "ROP", "Prod MOQ", "Gap to MTS", "Status",
    ])
    for mat_id, raw in all_raw.items():
        cfg    = raw.get("current_fg", 0)
        tgt    = raw.get("target_fg", 0)
        ss     = raw.get("safety_stock_fixed", 0)
        moq    = raw.get("prod_moq", 0)
        max_fg = tgt + ss
        rop    = max(ss, max_fg - moq)
        gap    = max(0, tgt - cfg)
        status = "Below ROP" if cfg <= rop else ("Below MTS" if cfg < tgt else "OK")
        ws1.append([
            mat_id, raw.get("mat_desc",""), raw.get("factory",""),
            cfg, tgt, ss, max_fg, rop, moq, gap, status,
        ])
    _auto_width(ws1)

    # ── Sheet 2: WOs In Transit ──────────────────────────────
    ws2 = wb.create_sheet("WOs In Transit")
    _hdr(ws2, [
        "Material ID", "Description", "WO #", "SO #",
        "Release Date", "Planned LT (days)", "Expected Arrival",
        "Target Qty", "Status",
    ])
    for mat_id, raw in all_raw.items():
        issued = all_issued_wos.get(mat_id, {})
        for ds, info in sorted(issued.items()):
            if not info.get("issued"):
                continue
            try:
                rel = datetime.date.fromisoformat(ds)
                lt  = int(info.get("actual_lt_override", raw.get("prod_lt", 10)))
                arr = rel + datetime.timedelta(days=lt)
            except Exception:
                continue
            # Only in-transit: arrival in the future, or flagged as in_transit
            is_it = info.get("in_transit") or arr > today
            if not is_it:
                continue
            status = "🏭 In Production" if arr > today else "✅ Arrived"
            ws2.append([
                mat_id, raw.get("mat_desc", ""),
                info.get("wo_num", ""), info.get("so_num", ""),
                _d(rel), lt, _d(arr),
                info.get("wo_qty", 0), status,
            ])
        # Also add WO Log in-transit entries not yet in issued_wos
        for wo in raw.get("wo_log", []):
            if not wo.get("in_transit"):
                continue
            rel_raw = wo.get("release")
            est_raw = wo.get("est_complete")
            if isinstance(rel_raw, (datetime.date, datetime.datetime)):
                rel = rel_raw.date() if isinstance(rel_raw, datetime.datetime) else rel_raw
            else:
                try: rel = pd.to_datetime(str(rel_raw), dayfirst=True).date()
                except: continue
            if isinstance(est_raw, (datetime.date, datetime.datetime)):
                arr = est_raw.date() if isinstance(est_raw, datetime.datetime) else est_raw
            else:
                try: arr = pd.to_datetime(str(est_raw), dayfirst=True).date()
                except: arr = None
            ws2.append([
                mat_id, raw.get("mat_desc", ""),
                str(wo.get("wo_num","")), str(wo.get("so_num","")),
                _d(rel), wo.get("planned_lt", ""),
                _d(arr) if arr else "—",
                wo.get("target_qty", 0), "🏭 In Production (WO Log)",
            ])
    _auto_width(ws2)

    # ── Sheet 3: Customer PO & Delivery (pending only) ───────
    ws3 = wb.create_sheet("Customer PO & Delivery (Pending)")
    _hdr(ws3, [
        "Customer PO No.", "PO Date", "SKU Material ID", "SKU Description",
        "Qty (Cartons)", "Invoice No.", "GST Invoice No.",
        "Delivery Partner", "LR Number", "Ship-to Party", "Destination", "Factory",
        "No. of Days to TCI", "Factory Pickup", "Actual TCI Delivery",
        "Scheduled HP Delivery", "Actual HP Delivery",
    ])

    # Build bill→tracking_row lookup — apply any pending delivery updates first
    _bill_track: dict = {}
    _pending_bills: set | None = None
    try:
        _dtk_snap = st.session_state.get("_del_track_df", pd.DataFrame()).copy()
        # Apply any pending arrival updates that haven't been saved yet
        _pending_updates = st.session_state.get("del_arrival_updates", {})
        if not _dtk_snap.empty and _pending_updates:
            _col_tci_s = next((c for c in _dtk_snap.columns if "actual tci" in c.lower()), None)
            _col_hp_s  = next((c for c in _dtk_snap.columns if "actual hyperpure" in c.lower()), None)
            for _pu_idx, _pu_upd in _pending_updates.items():
                if not isinstance(_pu_upd, dict) or _pu_idx >= len(_dtk_snap):
                    continue
                if _pu_upd.get("action") == "tci" and _col_tci_s:
                    _dtk_snap.at[_pu_idx, _col_tci_s] = _pu_upd["date"]
                    if _col_hp_s:
                        _hp_c = str(_dtk_snap.at[_pu_idx, _col_hp_s]).strip().lower()
                        if _hp_c in ("", "not arrived"):
                            _dtk_snap.at[_pu_idx, _col_hp_s] = "in transit"
                elif _pu_upd.get("action") == "hp" and _col_hp_s:
                    _dtk_snap.at[_pu_idx, _col_hp_s] = _pu_upd["date"]
                    if _col_tci_s and str(_dtk_snap.at[_pu_idx, _col_tci_s]).strip().lower() == "in transit":
                        _dtk_snap.at[_pu_idx, _col_tci_s] = _pu_upd["date"]

        if not _dtk_snap.empty:
            _cb = next((c for c in _dtk_snap.columns if "bill.no" in c.lower()), None)
            _cd = next((c for c in _dtk_snap.columns if "no. of days" in c.lower() or "days to tci" in c.lower()), None)
            _cf = next((c for c in _dtk_snap.columns if "factory pickup" in c.lower()), None)
            _ct = next((c for c in _dtk_snap.columns if "actual tci" in c.lower()), None)
            _cs = next((c for c in _dtk_snap.columns if "scheduled hyperpure" in c.lower()), None)
            _ch = next((c for c in _dtk_snap.columns if "actual hyperpure" in c.lower()), None)

            def _dv(v):
                """Format a tracking cell value — date as dd-mm-yyyy, text preserved."""
                if v is None: return ""
                if isinstance(v, (datetime.date, datetime.datetime)):
                    return v.strftime("%d-%m-%Y")
                s = str(v).strip()
                if s.lower() in ("in transit", "not arrived", ""):
                    return s
                for _fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%Y-%m-%d"):
                    try:
                        return datetime.datetime.strptime(s[:10], _fmt).strftime("%d-%m-%Y")
                    except Exception:
                        pass
                return s

            if _cb:
                # Build pending set: bills where HP is still blank/in transit/not arrived
                _col_hp_chk = _ch
                if _col_hp_chk:
                    _pending_bills = set(
                        str(r[_cb]) for _, r in _dtk_snap.iterrows()
                        if str(r.get(_col_hp_chk, "") or "").strip().lower()
                           in ("", "in transit", "not arrived")
                    )
                # One tracking entry per bill (first row wins for header fields)
                for _, _tr in _dtk_snap.iterrows():
                    _bk = str(_tr[_cb])
                    if _bk not in _bill_track:
                        _bill_track[_bk] = {
                            "days_tci": _dv(_tr.get(_cd)) if _cd else "",
                            "fac_pu":   _dv(_tr.get(_cf)) if _cf else "",
                            "tci_act":  _dv(_tr.get(_ct)) if _ct else "",
                            "hp_sched": _dv(_tr.get(_cs)) if _cs else "",
                            "hp_act":   _dv(_tr.get(_ch)) if _ch else "",
                        }
    except Exception:
        pass

    for mat_id, ev_list in all_demand_events.items():
        raw_sku  = all_raw.get(mat_id, {})
        factory  = raw_sku.get("factory", "")
        mat_desc = raw_sku.get("mat_desc", "")
        for ev in ev_list:
            _inv_no = str(ev.get("invoice_no", "") or "")
            if _pending_bills is not None and _inv_no and _inv_no not in _pending_bills:
                continue
            _tr = _bill_track.get(_inv_no, {})
            ws3.append([
                ev.get("customer_po", ""),
                _d(ev.get("date")),
                mat_id, mat_desc,
                int(ev.get("qty", 0)),
                _inv_no,
                ev.get("gst_invoice", ""),
                ev.get("transporter", ""),
                ev.get("lr_number", ""),
                ev.get("ship_to", ""),
                ev.get("destination", ""),
                factory,
                _tr.get("days_tci", ""),
                _tr.get("fac_pu",   ""),
                _tr.get("tci_act",  ""),
                _tr.get("hp_sched", ""),
                _tr.get("hp_act",   ""),
            ])
    _auto_width(ws3)

    wb.save(OUTPUT_EXCEL_PATH)
    # Also return as bytes for direct download
    from io import BytesIO as _BytesIO
    _buf = _BytesIO()
    wb.save(_buf)
    return _buf.getvalue()


def _cell_val(v):
    if isinstance(v, str) and v.startswith("="):
        return None
    return v


# ─────────────────────────────────────────────────────────────
# 2.  HELPERS
# ─────────────────────────────────────────────────────────────
def get_months_ahead(n=6):
    today = datetime.date.today()
    months = []
    for i in range(n):
        m = today.month + i
        y = today.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        start = datetime.date(y, m, 1)
        end   = (datetime.date(y, m + 1, 1) if m < 12
                 else datetime.date(y + 1, 1, 1)) - datetime.timedelta(days=1)
        months.append((start.strftime("%b %Y"), start, end))
    return months


def _safe_date_str(d) -> str:
    if d is None:
        return ""
    if hasattr(d, "isoformat"):
        return d.isoformat()
    return str(d)[:10]


def _fmt_date(d) -> str:
    if d is None or (isinstance(d, float) and math.isnan(d)):
        return "—"
    if hasattr(d, "strftime"):
        return d.strftime("%d %b %Y")
    s = str(d)
    return s[:10] if s else "—"


def _iso_in_week(iso_str: str, ws: datetime.date, we: datetime.date) -> bool:
    try:
        d = datetime.date.fromisoformat(iso_str)
        return ws <= d <= we
    except:
        return False


# ─────────────────────────────────────────────────────────────
# FIX D HELPER — per-SKU session state keys
# ─────────────────────────────────────────────────────────────
def _sku_key(sku_label: str, suffix: str) -> str:
    safe = sku_label.replace(" ", "_").replace("/", "_").replace("—", "-")
    return f"{safe}__{suffix}"


def ensure_sku_state(sku_label: str):
    for suffix, default in [
        ("issued_wos",    {}),
        ("demand_events", []),
        ("po_released",   {"reel": {}, "pouch": {}}),
    ]:
        k = _sku_key(sku_label, suffix)
        if k not in st.session_state:
            st.session_state[k] = default


# ─────────────────────────────────────────────────────────────
# 3.  COMPUTATION ENGINE  (MOQ-aware, fixed SS)
# ─────────────────────────────────────────────────────────────
def compute(raw: dict, demand_events: list, issued_wos: dict,
            po_history: "pd.DataFrame | None" = None) -> dict:
    z          = raw["z_score"]
    lt         = raw["prod_lt"]
    lt_var     = raw["lt_variability"]   # default from excel col H (SKU Basic Details)
    w_pct      = raw["wastage_pct"] / 100
    target_fg  = raw["target_fg"]
    current_fg = raw["current_fg"]
    prod_moq   = raw["prod_moq"]

    # ── Override lt_var from actual WO Log deviations ────────
    # When ≥6 completed WOs exist for this SKU, compute σ of (actual − planned LT)
    # from the Deviation (days) column instead of using the excel static value.
    _wo_log = raw.get("wo_log", [])
    _completed_deviations = [
        wo["deviation"]
        for wo in _wo_log
        if not wo.get("in_transit") and wo.get("deviation") is not None
        and wo.get("cycle_time", 0) > 0   # only use rows with actual cycle time
    ]
    if len(_completed_deviations) >= 6:
        import statistics as _stats
        try:
            lt_var = round(_stats.stdev(_completed_deviations), 1)
        except Exception:
            pass   # keep the excel value if calculation fails
    _lt_var_source = "WO Log" if len(_completed_deviations) >= 6 else "Excel"

    # ── Fixed Safety Stock ──────────────────────────────────
    # Use the explicitly set SS from parameters; compute fallback if 0
    if raw["safety_stock_fixed"] > 0:
        safety_stock = raw["safety_stock_fixed"]
    else:
        avg_daily_dem_ss = target_fg / 30
        safety_stock = round(z * avg_daily_dem_ss * (lt_var))

    # ── Inventory model ─────────────────────────────────────
    # Max FG (replenish ceiling) = MTS + SS
    #   e.g. 34,100 + 8,000 = 42,100
    # ROP = MTS + SS - Production MOQ
    #   One MOQ run from ROP brings you exactly back to Max FG.
    #   e.g. 34,100 + 8,000 - 8,000 = 34,100
    max_fg = target_fg + safety_stock
    rop    = max(safety_stock, max_fg - prod_moq)

    months_info = get_months_ahead(6)

    date_demand_map: dict[datetime.date, int] = {}
    for ev in demand_events:
        d = ev["date"]
        date_demand_map[d] = date_demand_map.get(d, 0) + int(ev["qty"])

    total_event_qty = sum(ev["qty"] for ev in demand_events)
    avg_daily_dem   = target_fg / 30   # for SS/cap reference only

    # ── Process WOs already issued ──────────────────────────
    pending_arrivals: dict[datetime.date, int] = {}
    wo_release_log:   dict[datetime.date, dict] = {}
    suppress_ranges:  list[tuple] = []

    for date_str, info in issued_wos.items():
        if info.get("issued"):
            rel_date    = datetime.date.fromisoformat(date_str)
            actual_lt_d = info.get("actual_lt_override", lt)
            so_wo_days  = info.get("so_to_wo_days", raw.get("so_to_wo_days", 1))

            # Use pre-computed arrival date if saved (includes SO→WO days)
            if info.get("wo_arrival_date"):
                arr_date = datetime.date.fromisoformat(info["wo_arrival_date"])
            elif info.get("wo_release_date"):
                _wrel    = datetime.date.fromisoformat(info["wo_release_date"])
                arr_date = _wrel + datetime.timedelta(days=int(actual_lt_d))
            else:
                arr_date = rel_date + datetime.timedelta(days=int(so_wo_days) + int(actual_lt_d))

            produced    = info.get("actual_produced", info.get("wo_qty", 0))
            wo_release_log[rel_date] = {
                "arrival_date":    arr_date,
                "wo_qty":          info.get("wo_qty", 0),
                "gross_qty":       info.get("gross_qty", 0),
                "issued":          True,
                "actual_produced": produced,
                "actual_lt":       actual_lt_d,
                "so_to_wo_days":   so_wo_days,
            }
            # Suppress new WO alerts from release date until arrival
            suppress_ranges.append(
                (rel_date, arr_date - datetime.timedelta(days=1))
            )
            today_pre = datetime.date.today()
            if arr_date > today_pre:
                pending_arrivals[arr_date] = pending_arrivals.get(arr_date, 0) + produced

    def is_suppressed(d: datetime.date) -> bool:
        return any(s <= d <= e for s, e in suppress_ranges)

    today    = datetime.date.today()
    end_date = today + datetime.timedelta(days=183)  # 6 months forward

    # ── Determine simulation start: earliest of (first demand event, first WO release) ──
    all_event_dates = [ev["date"] for ev in demand_events if isinstance(ev["date"], datetime.date)]
    all_wo_dates    = []
    for ds, info in issued_wos.items():
        if info.get("issued"):
            try:
                all_wo_dates.append(datetime.date.fromisoformat(ds))
            except Exception:
                pass

    if all_event_dates or all_wo_dates:
        earliest_hist = min(all_event_dates + all_wo_dates)
        # Snap to first of that month so the chart starts cleanly
        sim_start = datetime.date(earliest_hist.year, earliest_hist.month, 1)
    else:
        sim_start = today

    all_dates = [sim_start + datetime.timedelta(days=i)
                 for i in range((end_date - sim_start).days + 1)]

    def is_suppressed(d: datetime.date) -> bool:
        return any(s <= d <= e for s, e in suppress_ranges)

    today    = datetime.date.today()
    end_date = today + datetime.timedelta(days=183)

    # ── Build all_arrivals map (all WOs past + future) ───────
    # This iterates ALL issued_wos — past and future arrivals.
    # Do NOT also add pending_arrivals here: pending_arrivals contains
    # the same future WOs already included above, so merging would double-count.
    all_arrivals: dict[datetime.date, int] = {}
    for ds, info in issued_wos.items():
        if info.get("issued"):
            try:
                # Use pre-computed arrival date if saved, else compute from release + LT
                if info.get("wo_arrival_date"):
                    arr = datetime.date.fromisoformat(info["wo_arrival_date"])
                else:
                    # wo_release_date may differ from ds by so_to_wo_days
                    rel_str = info.get("wo_release_date", ds)
                    rel  = datetime.date.fromisoformat(rel_str)
                    lt_d = info.get("actual_lt_override", lt)
                    arr  = rel + datetime.timedelta(days=int(lt_d))
                prod = info.get("actual_produced", info.get("wo_qty", 0))
                all_arrivals[arr] = all_arrivals.get(arr, 0) + prod
            except Exception:
                pass

    # ── sim_demand_map: always use Customer PO Date ───────────
    # This is the date the customer placed the order, which is the
    # demand signal we plan against. bill_date is only for the Delivery tab.
    sim_demand_map = date_demand_map   # already built from ev["date"] (PO date)

    # ── Determine simulation start ────────────────────────────
    all_event_dates = list(sim_demand_map.keys())
    all_wo_dates    = []
    for ds, info in issued_wos.items():
        if info.get("issued"):
            try:
                all_wo_dates.append(datetime.date.fromisoformat(ds))
            except Exception:
                pass

    if all_event_dates or all_wo_dates:
        earliest = min(all_event_dates + all_wo_dates) if (all_event_dates and all_wo_dates) \
                   else (min(all_event_dates) if all_event_dates else min(all_wo_dates))
        sim_start = datetime.date(earliest.year, earliest.month, 1)
    else:
        sim_start = today

    all_dates = [sim_start + datetime.timedelta(days=i)
                 for i in range((end_date - sim_start).days + 1)]

    # ── Starting FG ──────────────────────────────────────────
    # Historical segment always starts at 0 (before any WO has run).
    # WO arrivals build FG up; demand events pull it down.
    # At today, if the FG inventory file was uploaded (current_fg > 0),
    # we snap the running fg to current_fg so the forward projection
    # uses the actual physical stock — not the simulated value.
    fg_start = 0

    sim_rows = []
    fg              = fg_start
    backorder_accum = 0

    for sim_date in all_dates:
        # ── At today's boundary: snap to actual FG if file was uploaded ──
        if sim_date == today and current_fg > 0:
            fg = current_fg
            backorder_accum = 0   # reset backorder; actual stock is the ground truth

        is_past = sim_date < today

        arrival_qty = all_arrivals.get(sim_date, 0)

        # Fulfill backorder from arrival first
        bo_fulfilled_today = 0
        if arrival_qty > 0 and backorder_accum > 0:
            bo_fulfilled_today = min(arrival_qty, backorder_accum)
            arrival_qty        = arrival_qty - bo_fulfilled_today
            backorder_accum    -= bo_fulfilled_today

        opening = fg + arrival_qty

        demand_today = date_demand_map.get(sim_date, 0)

        closing = opening - demand_today

        new_backorder = max(0, -closing)
        backorder_accum += new_backorder

        # EOM projection
        month_end = None
        for lbl, ms, me in months_info:
            if ms <= sim_date <= me:
                month_end = me
                break
        if month_end is None:
            month_end = end_date

        future_demand = sum(
            date_demand_map.get(sim_date + datetime.timedelta(days=k), 0)
            for k in range((month_end - sim_date).days + 1)
        )
        projected_fg_eom = max(0, closing) - (future_demand - demand_today)

        # ── WO trigger (future dates only) ────────────────────
        # In the past segment we never generate new WOs — they already happened.
        in_transit_fg = sum(
            qty for arr_d, qty in all_arrivals.items()
            if arr_d > sim_date
        )
        projected_after_transit = max(0, closing) + in_transit_fg - backorder_accum

        urgent_override = (not is_past) and projected_after_transit <= rop and backorder_accum > 0

        rop_hit = (not is_past) and (
            (closing <= rop or backorder_accum >= prod_moq or urgent_override)
            and (not is_suppressed(sim_date) or urgent_override)
        )

        wo_gross_qty  = 0
        wo_net_qty    = 0
        wo_arrival_dt = None

        if rop_hit:
            # WO Qty = Max FG − projected_after_transit
            # projected_after_transit = closing + in_transit_fg − backorder_accum
            # This correctly accounts for what's already coming in and what's owed to customers
            gap          = max_fg - projected_after_transit
            net_needed   = max(gap, prod_moq)
            moq_multiple = math.ceil(net_needed )
            wo_net_qty   = moq_multiple
            wo_gross_qty = math.ceil(wo_net_qty / (1 - w_pct)) if w_pct < 1 else wo_net_qty
            wo_arrival_dt = sim_date + datetime.timedelta(days=lt)

        date_str  = _safe_date_str(sim_date)
        user_info = issued_wos.get(date_str, {})
        wo_issued = user_info.get("issued", False)

        actual_lt_d  = user_info.get("actual_lt_override", lt) if wo_issued else lt
        so_wo_d      = user_info.get("so_to_wo_days", raw.get("so_to_wo_days", 1)) if wo_issued else raw.get("so_to_wo_days", 1)

        if wo_issued:
            display_gross = user_info.get("gross_qty", wo_gross_qty)
            display_net   = user_info.get("actual_produced", wo_net_qty)
            # Use pre-computed arrival date if saved; otherwise compute release + prod_lt
            if user_info.get("wo_arrival_date"):
                display_arr = datetime.date.fromisoformat(user_info["wo_arrival_date"])
            elif user_info.get("wo_release_date"):
                _rel = datetime.date.fromisoformat(user_info["wo_release_date"])
                display_arr = _rel + datetime.timedelta(days=int(actual_lt_d))
            else:
                display_arr = sim_date + datetime.timedelta(days=int(so_wo_d) + int(actual_lt_d))
        else:
            display_gross = wo_gross_qty
            display_net   = wo_net_qty
            display_arr   = wo_arrival_dt

        actual_produced = user_info.get("actual_produced", wo_net_qty) if wo_issued else 0

        if wo_issued and sim_date not in wo_release_log:
            # suppress_ranges: don't trigger new WOs while this one is in production
            suppress_ranges.append((sim_date, display_arr - datetime.timedelta(days=1)))
            wo_release_log[sim_date] = {
                "arrival_date":    arr,
                "wo_qty":          wo_net_qty,
                "gross_qty":       wo_gross_qty,
                "issued":          True,
                "actual_produced": actual_produced,
                "actual_lt":       actual_lt_d,
            }

        sim_rows.append({
            "date":             sim_date,
            "month":            sim_date.strftime("%b %Y"),
            "is_historical":    is_past,
            "opening_fg":       round(opening),
            "arrival_qty":      round(arrival_qty + bo_fulfilled_today),
            "demand":           round(demand_today),
            "closing_fg":       round(max(0, closing)),
            "backorder":        round(backorder_accum),
            "rop":              rop,
            "rop_hit":          rop_hit,
            "projected_fg_eom": round(projected_fg_eom),
            "target_fg":        target_fg,
            "safety_stock":     safety_stock,
            "max_fg":           max_fg,
            "wo_gross_qty":     display_gross,
            "wo_net_qty":       display_net,
            "wo_arrival_date":  display_arr,
            "wo_issued":        wo_issued,
            "actual_produced":  actual_produced,
            "is_month_end":     sim_date == month_end,
            "below_ss":         closing < safety_stock,
        })

        fg = max(0, closing)

    sim_df = pd.DataFrame(sim_rows)

    # ── Historical WO log ───────────────────────────────────
    wo_df_rows = []
    for w_ in raw["wo_log"]:
        # Skip in-transit WOs — they have no actuals
        if w_.get("in_transit"):
            continue
        # Parse release date — may be a datetime string, ISO string, or datetime object
        def _parse_date(val):
            if val is None or str(val).strip() in ("", "nan", "None"):
                return None
            if isinstance(val, (datetime.date, datetime.datetime)):
                return val.date() if isinstance(val, datetime.datetime) else val
            s = str(val).strip()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.datetime.strptime(s[:10], fmt[:8] if "H" in fmt else fmt).date()
                except:
                    pass
            try:
                return pd.to_datetime(s, dayfirst=True, errors="coerce").date()
            except:
                return None

        rel = _parse_date(w_.get("release"))
        cmp = _parse_date(w_.get("complete"))

        # Use deviation from WO Log if available, else compute from dates
        if w_.get("cycle_time") and w_["cycle_time"] > 0:
            actual_lt = w_["cycle_time"]
            deviation = w_.get("deviation", 0)
        elif rel and cmp:
            actual_lt = (cmp - rel).days
            deviation = actual_lt - w_.get("planned_lt", lt)
        else:
            actual_lt = None
            deviation = None

        # Use yield/waste from WO Log if populated, else compute
        tgt  = w_.get("target_qty", 0) or 0
        prod = w_.get("produced", 0)    or 0
        if w_.get("wastage_pct"):
            waste_pct = round(float(w_["wastage_pct"]), 1)
            yield_pct = round(float(w_.get("yield_pct", 0)), 1)
        else:
            waste_pct = round((tgt - prod) / tgt * 100, 1) if tgt else 0
            yield_pct = round(prod / tgt * 100, 1) if tgt else 0

        on_time = ("✅ Yes" if w_.get("on_time") else
                   "✅ Yes" if (actual_lt is not None and actual_lt <= lt) else "❌ No")

        wo_df_rows.append({
            "WO #":            w_.get("wo_num", ""),
            "SO #":            w_.get("so_num", ""),
            "Release":         rel,
            "Est. Complete":   _parse_date(w_.get("est_complete")),
            "Complete":        cmp,
            "Target":          tgt,
            "Produced":        prod,
            "Planned LT":      w_.get("planned_lt", lt),
            "Actual LT (days)":actual_lt,
            "Deviation (days)":deviation,
            "On Time?":        on_time,
            "Waste %":         waste_pct,
            "Yield %":         yield_pct,
            "Notes":           w_.get("notes", ""),
        })
    wo_df = pd.DataFrame(wo_df_rows)

    # ── RM planning ─────────────────────────────────────────
    reel_weight_monthly: dict[str, int]           = {}
    pouch_monthly:       dict[str, int]           = {}
    reel_pos:            dict[str, datetime.date] = {}
    pouch_pos:           dict[str, datetime.date] = {}

    reel_sheets = raw["reel_sheets"]
    # Recalculate if 0 — sum KG per component (each has different dims), then back-derive sheets
    if reel_sheets == 0 and raw.get("target_fg") and raw.get("bom_components"):
        _primary = raw.get("reel_name", "")
        _tgt     = raw["target_fg"]
        _total_kg_fallback = sum(
            math.ceil(
                math.ceil(_tgt / comp["ups"])
                * comp["width"] * comp["length"]
                * comp["gsm"] / ((1 - comp["wastage"] / 100) * 1_000_000_000)
            )
            for comp in raw["bom_components"]
            if comp.get("reel_name") == _primary and comp.get("ups", 0) > 0
        )
        _kps = raw["reel_width"] * raw["reel_length"] * raw["reel_gsm"] * (1 + raw["reel_waste"] / 100) / 1_000_000_000
        reel_sheets = math.ceil(_total_kg_fallback / _kps) if _kps > 0 else math.ceil(_tgt / max(raw.get("reel_ups", 1), 1))
    pouch_req = raw["pouch_req"]
    reel_sheets_to_kg = math.ceil(
        reel_sheets * raw["reel_width"] * raw["reel_length"]
        * raw["reel_gsm"] / ((1 - raw["reel_waste"] / 100) * 1_000_000_000)
    )

    for lbl, ms, me in months_info:
        month_demand   = sum(ev["qty"] for ev in demand_events if ms <= ev["date"] <= me)
        sheets_monthly = math.ceil(month_demand / raw["reel_ups"]) if raw["reel_ups"] else 0
        rw_kg = math.ceil(
            sheets_monthly * raw["reel_width"] * raw["reel_length"]
            * raw["reel_gsm"] / ((1 - raw["reel_waste"] / 100) * 1_000_000_000)
        ) if sheets_monthly else 0
        reel_weight_monthly[lbl] = rw_kg

        pouches_per_carton = raw["pouch_per_1m"] / 1_000_000
        pouch_monthly[lbl] = round(month_demand * pouches_per_carton)

        wo_rel_for_month = me - datetime.timedelta(days=lt)
        reel_pos[lbl]    = wo_rel_for_month - datetime.timedelta(days=raw["reel_total_lt"])
        pouch_pos[lbl]   = wo_rel_for_month - datetime.timedelta(days=int(raw["pouch_lt"]))

    # ── Per-reel ROP: each unique reel in the BOM gets its own SS and ROP ──
    # For each reel name, sum KG across all components that use it (each has
    # different dims/ups), then compute SS and ROP from that total monthly KG.
    reel_rop_map: dict[str, int] = {}   # reel_name → ROP in KG
    reel_ss_map:  dict[str, int] = {}   # reel_name → SS in KG

    bom = raw.get("bom_components", [])
    from collections import defaultdict
    reel_comps: dict[str, list] = defaultdict(list)
    for comp in bom:
        reel_comps[comp["reel_name"]].append(comp)

    # ── Demand variance from PO history (used when ≥10 months of data) ──────
    # po_df_global is parsed at the top-level from sidebar upload
    # sigma_demand_kg[reel_nm] = standard deviation of monthly KG demand
    def _get_demand_sigma_kg(reel_nm: str, comps: list, daily_kg: float) -> float:
        """
        Returns σ_demand (in KG/day) from PO history if ≥10 months of data exist.
        σ_demand = std dev of |actual monthly demand − MTS target|.
        This measures how consistently demand deviates from the plan,
        regardless of direction. SS protects against this uncertainty.
        If demand always = MTS exactly, σ_demand = 0 → SS = Z×D×σ_LT only.
        """
        try:
            if po_history is None or po_history.empty:
                return 0.0
            mat_id_s = raw.get("mat_id", "")
            mts      = raw.get("target_fg", 0)
            sku_df   = po_history[po_history["mat_id"] == mat_id_s]
            if sku_df.empty:
                return 0.0
            first_date = sku_df["date"].min()
            months_available = (datetime.date.today() - first_date.date()).days / 30
            if months_available < 10:
                return 0.0
            sku_df = sku_df.copy()
            sku_df["month"] = sku_df["date"].dt.to_period("M")
            monthly_qty = sku_df.groupby("month")["qty"].sum()
            if len(monthly_qty) < 2:
                return 0.0
            # Absolute deviations from MTS — std dev of these is σ_demand
            abs_deviations = np.abs(monthly_qty.values.astype(float) - mts)
            sigma_cartons_monthly = float(np.std(abs_deviations, ddof=1))
            # Convert monthly carton σ to monthly KG σ for this reel
            # Use ratio: KG_per_carton = daily_kg / (mts/30)
            kg_per_carton = daily_kg / (mts / 30) if mts > 0 else 0
            sigma_monthly_kg = sigma_cartons_monthly * kg_per_carton
            return sigma_monthly_kg / 30   # convert to daily σ
        except:
            return 0.0

    for reel_nm, comps in reel_comps.items():
        monthly_kg = 0
        for comp in comps:
            sheets_per_month = math.ceil(target_fg / comp["ups"]) if comp["ups"] else 0
            _mult = comp.get("multiplier", 1) or 1
            monthly_kg += math.ceil(
                sheets_per_month * comp["width"] * comp["length"]
                * comp["gsm"] / ((1 - comp["wastage"] / 100) * 1_000_000_000) * _mult
            )
        comp0    = comps[0]
        rm_lt_r  = comp0["total_lt"]
        # Use per-reel lt_var from reel_lt_var_map; fall back to comp field then raw scalar
        lt_var_r = raw.get("reel_lt_var_map", {}).get(reel_nm,
                   comp0.get("reel_lt_var", raw.get("reel_lt_var", 5)))
        daily_kg = monthly_kg / 30

        # Full SS formula: Z × √(LT × σ_demand² + D² × σ_LT²)
        # σ_LT = lt_var_r in days (std dev of LT)
        # σ_demand = daily demand std dev in KG (0 until ≥10 months of data)
        sigma_d  = _get_demand_sigma_kg(reel_nm, comps, daily_kg)
        sigma_lt = lt_var_r
        ss_r     = math.ceil(z * math.sqrt(
            rm_lt_r * sigma_d ** 2 + daily_kg ** 2 * sigma_lt ** 2
        ))
        rop_r    = math.ceil(daily_kg * rm_lt_r + ss_r)
        reel_ss_map[reel_nm]  = ss_r
        reel_rop_map[reel_nm] = rop_r

    # Total reel SS = sum across all distinct reels (additive, different lead times)
    total_reel_ss = sum(reel_ss_map.values())

    # Legacy single values from first/primary reel for backward compat
    _primary_reel = raw.get("reel_name", "")
    reel_ss     = reel_ss_map.get(_primary_reel, 0)
    reel_rop_kg = reel_rop_map.get(_primary_reel, 0)
    avg_pouch_monthly = sum(pouch_monthly.values()) / 6 if pouch_monthly else 0
    pouch_ss          = math.ceil(pouch_req / 30 * (raw["reel_lt_var"]))
    pouch_rop         = math.ceil(pouch_req / 30 * raw["pouch_lt"] + pouch_ss)

    days_of_supply = round(current_fg / avg_daily_dem, 1) if avg_daily_dem else 0

    return {
        "raw":                raw,
        "safety_stock":       safety_stock,
        "max_fg":             max_fg,
        "rop":                rop,
        "prod_moq":           prod_moq,
        "days_of_supply":     days_of_supply,
        "avg_daily_dem":      round(avg_daily_dem, 1),
        "lt_var_used":        lt_var,
        "lt_var_source":      _lt_var_source,
        "lt_var_n_wos":       len(_completed_deviations),
        "cap_data":           [],
        "bottleneck":         "—",
        "sim_df":             sim_df,
        "wo_df":              wo_df,
        "wo_release_log":     wo_release_log,
        "months_info":        months_info,
        "reel_weight_monthly":reel_weight_monthly,
        "reel_ss":            reel_ss,
        "reel_rop_kg":        round(reel_rop_kg),
        "reel_rop_map":       reel_rop_map,   # {reel_name → ROP kg} for all reels
        "reel_ss_map":        reel_ss_map,    # {reel_name → SS kg}
        "total_reel_ss":      total_reel_ss,  # sum SS across all reels
        "pouch_monthly":      pouch_monthly,
        "pouch_ss":           pouch_ss,
        "pouch_rop":          pouch_rop,
        "reel_pos":           reel_pos,
        "pouch_pos":          pouch_pos,
    }


def compute_capacity(raw: dict, demand_events: list) -> dict:
    """Capacity computation using actual demand events per month (not avg daily)."""
    months_info = get_months_ahead(6)
    processes   = raw["processes"]

    # Monthly demand from events
    month_demand_map = {}
    for lbl, ms, me in months_info:
        month_demand_map[lbl] = sum(
            ev["qty"] for ev in demand_events if ms <= ev["date"] <= me
        )

    cap_rows = []
    for lbl, ms, me in months_info:
        days_in      = (me - ms).days + 1
        month_dem    = month_demand_map.get(lbl, 0)
        daily_dem    = month_dem if days_in else 0
        sheets_daily = math.ceil(daily_dem / raw["number_of_ups"]) if raw["number_of_ups"] else 0

        proc_utils = {}
        for proc in processes:
            demand_unit = sheets_daily if proc["cap_unit"] == "Sheets" else daily_dem
            # Compare against per-shift capacity (capacity = total daily = shift_cap × shifts)
            shift_cap = proc["shift_cap"]
            daily_cap = proc["capacity"]
            util_daily = (demand_unit / daily_cap * 100) if daily_cap > 0 else 0
            util_shift = (demand_unit / shift_cap * 100) if shift_cap > 0 else 0
            proc_utils[proc["name"]] = {
                "daily_dem": round(demand_unit, 1),
                "daily_cap": daily_cap,
                "shift_cap": shift_cap,
                "shifts":    proc["shifts"],
                "util_pct":  round(util_daily, 2),
            }
        cap_rows.append({
            "month":       lbl,
            "month_dem":   month_dem,
            "daily_dem":   round(daily_dem, 1),
            "sheets_daily":sheets_daily,
            "proc_utils":  proc_utils,
        })

    # Summary per process across 6M
    proc_summary = []
    for proc in processes:
        max_util  = max((r["proc_utils"][proc["name"]]["util_pct"] for r in cap_rows), default=0)
        avg_util  = (sum(r["proc_utils"][proc["name"]]["util_pct"] for r in cap_rows)
                     / len(cap_rows)) if cap_rows else 0
        proc_summary.append({
            "name":       proc["name"],
            "daily_cap":  proc["capacity"],
            "shift_cap":  proc["shift_cap"],
            "shifts":     proc["shifts"],
            "cap_unit":   proc["cap_unit"],
            "days":       proc["days"],
            "max_util":   round(max_util, 2),
            "avg_util":   round(avg_util, 2),
        })

    bottleneck = max(proc_summary, key=lambda x: x["max_util"])["name"] if proc_summary else "—"

    return {
        "cap_rows":    cap_rows,
        "proc_summary":proc_summary,
        "bottleneck":  bottleneck,
    }


# ─────────────────────────────────────────────────────────────
# 4.  WEEKLY MRP TABLE BUILDER  (with Order-By date)
# ─────────────────────────────────────────────────────────────
def compute_global_rm_ss(all_raw: dict, po_history: "pd.DataFrame | None" = None) -> dict:
    """
    Compute Safety Stock and ROP for every unique RM (by rm_material_id) across ALL SKUs.

    When the same rm_material_id is shared by multiple SKUs, their daily KG demands
    and demand variabilities are SUMMED before computing SS/ROP — because the warehouse
    holds one stock of that RM serving all lines simultaneously.

    Returns:
        {rm_material_id → {
            "ss_kg":        int,   # Safety Stock in KG
            "rop_kg":       int,   # Reorder Point in KG
            "daily_kg":     float, # Combined daily consumption across all SKUs
            "lt_days":      int,   # Lead time (days)
            "lt_var":       int,   # LT variability σ (days)
            "moq_kg":       int,
            "reel_name":    str,   # Primary reel name (for registry lookup)
            "sku_list":     list,  # SKUs sharing this RM
            "rm_desc":      str,
        }}
    """
    import numpy as np
    today = datetime.date.today()

    # ── Step 1: Group all BOM components by rm_material_id ───
    # rm_id → list of (sku_label, comp, raw_dict)
    rm_groups: dict[str, list] = {}
    for sku_label, raw in all_raw.items():
        for comp in raw.get("bom_components", []):
            rm_id = (comp.get("rm_material_id") or "").strip()
            if not rm_id:
                # Fall back to reel_name as key if no RM ID
                rm_id = comp.get("reel_name", "").strip()
            if not rm_id:
                continue
            if rm_id not in rm_groups:
                rm_groups[rm_id] = []
            rm_groups[rm_id].append((sku_label, comp, raw))

    result: dict[str, dict] = {}

    for rm_id, entries in rm_groups.items():
        # ── Step 2: Aggregate daily KG demand across all sharing SKUs ──
        # For each SKU using this RM, compute its daily KG from target_fg + BOM dims.
        # Sum across SKUs — the warehouse needs to cover all lines.
        total_daily_kg = 0.0
        lt_days  = entries[0][1].get("total_lt", 22)     # LT same regardless of SKU
        lt_var   = entries[0][1].get("reel_lt_var", 5)
        moq_kg   = entries[0][1].get("moq_kg", 3000)
        reel_name = entries[0][1].get("reel_name", "")
        rm_desc  = entries[0][1].get("rm_material_desc", "")
        sku_list = []

        for sku_label, comp, raw in entries:
            if sku_label not in sku_list:
                sku_list.append(sku_label)
            target_fg = raw.get("target_fg", 0)
            ups       = comp.get("ups", 1) or 1
            width     = comp.get("width", 0)
            length    = comp.get("length", 0)
            gsm       = comp.get("gsm", 0)
            wastage   = comp.get("wastage", 15) / 100
            if target_fg > 0 and ups > 0 and width > 0 and length > 0 and gsm > 0:
                sheets_per_month = math.ceil(target_fg / ups)
                _mult_g = comp.get("multiplier", 1) or 1
                kg_per_month = math.ceil(
                    sheets_per_month * width * length * gsm
                    / ((1 - wastage) * 1_000_000_000) * _mult_g
                )
                total_daily_kg += kg_per_month / 30

            # Use the tightest (shortest) LT available — most conservative
            lt_days = min(lt_days, comp.get("total_lt", lt_days))
            lt_var  = max(lt_var,  comp.get("reel_lt_var", lt_var))   # worst-case variability

        # ── Step 3: Demand variability σ from PO history (combined across SKUs) ──
        sigma_d = 0.0
        if po_history is not None and not po_history.empty and total_daily_kg > 0:
            try:
                sku_dfs = []
                for sku_label, comp, raw in entries:
                    mat_id = raw.get("mat_id", "")
                    df_sku = po_history[po_history["mat_id"] == mat_id]
                    if not df_sku.empty:
                        sku_dfs.append(df_sku)
                if sku_dfs:
                    combined = pd.concat(sku_dfs)
                    months_avail = (today - combined["date"].min().date()).days / 30
                    if months_avail >= 10:
                        combined["month"] = combined["date"].dt.to_period("M")
                        monthly_qty = combined.groupby("month")["qty"].sum()
                        if len(monthly_qty) >= 2:
                            total_mts = sum(raw.get("target_fg", 0)
                                           for _, _, raw in entries
                                           if raw.get("mat_id") in [e[2].get("mat_id") for e in entries])
                            abs_dev = np.abs(monthly_qty.values.astype(float) - total_mts)
                            sigma_cartons_monthly = float(np.std(abs_dev, ddof=1))
                            kg_per_carton = total_daily_kg / (total_mts / 30) if total_mts > 0 else 0
                            sigma_d = sigma_cartons_monthly * kg_per_carton / 30
            except Exception:
                sigma_d = 0.0

        # ── Step 4: Compute SS and ROP ────────────────────────
        # Formula: SS = Z × √(LT × σ_demand² + D² × σ_LT²)
        # Use Z = 1.65 (95% service level) as default across all RMs
        # Individual SKU z_scores aren't meaningful at the RM pool level
        z = max((entries[i][2].get("z_score", 1.65) for i in range(len(entries))), default=1.65)
        sigma_lt = lt_var

        ss_kg  = math.ceil(z * math.sqrt(
            lt_days * sigma_d ** 2 + total_daily_kg ** 2 * sigma_lt ** 2
        ))
        rop_kg = math.ceil(total_daily_kg * lt_days + ss_kg)

        result[rm_id] = {
            "ss_kg":     ss_kg,
            "rop_kg":    rop_kg,
            "daily_kg":  round(total_daily_kg, 2),
            "lt_days":   lt_days,
            "lt_var":    lt_var,
            "moq_kg":    moq_kg,
            "reel_name": reel_name,
            "rm_desc":   rm_desc,
            "sku_list":  sku_list,
        }

    return result


def compute_bom_requirements(all_raw: dict, all_issued_wos: dict) -> dict:
    """
    Aggregate weekly KG/unit requirements per reel/pouch across ALL loaded SKUs.
    Requirements are WO-driven: RM is consumed on the WO release date.

    all_raw:        {sku_label → raw_dict}
    all_issued_wos: {sku_label → issued_wos_dict}
    """
    today      = datetime.date.today()
    end_date   = today + datetime.timedelta(days=183)
    week_start = today - datetime.timedelta(days=today.weekday())
    weeks = []
    ws = week_start
    while ws <= end_date:
        we = ws + datetime.timedelta(days=6)
        weeks.append((ws, min(we, end_date)))
        ws += datetime.timedelta(days=7)

    reel_req  = {}
    pouch_req = {}
    reel_meta  = {}
    pouch_meta = {}

    for sku_label, raw in all_raw.items():
        issued_wos = all_issued_wos.get(sku_label, {})

        # Auto-register reels and pouches
        for comp in raw.get("bom_components", []):
            _ensure_reel_in_registry(comp)
        _ensure_pouch_in_registry(raw)

        # Weekly WO quantities for this SKU
        for ws, we in weeks:
            ws_iso   = ws.isoformat()
            week_wo_qty = sum(
                info.get("gross_qty", info.get("gross_qty", info.get("actual_produced", info.get("wo_qty", 0))))
                for date_str, info in issued_wos.items()
                if info.get("issued") and ws <= datetime.date.fromisoformat(date_str) <= we
            )
            if week_wo_qty == 0:
                continue

            # Reel requirements per BOM component
            for comp in raw.get("bom_components", []):
                # Key by rm_material_id so each unique RM gets its own bucket
                nm = (comp.get("rm_material_id") or "").strip() or comp["reel_name"]
                sheets = math.ceil(week_wo_qty / comp["ups"]) if comp["ups"] else 0
                kg     = math.ceil(
                    sheets * comp["width"] * comp["length"]
                    * comp["gsm"] / 1_000_000_000 * (comp.get("multiplier", 1) or 1)
                ) if sheets else 0

                if nm not in reel_req:
                    reel_req[nm]  = {}
                    reel_meta[nm] = {
                        "total_lt":   comp["total_lt"],
                        "moq":        comp["moq_kg"],
                        "components": [],
                    }
                reel_req[nm][ws_iso] = reel_req[nm].get(ws_iso, 0) + kg
                entry = {"sku": sku_label, "component": comp["component"],
                         "ups": comp["ups"], "width": comp["width"],
                         "length": comp["length"], "gsm": comp["gsm"]}
                if entry not in reel_meta[nm]["components"]:
                    reel_meta[nm]["components"].append(entry)

            # Pouch requirement
            nm_p  = raw["pouch_name"]
            upc   = raw.get("pouch_per_1m", 20000) / 1_000_000
            units = round(week_wo_qty * upc)
            if nm_p not in pouch_req:
                pouch_req[nm_p]  = {}
                pouch_meta[nm_p] = {
                    "lt":   int(raw.get("pouch_lt", 3)),
                    "moq":  int(raw.get("pouch_moq", 20000)),
                    "skus": [],
                }
            pouch_req[nm_p][ws_iso] = pouch_req[nm_p].get(ws_iso, 0) + units
            if sku_label not in pouch_meta[nm_p]["skus"]:
                pouch_meta[nm_p]["skus"].append(sku_label)

    return {
        "reel":  {nm: {"weekly": reel_req[nm],  **reel_meta[nm]}  for nm in reel_req},
        "pouch": {nm: {"weekly": pouch_req[nm], **pouch_meta[nm]} for nm in pouch_req},
    }


def build_consolidated_mrp(rm_name: str, rm_type: str,
                            weekly_req: dict, weeks: list,
                            po_key: str = None,
                            actual_arrivals: dict = None) -> pd.DataFrame:
    """
    po_key: reel_pos_released key (rm_material_id when available, else reel_name).
    actual_arrivals: {rel_iso → actual_date} overrides rel_d + rm_lt.
    """
    today = datetime.date.today()
    _po_lookup = po_key or rm_name
    if rm_type == "reel":
        # Registry keyed by rm_material_id (= po_key when available, else rm_name)
        reg      = (st.session_state["reel_registry"].get(_po_lookup) or
                    st.session_state["reel_registry"].get(rm_name) or {})
        pos      = st.session_state["reel_pos_released"].get(_po_lookup, {})
        stock    = reg.get("stock_kg", 0)
        moq      = reg.get("moq", 3000)
        rm_lt    = reg.get("total_lt", 22)
        unit     = "KG"
        # Use global_rm_ss ROP when available, else fall back to demand-driven calc
        _global_rop = (st.session_state.get("_global_rm_ss_cache") or {}).get(_po_lookup, {}).get("rop_kg", 0)
        if _global_rop:
            rop = _global_rop
        else:
            total_req  = sum(weekly_req.values())
            weeks_cnt  = max(len([v for v in weekly_req.values() if v > 0]), 1)
            avg_weekly = total_req / weeks_cnt if weeks_cnt else 0
            rop        = math.ceil(avg_weekly / 7 * rm_lt)
    else:
        reg      = st.session_state["pouch_registry"].get(rm_name, {})
        pos      = st.session_state["pouch_pos_released"].get(rm_name, {})
        stock    = reg.get("stock_units", 0)
        moq      = reg.get("moq", 20000)
        rm_lt    = reg.get("lt", 3)
        unit     = "units"
        total_req = sum(weekly_req.values())
        weeks_cnt = max(len([v for v in weekly_req.values() if v > 0]), 1)
        avg_weekly = total_req / weeks_cnt if weeks_cnt else 0
        rop       = math.ceil(avg_weekly / 7 * rm_lt)

    # PO arrival map — respect actual_arrivals overrides
    po_arrival_map = {}
    for rel_iso, qty in pos.items():
        try:
            if actual_arrivals and rel_iso in actual_arrivals:
                arr = actual_arrivals[rel_iso]
            else:
                arr = datetime.date.fromisoformat(rel_iso) + datetime.timedelta(days=rm_lt)
            po_arrival_map[arr] = po_arrival_map.get(arr, 0) + qty
        except:
            pass

    rows  = []
    cur_stock = stock
    for wk_num, (ws, we) in enumerate(weeks, 1):
        ws_iso    = ws.isoformat()
        gross_req = weekly_req.get(ws_iso, 0)
        po_arrive = sum(q for d, q in po_arrival_map.items() if ws <= d <= we)
        opening   = cur_stock + po_arrive
        net_req   = max(0, gross_req - opening)
        closing   = opening - gross_req

        # In-transit: POs released on or before this week end that haven't arrived yet
        in_transit = sum(
            qty for rel_iso, qty in pos.items()
            if (lambda rd, ad: rd <= we and ad > we)(
                datetime.date.fromisoformat(rel_iso),
                datetime.date.fromisoformat(rel_iso) + datetime.timedelta(days=rm_lt)
            )
        )

        # Suppress reorder if any PO (past or future) released by ws hasn't arrived yet
        in_transit_not_yet_here_c = sum(
            qty for rel_iso, qty in pos.items()
            if (lambda rd, ad: rd <= ws and ad > ws)(
                datetime.date.fromisoformat(rel_iso),
                datetime.date.fromisoformat(rel_iso) + datetime.timedelta(days=rm_lt)
            )
        )
        po_will_cover_c = (opening + in_transit_not_yet_here_c) >= rop
        order_by_date = ws - datetime.timedelta(days=rm_lt)
        if opening <= rop and in_transit_not_yet_here_c > 0 and po_will_cover_c:
            reorder = "🚚 PO in production"
            ob_str  = "—"
        elif opening <= rop:
            reorder = "🔴 YES"
            if order_by_date >= today:
                ob_str = order_by_date.strftime("%d %b %Y")
            else:
                days_od = (today - order_by_date).days
                ob_str  = f"{order_by_date.strftime('%d %b %Y')} ⚠️ ({days_od}d overdue)"
        else:
            reorder = "—"
            ob_str  = "—"

        po_this_week = sum(q for ri, q in pos.items() if _iso_in_week(ri, ws, we))

        rows.append({
            "Week":                    wk_num,
            "Date Range":              f"{ws.strftime('%d %b')} – {we.strftime('%d %b %Y')}",
            f"Gross Req ({unit})":     gross_req,
            f"Opening Stock ({unit})": round(opening),
            f"Net Req ({unit})":       net_req,
            f"Closing Stock ({unit})": round(max(0, closing)),
            f"🏭 In Production ({unit})": in_transit,
            "Reorder?":                reorder,
            "📅 Order By":             ob_str,
            "PO Released?":            f"{po_this_week:,} {unit}" if po_this_week else "—",
            "Receiving Date":          (ws + datetime.timedelta(days=rm_lt)).strftime("%d %b %Y"),
        })
        cur_stock = closing

    return pd.DataFrame(rows)


def build_weekly_mrp(raw: dict, r: dict, issued_wos: dict,
                     rm_type: str = "reel",
                     po_released: dict = None,
                     stock_offset: int = 0,
                     use_raw_stock: bool = False,
                     other_weekly_kg: dict = None,
                     actual_arrivals: dict = None) -> pd.DataFrame:
    """
    actual_arrivals: {rel_iso → actual_date} — overrides rel_d + rm_lt for arrival date per PO.
    """
    if po_released is None:
        po_released = {}

    today    = datetime.date.today()
    end_date = today + datetime.timedelta(days=183)

    if rm_type == "reel":
        reel_nm = raw.get("reel_name", "")
        moq   = raw["reel_moq"]
        rm_lt = raw["reel_total_lt"]
        # Use rm_material_id from comp_raw directly — don't scan bom_components
        # which would find the first matching reel_name (may be wrong component)
        _rm_id_for_rop = (raw.get("rm_material_id") or "").strip() or reel_nm
        # Stock from registry using rm_material_id key
        _reg_entry = (st.session_state.get("reel_registry", {}).get(_rm_id_for_rop) or
                      st.session_state.get("reel_registry", {}).get(reel_nm) or {})
        # If this rm_id has no stock, find another registry entry with same reel_name
        if _reg_entry.get("stock_kg", 0) == 0:
            _siblings = [info for k, info in st.session_state.get("reel_registry", {}).items()
                         if info.get("reel_name") == reel_nm and info.get("stock_kg", 0) > 0]
            if _siblings:
                _reg_entry = max(_siblings, key=lambda x: x.get("stock_kg", 0))
        if use_raw_stock:
            opening_stock = max(0, raw["reel_stock"] - stock_offset)
        else:
            opening_stock = max(0, _reg_entry.get("stock_kg", raw["reel_stock"]) - stock_offset)
        moq   = raw["reel_moq"]
        rm_lt = _reg_entry.get("total_lt", raw["reel_total_lt"])
        rop   = global_rm_ss.get(_rm_id_for_rop, {}).get("rop_kg") or \
                r.get("reel_rop_map", {}).get(reel_nm, r["reel_rop_kg"])
        unit  = "KG"
    else:
        pouch_nm = raw.get("pouch_name", "")
        # Always read pouch stock from registry (never from raw which may be 0/stale)
        opening_stock = max(0,
            st.session_state.get("pouch_registry", {})
            .get(pouch_nm, {}).get("stock_units", raw.get("pouch_stock", 0)) - stock_offset
        )
        moq   = raw["pouch_moq"]
        rm_lt = int(raw["pouch_lt"])
        rop   = r["pouch_rop"]
        unit  = "units"

    week_start = today - datetime.timedelta(days=today.weekday())
    weeks: list[tuple[datetime.date, datetime.date]] = []
    ws = week_start
    while ws <= end_date:
        we = ws + datetime.timedelta(days=6)
        weeks.append((ws, min(we, end_date)))
        ws += datetime.timedelta(days=7)

    po_arrival_map: dict[datetime.date, int] = {}
    for rel_iso, qty in po_released.items():
        try:
            rel_date = datetime.date.fromisoformat(rel_iso)
        except:
            continue
        # Use actual receiving date if set, otherwise fall back to rel + LT
        if actual_arrivals and rel_iso in actual_arrivals:
            arr_date = actual_arrivals[rel_iso]
        else:
            arr_date = rel_date + datetime.timedelta(days=rm_lt)
        po_arrival_map[arr_date] = po_arrival_map.get(arr_date, 0) + qty

    rows  = []
    stock = opening_stock
    rop_already_triggered = False   # track if we've already shown the first reorder alert

    for wk_num, (ws, we) in enumerate(weeks, 1):
        # RM consumed = WOs released this week
        week_wo_fg = sum(
            info.get("gross_qty", info.get("gross_qty", info.get("actual_produced", info.get("wo_qty", 0))))
            for date_str, info in issued_wos.items()
            if info.get("issued") and ws <= datetime.date.fromisoformat(date_str) <= we
        )

        if rm_type == "reel":
            # Use BOM component dimensions from raw (= comp_raw), not primary reel fields
            # comp_raw has reel_name, reel_ups, reel_width, etc. set per-component
            _comp_ups    = raw.get("reel_ups", 1) or 1
            _comp_width  = raw.get("reel_width",  raw.get("width", 1))
            _comp_length = raw.get("reel_length", raw.get("length", 1))
            _comp_gsm    = raw.get("reel_gsm",    raw.get("gsm", 120))
            _comp_waste  = raw.get("reel_waste",   raw.get("wastage", 15)) / 100
            _comp_mult   = raw.get("multiplier", 1) or 1
            sheets    = math.ceil(week_wo_fg / _comp_ups) if _comp_ups else 0
            gross_req = math.ceil(
                sheets * _comp_width * _comp_length
                * _comp_gsm / ((1 - _comp_waste) * 1_000_000_000) * _comp_mult
            ) if sheets else 0
        else:
            pouches_per = raw["pouch_per_1m"] / 1_000_000
            gross_req   = round(week_wo_fg * pouches_per)

        ws_iso    = ws.isoformat()
        also_used = int(other_weekly_kg.get(ws_iso, 0)) if other_weekly_kg else 0
        total_consumption = gross_req + also_used

        po_arrive = sum(qty for arr_dt, qty in po_arrival_map.items() if ws <= arr_dt <= we)
        opening   = stock + po_arrive
        net_req   = max(0, total_consumption - opening)
        closing   = opening - total_consumption

        def _arr(rel_iso):
            """Get actual or default arrival date for a PO."""
            try:
                rd = datetime.date.fromisoformat(rel_iso)
            except:
                return datetime.date.today()
            if actual_arrivals and rel_iso in actual_arrivals:
                return actual_arrivals[rel_iso]
            return rd + datetime.timedelta(days=rm_lt)

        in_transit_qty = sum(
            qty for rel_iso, qty in po_released.items()
            if (lambda ad: _arr(rel_iso) > we and datetime.date.fromisoformat(rel_iso) <= we)(None)
        )
        in_transit_not_yet_here = sum(
            qty for rel_iso, qty in po_released.items()
            if (lambda ad: datetime.date.fromisoformat(rel_iso) <= ws and ad > ws)(_arr(rel_iso))
        )
        po_will_cover = (opening + in_transit_not_yet_here) >= rop

        # ROP check: trigger when closing drops below ROP (consumption this week caused the breach)
        # OR when opening is already below ROP (carried over from previous week)
        closing_below_rop = closing < rop
        opening_below_rop = opening <= rop

        if opening_below_rop and in_transit_not_yet_here > 0 and po_will_cover:
            reorder_status = "🚚 PO in production"
            order_by_str   = "—"
            rop_already_triggered = False
        elif closing_below_rop and not rop_already_triggered:
            # First week where stock crosses below ROP — show the trigger date
            reorder_status = "🔴 YES"
            if total_consumption > 0:
                # WO consumed RM this week causing the breach — order on WO date
                wo_dates_this_week = sorted([
                    datetime.date.fromisoformat(ds)
                    for ds, info in issued_wos.items()
                    if info.get("issued") and ws <= datetime.date.fromisoformat(ds) <= we
                ])
                order_by_str = wo_dates_this_week[0].strftime("%d %b %Y") if wo_dates_this_week else today.strftime("%d %b %Y")
            else:
                # Already below ROP with no consumption — order today
                order_by_str = today.strftime("%d %b %Y")
            rop_already_triggered = True
        elif opening_below_rop and rop_already_triggered:
            # Still below ROP, already alerted once — keep red but suppress date
            reorder_status = "🔴 YES"
            order_by_str   = "—"
        else:
            reorder_status = "—"
            order_by_str   = "—"
            rop_already_triggered = False   # stock is fine, reset

        po_qty_week = sum(qty for rel_iso, qty in po_released.items() if _iso_in_week(rel_iso, ws, we))
        po_released_str = f"{po_qty_week:,} {unit}" if po_qty_week else "—"

        # Est. Receiving Date = actual arrival of PO released this week (if any)
        if po_qty_week > 0:
            released_this_week_arrivals = [
                _arr(rel_iso)
                for rel_iso in po_released
                if _iso_in_week(rel_iso, ws, we)
            ]
            receiving_date = min(released_this_week_arrivals) if released_this_week_arrivals else ws + datetime.timedelta(days=rm_lt)
        elif in_transit_qty > 0:
            in_transit_arrivals = [
                _arr(rel_iso)
                for rel_iso in po_released
                if (lambda ad: datetime.date.fromisoformat(rel_iso) <= we and ad > we)(_arr(rel_iso))
            ]
            receiving_date = min(in_transit_arrivals) if in_transit_arrivals else ws + datetime.timedelta(days=rm_lt)
        else:
            receiving_date = ws + datetime.timedelta(days=rm_lt)

        # Actual vs estimated note
        default_date = ws + datetime.timedelta(days=rm_lt)
        actual_arr_note = ""
        if po_qty_week > 0:
            for rel_iso in po_released:
                if _iso_in_week(rel_iso, ws, we) and actual_arrivals and rel_iso in actual_arrivals:
                    actual_arr_note = "📦 Actual"
                    break

        row = {
            "Week":                           wk_num,
            "Date Range":                     f"{ws.strftime('%d %b')} – {we.strftime('%d %b %Y')}",
            f"This Comp Req ({unit})":        gross_req,
            f"Opening Stock ({unit})":        round(opening),
            f"Net Requirement ({unit})":      net_req,
            f"Closing Stock ({unit})":        round(max(0, closing)),
            f"🏭 In Production ({unit})":        in_transit_qty if in_transit_qty > 0 else 0,
            "Reorder?":                       reorder_status,
            "📅 Order By":                    order_by_str,
            "PO Released?":                   po_released_str,
            "Receiving Date":                 f"{receiving_date.strftime('%d %b %Y')} {'📦' if actual_arr_note else ''}".strip(),
        }
        if rm_type == "reel":
            row[f"Also Used ({unit})"] = also_used if also_used > 0 else 0
            row[f"Total Pool Req ({unit})"] = total_consumption

        rows.append(row)
        stock = closing

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────
# 5.  CHART HELPERS
# ─────────────────────────────────────────────────────────────
def fg_chart(r: dict, demand_events: list) -> go.Figure:
    df = r["sim_df"].copy()
    if df.empty:
        return go.Figure()
    rop    = r["rop"]
    tgt    = r["raw"]["target_fg"]
    ss     = r["safety_stock"]
    max_fg = r.get("max_fg", tgt + ss)
    today  = datetime.date.today()
    # Ensure date column is always pd.Timestamp so scatter x-axis works correctly
    df["date"] = pd.to_datetime(df["date"])

    def _ms(val) -> int:
        """Convert any date/Timestamp to Unix milliseconds for plotly shape functions."""
        return int(pd.Timestamp(val).value // 1_000_000)

    today_ms = _ms(today)

    fig = go.Figure()

    df_hist   = df[df["is_historical"]]
    df_future = df[~df["is_historical"]]

    # Historical shading
    if not df_hist.empty:
        fig.add_vrect(
            x0=_ms(df_hist["date"].min()), x1=today_ms,
            fillcolor="rgba(100,100,100,0.06)", line_width=0,
            annotation_text="Historical", annotation_position="top left",
            annotation_font_color="#9CA3AF", annotation_font_size=10,
        )

    # SS zone
    fig.add_hrect(
        y0=0, y1=ss, fillcolor="rgba(239,68,68,0.05)", line_width=0,
        annotation_text="Safety Stock Zone", annotation_position="right",
        annotation_font_color="#EF4444", annotation_font_size=10,
    )

    # Historical FG line (grey)
    if not df_hist.empty:
        fig.add_trace(go.Scatter(
            x=df_hist["date"], y=df_hist["closing_fg"],
            mode="lines", name="FG (Historical)",
            line=dict(color="#9CA3AF", width=2),
            fill="tozeroy", fillcolor="rgba(156,163,175,0.08)",
        ))

    # Future FG line (blue) with seamless connector
    if not df_future.empty:
        if not df_hist.empty:
            fig.add_trace(go.Scatter(
                x=[df_hist["date"].iloc[-1], df_future["date"].iloc[0]],
                y=[df_hist["closing_fg"].iloc[-1], df_future["closing_fg"].iloc[0]],
                mode="lines", line=dict(color="#2563EB", width=2.5), showlegend=False,
            ))
        fig.add_trace(go.Scatter(
            x=df_future["date"], y=df_future["closing_fg"],
            mode="lines", name="FG Projection (Future)",
            line=dict(color="#2563EB", width=2.5),
            fill="tozeroy", fillcolor="rgba(37,99,235,0.10)",
        ))

    # Today line
    fig.add_vline(x=today_ms, line_color="#374151", line_width=1.5, line_dash="dot",
                  annotation_text="Today", annotation_position="top",
                  annotation_font_color="#374151", annotation_font_size=10)

    # WO Arrivals
    arrivals = df[df["arrival_qty"] > 0]
    if not arrivals.empty:
        for subset, color, dark, name, size in [
            (arrivals[arrivals["is_historical"]],  "#16A34A", "#14532D", "WO Arrived (Historical)", 14),
            (arrivals[~arrivals["is_historical"]], "#34D399", "#065F46", "WO Arrival (Planned)",    12),
        ]:
            if not subset.empty:
                fig.add_trace(go.Scatter(
                    x=subset["date"], y=subset["opening_fg"],
                    mode="markers", name=name,
                    marker=dict(symbol="triangle-up", size=size, color=color,
                                line=dict(width=1, color=dark)),
                    customdata=subset["arrival_qty"].values,
                    hovertemplate="<b>WO Arrived</b>: %{x|%d %b %Y}<br>Qty: <b>%{customdata:,}</b><extra></extra>",
                ))

    # Demand events
    if demand_events:
        # Use Customer PO Date for demand event positions on the chart
        def _ev_date(ev):
            d = ev.get("date")
            if d is None: return None
            if hasattr(d, "date"): return d.date()
            return d

        ev_past   = [(_ev_date(ev), ev["qty"], ev.get("label","")) for ev in demand_events
                     if isinstance(_ev_date(ev), datetime.date) and _ev_date(ev) < today]
        ev_future = [(_ev_date(ev), ev["qty"], ev.get("label","")) for ev in demand_events
                     if isinstance(_ev_date(ev), datetime.date) and _ev_date(ev) >= today]

        for ev_group, color, name in [
            (ev_past,   "#F59E0B", "Customer PO (Historical)"),
            (ev_future, "#F97316", "Customer PO (Future)"),
        ]:
            if not ev_group:
                continue
            dates_raw, qtys, labels = zip(*ev_group)
            dates_ts = [pd.Timestamp(d) for d in dates_raw]
            ev_fg = []
            for d_ts in dates_ts:
                row = df[df["date"] == d_ts]
                ev_fg.append(int(row["closing_fg"].values[0]) if not row.empty else 0)
            fig.add_trace(go.Scatter(
                x=dates_ts, y=ev_fg, mode="markers", name=name,
                marker=dict(symbol="triangle-down", size=12, color=color,
                            line=dict(width=1.5, color="#92400E")),
                customdata=list(zip(qtys, labels)),
                hovertemplate="<b>Customer PO</b>: %{x|%d %b %Y}<br>Qty: <b>%{customdata[0]:,}</b><br>%{customdata[1]}<extra></extra>",
            ))

        # WO→PO fulfillment arrows (historical only)
        for ev in demand_events:
            ev_date = ev["date"] if isinstance(ev["date"], datetime.date) else None
            if ev_date is None or ev_date >= today:
                continue
            ev_ts = pd.Timestamp(ev_date)
            covering_arr_ts = None
            for wo_rel, wo_info in sorted(r.get("wo_release_log", {}).items()):
                if not isinstance(wo_rel, datetime.date):
                    continue
                arr_d = wo_info.get("arrival_date")
                if arr_d and arr_d <= ev_date:
                    covering_arr_ts = pd.Timestamp(arr_d)
            if covering_arr_ts is not None:
                arr_row = df[df["date"] == covering_arr_ts]
                ev_row  = df[df["date"] == ev_ts]
                if not arr_row.empty and not ev_row.empty:
                    fig.add_annotation(
                        x=_ms(covering_arr_ts), y=arr_row["opening_fg"].values[0],
                        ax=_ms(ev_ts),          ay=ev_row["closing_fg"].values[0],
                        xref="x", yref="y", axref="x", ayref="y",
                        arrowhead=2, arrowsize=1, arrowwidth=1.2,
                        arrowcolor="rgba(245,158,11,0.5)", showarrow=True, text="",
                    )

    # SS breaches (future only)
    ss_breach = df[df["below_ss"] & ~df["is_historical"]]
    if not ss_breach.empty:
        fig.add_trace(go.Scatter(
            x=ss_breach["date"], y=ss_breach["closing_fg"], mode="markers",
            name="⚠️ SS Breached",
            marker=dict(symbol="circle", size=8, color="#DC2626", line=dict(width=2, color="#DC2626")),
        ))

    rop_hits = df[df["rop_hit"]]
    if not rop_hits.empty:
        fig.add_trace(go.Scatter(
            x=rop_hits["date"], y=rop_hits["closing_fg"], mode="markers", name="ROP Trigger",
            marker=dict(symbol="x", size=10, color="#EF4444"),
        ))

    confirmed = df[df["wo_issued"]]
    if not confirmed.empty:
        fig.add_trace(go.Scatter(
            x=confirmed["date"], y=confirmed["closing_fg"], mode="markers", name="WO Confirmed",
            marker=dict(symbol="star", size=12, color="#7C3AED"),
        ))

    # Reference lines
    for y, name, color, dash in [
        (max_fg, f"Max FG ({max_fg:,})",  "#1D4ED8", "longdash"),
        (tgt,    f"MTS ({tgt:,})",        "#10B981", "dot"),
        (rop,    f"ROP ({rop:,})",        "#EF4444", "dash"),
        (ss,     f"SS ({ss:,})",          "#F59E0B", "dashdot"),
    ]:
        fig.add_hline(y=y, line_color=color, line_dash=dash,
                      annotation_text=name, annotation_position="right",
                      annotation_font_color=color)

    # Month dividers (using Timestamps since df["date"] is now Timestamp)
    seen_months: set = set()
    for ts in df["date"]:
        if ts.day == 1:
            key = (ts.year, ts.month)
            if key not in seen_months:
                seen_months.add(key)
                fig.add_vline(x=_ms(ts), line_dash="dot", line_color="rgba(100,100,100,0.20)")
                fig.add_annotation(
                    x=_ms(ts), y=1, yref="paper",
                    text=ts.strftime("%b %Y"), showarrow=False,
                    xanchor="left", yanchor="bottom",
                    font=dict(size=9, color="gray"),
                )

    fig.update_layout(
        height=480, margin=dict(l=0, r=160, t=30, b=0),
        legend=dict(orientation="h", y=-0.18),
        yaxis_title="Cartons", xaxis_title=None,
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
    )
    fig.update_xaxes(showgrid=True, gridcolor="rgba(0,0,0,0.06)")
    fig.update_yaxes(showgrid=True, gridcolor="rgba(0,0,0,0.06)")
    return fig

def capacity_chart_monthly(cap_result: dict, processes: list) -> go.Figure:
    """Bar chart: monthly utilisation per process."""
    months = [r["month"] for r in cap_result["cap_rows"]]
    fig    = go.Figure()
    colors = ["#2563EB", "#7C3AED", "#10B981", "#F59E0B"]

    for i, proc in enumerate(processes):
        utils = [r["proc_utils"][proc["name"]]["util_pct"] for r in cap_result["cap_rows"]]
        fig.add_trace(go.Bar(
            x=months, y=utils, name=proc["name"],
            marker_color=colors[i % len(colors)],
            text=[f"{u:.1f}%" for u in utils], textposition="outside",
        ))

    fig.add_hline(y=90, line_dash="dash", line_color="#EF4444",
                  annotation_text="90% limit", annotation_position="right")
    fig.update_layout(
        height=340, yaxis_title="Utilisation %", barmode="group",
        yaxis_range=[0, 120],
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=0, r=0, t=30, b=0),
        legend=dict(orientation="h", y=-0.18),
    )
    return fig


def mrp_reel_chart(r: dict) -> go.Figure:
    raw          = r["raw"]
    moq          = raw["reel_moq"]
    rop_kg       = r["reel_rop_kg"]
    months       = list(r["reel_weight_monthly"].keys())
    monthly_need = list(r["reel_weight_monthly"].values())
    stock = [raw["reel_stock"]]
    for need in monthly_need[:-1]:
        s = stock[-1] - need
        if s <= rop_kg:
            s += max(moq, need)
        stock.append(max(0, s))
    fig = go.Figure()
    fig.add_trace(go.Bar(x=months, y=stock,        name="Reel Stock (KG)",   marker_color="#3B82F6"))
    fig.add_trace(go.Bar(x=months, y=monthly_need, name="Monthly Need (KG)", marker_color="#93C5FD"))
    fig.add_hline(y=rop_kg, line_dash="dash", line_color="#EF4444",
                  annotation_text=f"ROP {rop_kg:,} KG", annotation_position="right")
    fig.add_hline(y=moq, line_dash="dot", line_color="#F59E0B",
                  annotation_text=f"MOQ {moq:,} KG", annotation_position="right")
    fig.update_layout(height=260, yaxis_title="KG", barmode="group",
                      paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                      margin=dict(l=0, r=120, t=30, b=0))
    return fig


def mrp_pouch_chart(r: dict) -> go.Figure:
    raw          = r["raw"]
    moq          = raw["pouch_moq"]
    rop          = r["pouch_rop"]
    months       = list(r["pouch_monthly"].keys())
    monthly_need = list(r["pouch_monthly"].values())
    stock = [raw["pouch_stock"]]
    for need in monthly_need[:-1]:
        s = stock[-1] - need
        if s <= rop:
            s += max(moq, need)
        stock.append(max(0, s))
    fig = go.Figure()
    fig.add_trace(go.Bar(x=months, y=stock,        name="Pouch Stock",  marker_color="#8B5CF6"))
    fig.add_trace(go.Bar(x=months, y=monthly_need, name="Monthly Need", marker_color="#C4B5FD"))
    fig.add_hline(y=rop, line_dash="dash", line_color="#EF4444",
                  annotation_text=f"ROP {rop:,}", annotation_position="right")
    fig.update_layout(height=260, yaxis_title="Units", barmode="group",
                      paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                      margin=dict(l=0, r=120, t=30, b=0))
    return fig


# ─────────────────────────────────────────────────────────────
# 6.  KPI CARD
# ─────────────────────────────────────────────────────────────
def kpi_card(label, value, delta=None, help_text=None, color=None):
    delta_html = ""
    if delta is not None:
        sign = "+" if delta >= 0 else ""
        dcol = "#16A34A" if delta >= 0 else "#DC2626"
        darr = "▲" if delta >= 0 else "▼"
        delta_html = f'<span style="color:{dcol};font-size:.8rem">{darr} {sign}{delta}</span>'
    bc = color or "#2563EB"
    st.markdown(f"""
    <div style="border:1px solid {bc}33;border-left:4px solid {bc};
    border-radius:8px;padding:14px 16px;background:#fff;min-height:80px;">
      <div style="font-size:.75rem;color:#6B7280;font-weight:600;
      letter-spacing:.05em;text-transform:uppercase">{label}</div>
      <div style="font-size:1.6rem;font-weight:700;color:#111827;
      line-height:1.2;margin-top:4px">{value}</div>
      {delta_html}
      {"<div style='font-size:.7rem;color:#9CA3AF;margin-top:3px'>"+help_text+"</div>" if help_text else ""}
    </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# 7.  SESSION STATE
# ─────────────────────────────────────────────────────────────
if "issued_wos"    not in st.session_state:
    st.session_state["issued_wos"]    = {}
if "demand_events" not in st.session_state:
    st.session_state["demand_events"] = []
if "po_released"   not in st.session_state:
    st.session_state["po_released"]   = {"reel": {}, "pouch": {}}

# ── Global RM Registry (shared across all SKUs) ──────────────
# reel_registry: {reel_name → {stock_kg, moq, total_lt, order_cycle, gsm, width, length}}
# pouch_registry: {pouch_name → {stock_units, moq, lt, units_per_carton}}
# reel_pos_released: {reel_name → {iso_date → qty_kg}}
# pouch_pos_released: {pouch_name → {iso_date → qty_units}}
if "reel_registry"       not in st.session_state:
    st.session_state["reel_registry"]       = {}
if "pouch_registry"      not in st.session_state:
    st.session_state["pouch_registry"]      = {}
if "reel_pos_released"   not in st.session_state:
    st.session_state["reel_pos_released"]   = {}
if "pouch_pos_released"  not in st.session_state:
    st.session_state["pouch_pos_released"]  = {}
if "reel_actual_arrivals" not in st.session_state:
    st.session_state["reel_actual_arrivals"] = {}   # {reel_nm → {rel_iso → actual_date}}


def check_rm_availability(raw: dict, wo_qty: int, wo_date: datetime.date,
                           reel_registry: dict, reel_pos_released: dict,
                           already_issued_wos: dict,
                           reel_actual_arrivals: dict = None,
                           all_issued_wos_map: dict = None,
                           all_raw_map: dict = None) -> dict:
    """
    Check if all RM components have sufficient stock for a WO of wo_qty cartons
    on wo_date, accounting for:
      - Current registry stock
      - KG already consumed by previously issued WOs
      - POs in production (released but not yet arrived)

    Returns:
    {
      "ok": bool,
      "earliest_ok_date": date or None,
      "components": [
        {
          "reel_name": str,
          "required_kg": int,
          "available_kg": int,
          "shortfall_kg": int,
          "po_arriving": date or None,  # earliest PO that covers shortfall
          "status": "✅ OK" | "🔴 Short" | "🚚 Covered by PO arriving dd Mon"
        }
      ]
    }
    """
    results = []
    overall_ok = True
    earliest_ok = wo_date  # latest "earliest possible date" across all components

    # Group components by rm_material_id (unique RM) — NOT by reel_name
    # reel_name is shared by Lid/Strip/Wall; rm_material_id is unique per component
    from collections import defaultdict
    rm_id_comps: dict[str, list] = defaultdict(list)
    for comp in raw.get("bom_components", []):
        rm_id = (comp.get("rm_material_id") or comp["reel_name"]).strip()
        rm_id_comps[rm_id].append(comp)

    for rm_id, comps in rm_id_comps.items():
        reel_nm = comps[0]["reel_name"]   # registry is still keyed by reel_name
        # Total KG needed for this reel for this WO
        total_req_kg = 0
        for comp in comps:
            sheets = math.ceil(wo_qty / comp["ups"]) if comp["ups"] else 0
            _mult  = comp.get("multiplier", 1) or 1
            kg     = math.ceil(
                sheets * comp["width"] * comp["length"]
                * comp["gsm"] / ( (1 - comp["wastage"]/100) * 1_000_000_000 ) * _mult
            ) if sheets else 0
            total_req_kg += kg

        # Current registry stock
        reg_stk = reel_registry.get(_reel_reg_key(comps[0]), reel_registry.get(reel_nm, {})).get("stock_kg", 0)
        # Fall back to sibling entry with same reel_name if this rm_id has 0 stock
        if reg_stk == 0:
            _siblings_stk = [info.get("stock_kg", 0) for k, info in reel_registry.items()
                              if info.get("reel_name") == reel_nm and info.get("stock_kg", 0) > 0]
            if _siblings_stk:
                reg_stk = max(_siblings_stk)

        # Subtract KG consumed by ALL SKUs' confirmed WOs on or before wo_date
        already_consumed = 0

        # Build map: sku_label → (issued_wos, [list of components using this reel])
        # Must use a list — multiple components per SKU can share the same reel
        skus_to_check: dict[str, tuple] = {}
        if all_issued_wos_map and all_raw_map:
            for _sl, _rd in all_raw_map.items():
                _sl_wos = all_issued_wos_map.get(_sl, {})
                _matching_comps = [
                    _comp for _comp in _rd.get("bom_components", [])
                    if _comp["reel_name"] == reel_nm
                ]
                if _matching_comps:
                    skus_to_check[_sl] = (_sl_wos, _matching_comps)
        else:
            skus_to_check["__current__"] = (already_issued_wos, comps)

        for _sl, (_wos, _comp_list) in skus_to_check.items():
            for ds, info in _wos.items():
                if not info.get("issued"):
                    continue
                try:
                    wo_d = datetime.date.fromisoformat(ds)
                except:
                    continue
                if wo_d > wo_date:
                    continue
                # Only subtract consumption from WOs whose RM was drawn AFTER
                # today's registry stock snapshot (i.e. future WOs that haven't
                # yet pulled from the physical stock).
                # Historical WOs (arrival already in the past) already consumed
                # their RM — the registry stock is already net of those.
                _lt_d  = info.get("actual_lt_override", 10)
                _arr_d = wo_d + datetime.timedelta(days=int(_lt_d))
                _today_check = datetime.date.today()
                if _arr_d <= _today_check:
                    # Already arrived → RM was consumed in the past → skip
                    continue
                prev_qty = info.get("gross_qty", info.get("actual_produced", info.get("wo_qty", 0)))
                for _comp in _comp_list:
                    sh = math.ceil(prev_qty / _comp["ups"]) if _comp["ups"] else 0
                    _mult_c = _comp.get("multiplier", 1) or 1
                    kg = math.ceil(
                        sh * _comp["width"] * _comp["length"]
                        * _comp["gsm"] / 1_000_000_000 * _mult_c
                    ) if sh else 0
                    already_consumed += kg

        available_kg = max(0, reg_stk - already_consumed)
        rm_lt = comps[0]["total_lt"]

        # Add PO quantities that arrive ON OR BEFORE the WO date
        # POs keyed by rm_material_id; fall back to reel_nm for backward compat
        _po_key_ra = (comps[0].get("rm_material_id") or reel_nm)
        pos_for_reel   = reel_pos_released.get(_po_key_ra) or reel_pos_released.get(reel_nm, {})
        actual_arr_map = ((reel_actual_arrivals or {}).get(_po_key_ra) or \
                         (reel_actual_arrivals or {}).get(reel_nm, {}))
        for rel_iso, po_qty in pos_for_reel.items():
            try:
                rel_d  = datetime.date.fromisoformat(rel_iso)
                arr_d  = actual_arr_map.get(rel_iso, rel_d + datetime.timedelta(days=rm_lt))
            except:
                continue
            if arr_d <= wo_date:
                available_kg += po_qty

        if available_kg >= total_req_kg:
            results.append({
                "reel_name":      reel_nm,
                "rm_material_id": comps[0].get("rm_material_id", ""),
                "rm_material_desc": comps[0].get("rm_material_desc", ""),
                "required_kg":    total_req_kg,
                "available_kg":   available_kg,
                "shortfall_kg":   0,
                "po_arriving":    None,
                "status":         "✅ OK",
                "has_shortfall":  False,
            })
        else:
            shortfall = total_req_kg - available_kg
            overall_ok = False   # stock is short — always mark as not immediately available
            pos = reel_pos_released.get(_po_key_ra) or reel_pos_released.get(reel_nm, {})
            covering_po_date = None
            cumulative = available_kg
            for rel_iso, qty in sorted(pos.items()):
                try:
                    rel_d  = datetime.date.fromisoformat(rel_iso)
                    arr_d  = actual_arr_map.get(rel_iso, rel_d + datetime.timedelta(days=rm_lt))
                except:
                    continue
                if arr_d > wo_date:
                    cumulative += qty
                    if cumulative >= total_req_kg:
                        covering_po_date = arr_d
                        break

            if covering_po_date:
                status = f"🚚 Covered by PO arriving {covering_po_date.strftime('%d %b %Y')} — WO must be issued on/after that date"
                earliest_ok = max(earliest_ok, covering_po_date) if earliest_ok is not None else covering_po_date
            else:
                status = f"🔴 Short by {shortfall:,} KG — place RM order first"
                earliest_ok = None

            results.append({
                "reel_name":      reel_nm,
                "rm_material_id": comps[0].get("rm_material_id", ""),
                "rm_material_desc": comps[0].get("rm_material_desc", ""),
                "required_kg":    total_req_kg,
                "available_kg":   available_kg,
                "shortfall_kg":   shortfall,
                "po_arriving":    covering_po_date,
                "status":         status,
                "has_shortfall":  True,
            })

    has_any_shortfall = any(c["has_shortfall"] for c in results)
    po_covered_all    = all(c["status"].startswith("✅") or c["status"].startswith("🚚") for c in results)
    all_covered_by    = max((c["po_arriving"] for c in results if c.get("po_arriving")), default=None) if po_covered_all else None

    return {
        "stock_ok":         not has_any_shortfall,           # True = enough stock right now, WO can go today
        "po_covered":       has_any_shortfall and po_covered_all,  # True = short but PO will cover, wait for arrival
        "earliest_ok_date": wo_date if not has_any_shortfall else all_covered_by,
        "components":       results,
    }


def _reel_reg_key(comp: dict) -> str:
    """Registry key = rm_material_id when present, else reel_name."""
    return (comp.get("rm_material_id") or "").strip() or comp["reel_name"]

def _reel_display(comp: dict) -> str:
    """Human-readable label: 'rm_id — description' or reel_name if no rm_id."""
    rid  = (comp.get("rm_material_id") or "").strip()
    desc = (comp.get("rm_material_desc") or "").strip()
    if rid:
        return f"{rid} — {desc}" if desc else rid
    return comp["reel_name"]

def _ensure_reel_in_registry(comp: dict):
    """Auto-register a reel from BOM if not already in registry.
    Key = rm_material_id when present, else reel_name."""
    nm = _reel_reg_key(comp)
    if nm not in st.session_state["reel_registry"]:
        st.session_state["reel_registry"][nm] = {
            "stock_kg":    0,
            "moq":         comp["moq_kg"],
            "total_lt":    comp["total_lt"],
            "order_cycle": 60,
            "gsm":         comp["gsm"],
            "width":       comp["width"],
            "length":      comp["length"],
            "reel_name":   comp["reel_name"],          # preserve for display
            "rm_material_id": (comp.get("rm_material_id") or "").strip(),
            "rm_material_desc": (comp.get("rm_material_desc") or "").strip(),
            "display":     _reel_display(comp),        # "rm_id — desc" label
        }
    if nm not in st.session_state["reel_pos_released"]:
        st.session_state["reel_pos_released"][nm] = {}


def _ensure_pouch_in_registry(raw: dict):
    """Auto-register a pouch from SKU params if not already in registry."""
    nm = raw["pouch_name"]
    if nm not in st.session_state["pouch_registry"]:
        st.session_state["pouch_registry"][nm] = {
            "stock_units":    int(raw.get("pouch_stock", 0)),
            "moq":            int(raw.get("pouch_moq", 20000)),
            "lt":             int(raw.get("pouch_lt",  3)),
            "units_per_carton": raw.get("pouch_per_1m", 20000) / 1_000_000,
        }
    if nm not in st.session_state["pouch_pos_released"]:
        st.session_state["pouch_pos_released"][nm] = {}


# ─────────────────────────────────────────────────────────────
# 8.  SIDEBAR
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📦 ITC Ecobyte\n**Supply Chain Planner**")
    st.divider()

    st.markdown("### 📂 Upload Files")
    sku_file = st.file_uploader(
        "① SKU Master Excel (.xlsx)", type=["xlsx"], key="sku_master",
        help="Contains SKU Basic Details, BOM-RM Details, Capacity Details, WO Log sheets"
    )
    billing_file = st.file_uploader(
        "② Demand Events & Billing Excel (.xlsx / .xls)",
        type=["xlsx", "xls"],
        key="billing_upload",
        help="SAP billing export — Customer PO Date, Material Code, Billed Qty, Invoice, Transporter, LR Number etc."
    )
    fg_file = st.file_uploader(
        "③ Current FG Inventory (.xls / .xlsx)", type=["xls","xlsx"], key="fg_inv",
        help="FG stock list — Material ID in col A, Total Stock in col H"
    )
    rm_file = st.file_uploader(
        "④ Current RM Inventory (.xls / .xlsx)", type=["xls","xlsx"], key="rm_inv",
        help="RM stock list — RM Material Code in col A, Total Stock (KG) in col H"
    )
    del_track_sidebar = st.file_uploader(
        "⑤ Delivery Tracking Excel (.xlsx)", type=["xlsx","xls"], key="delivery_tracker_sidebar",
        help="Tracking file with 'Arrived at TCI?' column — pending shipments shown in Delivery tab"
    )
    # Keep po_hist_file pointing to billing_file for backward compat
    po_hist_file       = billing_file
    demand_events_file = billing_file
    st.divider()

    # ── Parse files — cache in session state, only re-parse when file changes ──
    # Key = (filename, filesize) so a new upload always triggers a fresh parse.
    # This prevents re-reading a 9MB Excel on every button click / widget interaction.

    def _file_key(f):
        if f is None:
            return None
        return (f.name, f.size)

    all_results_raw: dict = {}
    fg_inventory:    dict = {}
    rm_inventory:    dict = {}

    # ── SKU Master ───────────────────────────────────────────
    if sku_file:
        _sku_key_cache = _file_key(sku_file)
        if st.session_state.get("_cached_sku_key") != _sku_key_cache:
            try:
                _sku_bytes = sku_file.read()
                st.session_state["_cached_sku_data"] = parse_sku_master(_sku_bytes)
                st.session_state["_cached_sku_key"]  = _sku_key_cache
            except Exception as e:
                st.error(f"Error parsing SKU Master: {e}")
        all_results_raw = st.session_state.get("_cached_sku_data", {})

    # ── Billing / PO History ─────────────────────────────────
    if billing_file:
        _bill_key_cache = _file_key(billing_file)
        if st.session_state.get("_cached_bill_key") != _bill_key_cache:
            try:
                _bill_bytes = billing_file.read()
                st.session_state["_cached_po_df"]        = parse_po_history(_bill_bytes)
                st.session_state["_cached_demand_df"]    = parse_demand_events_excel(_bill_bytes)
                st.session_state["_cached_bill_key"]     = _bill_key_cache
            except Exception as e:
                st.sidebar.warning(f"Billing file parse error: {e}")
    po_df_global_cached            = st.session_state.get("_cached_po_df",     pd.DataFrame())
    demand_events_df_global_cached = st.session_state.get("_cached_demand_df", pd.DataFrame())

    # ── FG Inventory ─────────────────────────────────────────
    if fg_file:
        _fg_key_cache = _file_key(fg_file)
        if st.session_state.get("_cached_fg_key") != _fg_key_cache:
            try:
                _fg_bytes = fg_file.read()
                st.session_state["_cached_fg_inv"] = parse_fg_inventory(_fg_bytes)
                st.session_state["_cached_fg_key"] = _fg_key_cache
            except Exception as e:
                st.warning(f"FG file issue: {e}")
    fg_inventory = st.session_state.get("_cached_fg_inv", {})

    # Apply FG inventory to all_results_raw
    if fg_inventory and all_results_raw:
        matched = 0
        for mat_id, fg_stk in fg_inventory.items():
            if mat_id in all_results_raw:
                all_results_raw[mat_id]["current_fg"] = int(fg_stk)
                matched += 1
        if matched == 0:
            fg_keys  = list(fg_inventory.keys())
            sku_keys = list(all_results_raw.keys())
            for fk in fg_keys:
                fk_clean = fk.strip().upper()
                for sk in sku_keys:
                    sk_clean = sk.strip().upper()
                    if fk_clean == sk_clean or fk_clean in sk_clean or sk_clean in fk_clean:
                        all_results_raw[sk]["current_fg"] = int(fg_inventory[fk])
                        matched += 1
                        break
            if matched == 0:
                st.warning(
                    f"⚠️ FG file parsed {len(fg_inventory)} material(s) but none matched SKU IDs.  \n"
                    f"FG file IDs (first 3): {list(fg_inventory.keys())[:3]}  \n"
                    f"SKU Master IDs (first 3): {list(all_results_raw.keys())[:3]}"
                )
            else:
                st.sidebar.success(f"✅ FG matched {matched} SKU(s)")
        else:
            st.sidebar.success(f"✅ FG loaded: {matched} SKU(s) updated")
    elif fg_file and not fg_inventory:
        st.warning("⚠️ FG file uploaded but no stock data could be parsed.")

    # ── RM Inventory ──────────────────────────────────────────
    if rm_file:
        _rm_key_cache = _file_key(rm_file)
        if st.session_state.get("_cached_rm_key") != _rm_key_cache:
            try:
                _rm_bytes = rm_file.read()
                st.session_state["_cached_rm_inv"] = parse_rm_inventory(_rm_bytes)
                st.session_state["_cached_rm_key"] = _rm_key_cache
                st.sidebar.success(f"✅ RM file parsed: {len(st.session_state['_cached_rm_inv'])} materials")
            except Exception as e:
                st.warning(f"RM file issue: {e}")
    rm_inventory = st.session_state.get("_cached_rm_inv", {})

    # ── Delivery Tracking (sidebar upload — also accessible from Delivery tab) ──
    _del_src = del_track_sidebar
    if _del_src:
        _dtk_key = _file_key(_del_src)
        if st.session_state.get("_del_track_key") != _dtk_key:
            try:
                _dtk_bytes = _del_src.read()
                _dtk_df_parsed = pd.read_excel(io.BytesIO(_dtk_bytes), header=0, engine="openpyxl")
                _dtk_df_parsed.columns = [str(c).strip() for c in _dtk_df_parsed.columns]
                st.session_state["_del_track_df"]    = _dtk_df_parsed
                st.session_state["_del_track_bytes"] = _dtk_bytes
                st.session_state["_del_track_key"]   = _dtk_key
                st.sidebar.success(f"✅ Delivery tracking: {len(_dtk_df_parsed)} records loaded")
            except Exception as _e:
                st.sidebar.warning(f"Delivery tracking file issue: {_e}")

    def _rm_stock_for_comp(comp, rm_inv):
        for key in [
            comp.get("rm_material_id", ""),
            comp.get("rm_code", ""),
            comp.get("reel_name", ""),
        ]:
            key = (key or "").strip()
            if key and key in rm_inv:
                return rm_inv[key]
        return None

    # ── Track data source — reset registries only when files actually change ──
    # Use the cached file keys (name+size stored when file was parsed) NOT the live
    # file uploader objects — on Streamlit Cloud the uploader returns None briefly
    # between reruns, causing spurious registry wipes.
    if all_results_raw:
        _src_key = (
            st.session_state.get("_cached_sku_key",  ""),
            st.session_state.get("_cached_fg_key",   ""),
            st.session_state.get("_cached_rm_key",   ""),
            st.session_state.get("_cached_bill_key", ""),
        )
        if st.session_state.get("_data_source") != _src_key:
            st.session_state["_data_source"]       = _src_key
            st.session_state["reel_registry"]      = {}
            st.session_state["pouch_registry"]     = {}
            st.session_state["reel_pos_released"]  = {}
            st.session_state["pouch_pos_released"] = {}

        # Register all reels (keyed by rm_material_id) and apply RM inventory stock.
        # Always re-apply stock from rm_inventory so the Registry tab stays populated
        # even after reruns (registry stock is overwritten only from the parsed file,
        # never from a stale 0 — manual edits in the Registry tab are preserved because
        # _ensure_reel_in_registry skips already-registered keys).
        for mat_id, raw in all_results_raw.items():
            for comp in raw.get("bom_components", []):
                _ensure_reel_in_registry(comp)
                _rk = _reel_reg_key(comp)
                qty = _rm_stock_for_comp(comp, rm_inventory) if rm_inventory else None
                if qty is not None:
                    # Always apply file stock — this keeps the registry in sync with
                    # the uploaded RM file across reruns. Manual registry edits are
                    # intentionally overwritten on re-upload (file is the source of truth).
                    st.session_state["reel_registry"][_rk]["stock_kg"] = qty
            _ensure_pouch_in_registry(raw)

        # Purge any stale reel_name-keyed entries (from old sessions before rm_id keying)
        _valid_keys = set()
        for mat_id, raw in all_results_raw.items():
            for comp in raw.get("bom_components", []):
                _valid_keys.add(_reel_reg_key(comp))
        for _stale in [k for k in list(st.session_state["reel_registry"].keys()) if k not in _valid_keys]:
            del st.session_state["reel_registry"][_stale]
        for _stale in [k for k in list(st.session_state["reel_pos_released"].keys()) if k not in _valid_keys]:
            # Don't delete POs — they may be under rm_id keys that ARE valid
            pass

    # ── Demand events auto-population from cached billing data ──
    demand_events_df_global = demand_events_df_global_cached
    po_df_global            = po_df_global_cached

    if not demand_events_df_global.empty and all_results_raw:
        for _sl in all_results_raw:
            _k_de = _sku_key(_sl, "demand_events")
            if _k_de not in st.session_state:
                st.session_state[_k_de] = []
            _sku_rows = demand_events_df_global[demand_events_df_global["mat_id"] == _sl]
            if not _sku_rows.empty:
                _existing_labels = {(ev["date"], ev.get("label","")) for ev in st.session_state[_k_de]}
                for _, _row in _sku_rows.iterrows():
                    _ev_date   = _row["date"].date()   if pd.notna(_row.get("date"))      else None
                    _bill_date = None
                    if pd.notna(_row.get("bill_date")) and _row.get("bill_date") is not None:
                        try:
                            _bill_date = pd.Timestamp(_row["bill_date"]).date()
                        except Exception:
                            pass
                    if _ev_date is None and _bill_date is None:
                        continue
                    _label = f"Invoice {_row['invoice_no']}" if _row.get("invoice_no") else "From billing export"
                    _key   = (_ev_date or _bill_date, _label)
                    if _key not in _existing_labels:
                        _qty = int(_row["qty"]) if pd.notna(_row.get("qty")) and _row["qty"] > 0 else 0
                        st.session_state[_k_de].append({
                            "date":      _ev_date or _bill_date,
                            "bill_date": _bill_date,
                            "qty":       _qty,
                            "label":     _label,
                        })
                        _existing_labels.add(_key)

    if not all_results_raw:
        st.markdown("### Or enter FG manually")
        st.info("Upload SKU Master Excel to load all 20 SKUs, or manually set FG stock below after uploading SKU Master.")
        if not sku_file:
            st.warning("Please upload the SKU Master Excel file to begin.")
        else:
            st.warning("SKU Master uploaded but no data was parsed — please check the file format.")
        st.stop()

    st.divider()

    # ── FG manual override per SKU ───────────────────────────
    if all_results_raw:
        if not fg_file:
            with st.expander("✏️ Manual FG Stock Override", expanded=False):
                st.caption("Override FG stock per SKU (if not uploading FG inventory file)")
                for mid in list(all_results_raw.keys())[:5]:
                    v = st.number_input(
                        f"{mid[:20]}", min_value=0, max_value=2_000_000,
                        value=int(all_results_raw[mid].get("current_fg", 0)),
                        step=100, key=f"fg_override_{mid}",
                    )
                    all_results_raw[mid]["current_fg"] = v
                if len(all_results_raw) > 5:
                    st.caption(f"...and {len(all_results_raw)-5} more. Upload FG inventory file to set all at once.")

        sku_label = st.selectbox(
            "🔍 Select SKU to plan",
            options=list(all_results_raw.keys()),
            format_func=lambda x: f"{x} — {all_results_raw[x]['mat_desc'][:35]}"
        )

    st.divider()
    # ── Output Excel download ────────────────────────────────
    # (Download button rendered below after Excel is generated)
    import os


# ─────────────────────────────────────────────────────────────
# FIX D: Per-SKU state
# ─────────────────────────────────────────────────────────────
# Guard: stop if no SKU data is available (file not uploaded or parse failed)
if not all_results_raw:
    st.warning("⚠️ No SKU data loaded. Please upload the SKU Master Excel file in the sidebar.")
    st.stop()

if "sku_label" not in dir():
    sku_label = list(all_results_raw.keys())[0]

raw_data = all_results_raw[sku_label]

ensure_sku_state(sku_label)
_k_issued = _sku_key(sku_label, "issued_wos")
_k_demand = _sku_key(sku_label, "demand_events")
_k_po     = _sku_key(sku_label, "po_released")

def get_issued_wos():    return st.session_state[_k_issued]
def get_demand_events(): return st.session_state[_k_demand]
def get_po_released():   return st.session_state[_k_po]


# ─────────────────────────────────────────────────────────────
# 9.  HEADER + COMPUTE
# ─────────────────────────────────────────────────────────────
st.title(f"📦 {raw_data['mat_id']}  —  {raw_data['mat_desc']}")
st.caption(
    f"Factory: {raw_data.get('factory','—')}  |  "
    f"Production LT: {raw_data['prod_lt']} days  "
    f"(incl. SO→WO: {raw_data.get('so_to_wo_days',1)}d)  |  "
    f"MTS Target: {raw_data['target_fg']:,}  |  "
    f"Production MOQ: {raw_data['prod_moq']:,}  |  "
    f"Fixed SS: {raw_data['safety_stock_fixed']:,}  |  "
    f"Wastage: {raw_data['wastage_pct']}%"
)

# ── Auto-populate issued WOs from WO Log sheet in SKU Master ─
_wo_log_entries = raw_data.get("wo_log", [])
if _wo_log_entries:
    _k_issued_cur = _sku_key(sku_label, "issued_wos")
    if _k_issued_cur not in st.session_state:
        st.session_state[_k_issued_cur] = {}
    for _wo in _wo_log_entries:
        try:
            _rel_raw  = _wo.get("release")
            _is_intransit = _wo.get("in_transit", False)
            # release is now a raw datetime object (not stringified)
            if isinstance(_rel_raw, (datetime.date, datetime.datetime)):
                _rel_date = pd.Timestamp(_rel_raw)
            else:
                _rel_date = pd.to_datetime(str(_rel_raw or "").strip(), dayfirst=True, errors="coerce")
            if pd.isna(_rel_date):
                continue
            _ds = _rel_date.date().isoformat()

            if _is_intransit:
                # In-transit: arrival = est_complete date
                _est_raw = _wo.get("est_complete")
                if isinstance(_est_raw, (datetime.date, datetime.datetime)):
                    _arr_date = pd.Timestamp(_est_raw)
                else:
                    _arr_date = pd.to_datetime(str(_est_raw or ""), dayfirst=True, errors="coerce")
                if pd.isna(_arr_date):
                    _arr_date = _rel_date + pd.Timedelta(days=int(_wo.get("planned_lt", raw_data.get("prod_lt", 10))))
                _actual_lt = (_arr_date.date() - _rel_date.date()).days
            else:
                # Historical: use actual WO complete date as arrival — not release + lt
                _cmp_raw = _wo.get("complete")
                if isinstance(_cmp_raw, (datetime.date, datetime.datetime)):
                    _cmp_date = pd.Timestamp(_cmp_raw)
                elif _cmp_raw and "in production" not in str(_cmp_raw).lower():
                    _cmp_date = pd.to_datetime(str(_cmp_raw), dayfirst=True, errors="coerce")
                else:
                    _cmp_date = pd.NaT
                if not pd.isna(_cmp_date) and _cmp_date.date() > _rel_date.date():
                    _actual_lt = (_cmp_date.date() - _rel_date.date()).days
                    _arr_date  = _cmp_date
                else:
                    _actual_lt = int(_wo.get("cycle_time", 0)) or int(_wo.get("planned_lt", 0)) or raw_data.get("prod_lt", 10)
                    _arr_date  = _rel_date + pd.Timedelta(days=_actual_lt)

            if _ds not in st.session_state[_k_issued_cur]:
                _prod = int(_wo.get("target_qty", 0)) if _is_intransit else (int(_wo.get("produced", 0)) or int(_wo.get("target_qty", 0)))
                st.session_state[_k_issued_cur][_ds] = {
                    "issued":             True,
                    "wo_qty":             int(_wo.get("target_qty", 0)),
                    "gross_qty":          int(_wo.get("target_qty", 0)),
                    "actual_produced":    _prod,
                    "actual_lt_override": _actual_lt,
                    "from_wo_log":        True,
                    "in_transit":         _is_intransit,
                    "wo_num":             str(_wo.get("wo_num", "")),
                    "so_num":             str(_wo.get("so_num", "")),
                    "est_complete":       _arr_date.date().isoformat() if _arr_date is not None and not pd.isna(_arr_date) else "",
                }
        except Exception:
            continue

r       = compute(raw_data, get_demand_events(), get_issued_wos(), po_history=po_df_global)
cap_res = compute_capacity(raw_data, get_demand_events())


# Auto-register all reels and pouches from all loaded SKUs
for _sl, _rd in all_results_raw.items():
    for _comp in _rd.get("bom_components", []):
        _ensure_reel_in_registry(_comp)
    _ensure_pouch_in_registry(_rd)

# Compute consolidated BOM requirements across all SKUs
all_issued_wos_map = {sl: st.session_state.get(_sku_key(sl, "issued_wos"), {})
                      for sl in all_results_raw}
bom_req = compute_bom_requirements(all_results_raw, all_issued_wos_map)

# ── Global RM SS/ROP across all SKUs sharing the same rm_material_id ──
global_rm_ss = compute_global_rm_ss(all_results_raw, po_df_global)
st.session_state["_global_rm_ss_cache"] = global_rm_ss

# Build week list once (used in Registry tab)
_today    = datetime.date.today()
_end_date = _today + datetime.timedelta(days=183)
_week_start = _today - datetime.timedelta(days=_today.weekday())
_all_weeks  = []
_ws = _week_start
while _ws <= _end_date:
    _we = _ws + datetime.timedelta(days=6)
    _all_weeks.append((_ws, min(_we, _end_date)))
    _ws += datetime.timedelta(days=7)

# ── Auto-update output Excel on any WO/PO event ─────────────
try:
    # Pass demand events as list-of-dicts per mat_id for the Customer PO sheet
    _all_demand_for_excel = {}
    if not demand_events_df_global.empty:
        for _mat, _grp in demand_events_df_global.groupby("mat_id"):
            _all_demand_for_excel[_mat] = _grp.to_dict("records")
    else:
        _all_demand_for_excel = {sl: st.session_state.get(_sku_key(sl, "demand_events"), [])
                                  for sl in all_results_raw}
    _excel_bytes = write_output_excel(
        all_raw          = all_results_raw,
        all_issued_wos   = all_issued_wos_map,
        all_demand_events= _all_demand_for_excel,
        bom_req          = bom_req,
        reel_registry    = st.session_state.get("reel_registry", {}),
        reel_pos_released= st.session_state.get("reel_pos_released", {}),
    )
except Exception as _e:
    _excel_bytes = None

# ── Download button in sidebar (after Excel bytes are ready) ──
with st.sidebar:
    st.divider()
    if _excel_bytes:
        st.download_button(
            "⬇️ Download Snapshot Report",
            data=_excel_bytes,
            file_name="ecobyte_snapshot.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    elif os.path.exists(OUTPUT_EXCEL_PATH):
        with open(OUTPUT_EXCEL_PATH, "rb") as _f:
            st.download_button(
                "⬇️ Download Snapshot Report",
                data=_f.read(),
                file_name="ecobyte_snapshot.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    pass   # Don't crash the app if output write fails

# ─────────────────────────────────────────────────────────────
# 10.  TABS
# ─────────────────────────────────────────────────────────────
# ── Build reel_name → rm_material_id lookup (for PO key resolution) ──────────
_reel_rm_id_map: dict[str, str] = {}
for _sl_g, _rd_g in all_results_raw.items():
    for _c_g in _rd_g.get("bom_components", []):
        _rn_g  = _c_g.get("reel_name", "")
        _rid_g = _c_g.get("rm_material_id", "")
        if _rn_g and _rid_g:
            _reel_rm_id_map[_rn_g] = _rid_g
st.session_state["_reel_rm_id_map"] = _reel_rm_id_map

tab_overview, tab_fg_status, tab_mps, tab_mrp, tab_registry, tab_capacity, tab_wo, tab_variability, tab_forecast, tab_delivery, tab_outsource = st.tabs([
    "📊 Overview",
    "🟢 FG Status",
    "🏭 MPS / FG Planning",
    "📦 MRP / Raw Materials",
    "🗄️ RM Registry",
    "⚙️ Capacity",
    "📋 WO Log",
    "📉 Demand Variability",
    "📈 Demand Forecast",
    "🚚 Delivery",
    "🏭 Outsource",
])


# ══════════════════════════════════════════════════════════════
# TAB 1 — OVERVIEW
# ══════════════════════════════════════════════════════════════
with tab_overview:
    st.subheader("Key Performance Indicators")
    total_dem  = sum(ev["qty"] for ev in get_demand_events())
    sim_df_ov  = r["sim_df"]
    rop_alerts = int(sim_df_ov["rop_hit"].sum()) if not sim_df_ov.empty else 0
    ss_breaches = int(sim_df_ov["below_ss"].sum()) if not sim_df_ov.empty else 0

    col1, col2, col3, col4, col5, col6 = st.columns(6)
    with col1:
        kpi_card("Current FG", f"{raw_data['current_fg']:,}",
                 delta=raw_data['current_fg'] - raw_data['target_fg'],
                 help_text="vs. MTS Target", color="#2563EB")
    with col2:
        kpi_card("MTS Target", f"{raw_data['target_fg']:,}",
                 help_text="Desired EOM stock", color="#10B981")
    with col3:
        kpi_card("Reorder Point", f"{r['rop']:,}",
                 help_text=f"MTS − MOQ ({raw_data['prod_moq']:,})",
                 color="#EF4444" if raw_data["current_fg"] <= r["rop"] else "#10B981")
    with col4:
        kpi_card("Fixed Safety Stock", f"{r['safety_stock']:,}",
                 help_text="Hard floor — never consume in ideal plan",
                 color="#F59E0B")
    with col5:
        kpi_card("Total Demand (events)", f"{total_dem:,}",
                 help_text=f"{len(get_demand_events())} orders entered",
                 color="#8B5CF6")
    with col6:
        kpi_card("SS Breach Days (6M)", str(ss_breaches),
                 help_text="Days FG dips below safety stock",
                 color="#EF4444" if ss_breaches > 0 else "#10B981")

    st.divider()

    # ── Fill Rate & Service Level ─────────────────────────────
    st.subheader("📊 Fill Rate & Service Level")
    st.caption(
        "**Fill Rate** = fully filled orders ÷ total orders  &nbsp;|&nbsp;  "
        "**Service Level** = fulfilled quantity ÷ total demand quantity  &nbsp;|&nbsp;  "
        "Calculated from demand events entered in MPS vs. projected FG at each date."
    )

    _demand_evs = get_demand_events()
    if not _demand_evs:
        st.info("Enter demand events in the MPS tab to calculate fill rate and service level.")
    else:
        # For each demand event, check if FG at that date could fulfil it
        _sim = r["sim_df"].copy() if not r["sim_df"].empty else pd.DataFrame()

        def _fg_at(d: datetime.date) -> int:
            """Closing FG on date d from sim_df."""
            if _sim.empty: return raw_data["current_fg"]
            _row = _sim[_sim["date"] == d]
            if _row.empty:
                # Use closest prior date
                _prior = _sim[_sim["date"] <= d]
                if _prior.empty: return raw_data["current_fg"]
                return int(_prior.iloc[-1]["closing_fg"])
            return int(_row.iloc[0]["closing_fg"])

        _total_orders    = len(_demand_evs)
        _total_qty       = sum(ev["qty"] for ev in _demand_evs)
        _filled_orders   = 0
        _fulfilled_qty   = 0
        _running_fg      = raw_data["current_fg"]
        _unfilled_rows   = []

        # Process events chronologically, simulating FG drawdown
        for ev in sorted(_demand_evs, key=lambda x: x["date"]):
            _fg_now   = _fg_at(ev["date"])
            _demanded = int(ev["qty"])
            _can_fill = min(_fg_now, _demanded)
            _fulfilled_qty += _can_fill
            if _can_fill >= _demanded:
                _filled_orders += 1
            else:
                _shortfall = _demanded - _can_fill
                _unfilled_rows.append({
                    "Date":     ev["date"].strftime("%d %b %Y") if hasattr(ev["date"], "strftime") else str(ev["date"]),
                    "Label":    ev.get("label", "—"),
                    "Demanded": f"{_demanded:,}",
                    "Fulfilled":f"{_can_fill:,}",
                    "Shortfall":f"{_shortfall:,}",
                })

        _fill_rate   = (_filled_orders / _total_orders * 100) if _total_orders else 0
        _svc_level   = (_fulfilled_qty  / _total_qty   * 100) if _total_qty    else 0

        _fr_col, _sl_col, _tq_col, _fo_col = st.columns(4)
        with _fr_col:
            kpi_card("Fill Rate",
                     f"{_fill_rate:.1f}%",
                     help_text=f"{_filled_orders}/{_total_orders} orders fully filled",
                     color="#10B981" if _fill_rate >= 95 else "#F59E0B" if _fill_rate >= 80 else "#EF4444")
        with _sl_col:
            kpi_card("Service Level",
                     f"{_svc_level:.1f}%",
                     help_text=f"{_fulfilled_qty:,}/{_total_qty:,} cartons fulfilled",
                     color="#10B981" if _svc_level >= 95 else "#F59E0B" if _svc_level >= 80 else "#EF4444")
        with _tq_col:
            kpi_card("Total Demand",
                     f"{_total_qty:,}",
                     help_text=f"{_total_orders} demand events",
                     color="#8B5CF6")
        with _fo_col:
            kpi_card("Unfilled Orders",
                     str(_total_orders - _filled_orders),
                     help_text="Orders with at least partial shortfall",
                     color="#EF4444" if (_total_orders - _filled_orders) > 0 else "#10B981")

        if _unfilled_rows:
            with st.expander(f"⚠️ Unfilled / Partial Orders ({len(_unfilled_rows)})", expanded=False):
                st.dataframe(pd.DataFrame(_unfilled_rows), hide_index=True, use_container_width=True)

    st.divider()
    col_a, col_b = st.columns([3, 1])
    with col_a:
        st.subheader("FG Inventory — Historical + 12-Month Projection")
        st.plotly_chart(fg_chart(r, get_demand_events()), use_container_width=True, key="fg_chart_overview")
    with col_b:
        st.subheader("MPS Parameters")
        st.markdown(f"""
| Parameter | Value |
|-----------|-------|
| MTS Target | {raw_data['target_fg']:,} |
| Fixed Safety Stock | {r['safety_stock']:,} |
| **Max FG (MTS + SS)** | **{r['max_fg']:,}** |
| Production LT | {raw_data['prod_lt']} days |
| Production MOQ | {raw_data['prod_moq']:,} |
| **ROP (MTS + SS − MOQ)** | **{r['rop']:,}** |
| Wastage % | {raw_data['wastage_pct']}% |
| Number of Ups | {raw_data['number_of_ups']} |
        """)

    st.divider()
    st.subheader("ROP Logic Explained")
    st.info(
        f"**Inventory Model:**  \n"
        f"- Max FG (replenish ceiling) = MTS + SS = {raw_data['target_fg']:,} + {r['safety_stock']:,} = **{r['max_fg']:,}**  \n"
        f"- ROP = MTS + SS − MOQ = {raw_data['target_fg']:,} + {r['safety_stock']:,} − {raw_data['prod_moq']:,} = **{r['rop']:,}**  \n"
        f"- WO Qty = Max FG − current FG = {r['max_fg']:,} − FG at trigger, rounded up to MOQ ({raw_data['prod_moq']:,})  \n"
        f"- Safety Stock = **{r['safety_stock']:,}** — hard floor, never consumed in ideal plan"
    )

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        sc   = "#DC2626" if raw_data["current_fg"] <= r["rop"] else "#16A34A"
        stxt = "BELOW ROP — Release WO" if raw_data["current_fg"] <= r["rop"] else "Above ROP — Monitor"
        st.markdown(f"""<div style="background:{sc}18;border:1px solid {sc}44;
        border-radius:8px;padding:14px;text-align:center">
        <div style="font-size:.8rem;color:{sc};font-weight:700">FG STATUS</div>
        <div style="font-size:1.1rem;font-weight:600;color:{sc};margin-top:4px">{stxt}</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        _reg_reel_stk = st.session_state.get("reel_registry", {}).get(raw_data["reel_name"], {}).get("stock_kg", raw_data["reel_stock"])
        reel_ok = _reg_reel_stk > r["reel_rop_kg"]
        rc = "#16A34A" if reel_ok else "#DC2626"
        rs = "OK" if reel_ok else "BELOW ROP"
        st.markdown(f"""<div style="background:{rc}18;border:1px solid {rc}44;border-radius:8px;
        padding:14px;text-align:center">
        <div style="font-size:.8rem;color:{rc};font-weight:700">REEL STATUS</div>
        <div style="font-size:1.1rem;font-weight:600;color:{rc};margin-top:4px">
        {int(_reg_reel_stk):,} KG — {rs}</div></div>""", unsafe_allow_html=True)
    with c3:
        _reg_pouch_stk = st.session_state.get("pouch_registry", {}).get(raw_data["pouch_name"], {}).get("stock_units", raw_data["pouch_stock"])
        pok = _reg_pouch_stk > r["pouch_rop"]
        pc  = "#16A34A" if pok else "#DC2626"
        ps  = "OK" if pok else "BELOW ROP"
        st.markdown(f"""<div style="background:{pc}18;border:1px solid {pc}44;border-radius:8px;
        padding:14px;text-align:center">
        <div style="font-size:.8rem;color:{pc};font-weight:700">POUCH STATUS</div>
        <div style="font-size:1.1rem;font-weight:600;color:{pc};margin-top:4px">
        {int(_reg_pouch_stk):,} — {ps}</div></div>""", unsafe_allow_html=True)
    with c4:
        if cap_res["proc_summary"]:
            bn_util = max(p["max_util"] for p in cap_res["proc_summary"])
            bn_name = cap_res["bottleneck"]
        else:
            bn_util, bn_name = 0, "—"
        bc = "#DC2626" if bn_util > 90 else "#F59E0B" if bn_util > 60 else "#16A34A"
        st.markdown(f"""<div style="background:{bc}18;border:1px solid {bc}44;border-radius:8px;
        padding:14px;text-align:center">
        <div style="font-size:.8rem;color:{bc};font-weight:700">BOTTLENECK</div>
        <div style="font-size:1.1rem;font-weight:600;color:{bc};margin-top:4px">
        {bn_name} — {bn_util:.1f}%</div></div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# TAB 2 — FG STATUS  (NEW)
# ══════════════════════════════════════════════════════════════
with tab_fg_status:
    st.subheader("🟢 FG Inventory Status — Live View")

    today      = datetime.date.today()
    sim_df_fg  = r["sim_df"]

    current_fg  = raw_data["current_fg"]
    ss          = r["safety_stock"]
    rop         = r["rop"]
    target_fg   = raw_data["target_fg"]

    # Current backorder from simulation (today's row)
    _today_row  = sim_df_fg[sim_df_fg["date"] == pd.Timestamp(today)]
    backorder_now = int(_today_row["backorder"].values[0]) if not _today_row.empty else 0

    # WIP: only WOs whose arrival date is still in the future
    issued_wos  = get_issued_wos()
    wip_qty     = sum(
        info.get("actual_produced", info.get("wo_qty", 0))
        for ds, info in issued_wos.items()
        if info.get("issued") and (
            datetime.date.fromisoformat(ds)
            + datetime.timedelta(days=info.get("actual_lt_override", raw_data["prod_lt"]))
        ) > today
    )

    # Net effective FG = current + WIP arriving − backorder obligation
    net_effective_fg = max(0, current_fg + wip_qty - backorder_now)
    headroom         = current_fg - rop

    # ── KPI cards ────────────────────────────────────────────
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    with c1:
        kpi_card("Current FG", f"{current_fg:,}", color="#2563EB")
    with c2:
        kpi_card("WIP (in production)", f"{wip_qty:,}",
                 help_text="Confirmed WOs not yet arrived", color="#7C3AED")
    with c3:
        kpi_card("Backorder", f"{backorder_now:,}",
                 help_text="Unfulfilled customer demand",
                 color="#DC2626" if backorder_now > 0 else "#16A34A")
    with c4:
        kpi_card("Net Effective FG", f"{net_effective_fg:,}",
                 help_text="FG + WIP − Backorder",
                 color="#EF4444" if net_effective_fg < ss else "#10B981")
    with c5:
        kpi_card("Safety Stock", f"{ss:,}",
                 help_text="Hard floor — never consume", color="#F59E0B")
    with c6:
        kpi_card("Gap to MTS", f"{max(0, target_fg - net_effective_fg):,}",
                 help_text=f"Target: {target_fg:,}", color="#8B5CF6")

    st.divider()

    # # ── FG Level gauge ───────────────────────────────────────
    # st.markdown("#### FG Level Gauge")
    # max_val = max(r["max_fg"] * 1.1, max(current_fg, net_effective_fg) * 1.1)
    # fig_gauge = go.Figure()
    # fig_gauge.add_trace(go.Bar(x=["FG Level"], y=[ss],
    #     name=f"Safety Stock ({ss:,})", marker_color="#FEE2E2", base=0))
    # fig_gauge.add_trace(go.Bar(x=["FG Level"], y=[target_fg - ss],
    #     name=f"Working Stock (SS→MTS)", marker_color="#FEF3C7", base=ss))
    # fig_gauge.add_trace(go.Bar(x=["FG Level"], y=[r["max_fg"] - target_fg],
    #     name=f"Buffer (MTS→Max)", marker_color="#D1FAE5", base=target_fg))
    # fig_gauge.add_hline(y=current_fg, line_color="#2563EB", line_width=3,
    #                     annotation_text=f"Current FG: {current_fg:,}",
    #                     annotation_position="right", annotation_font_color="#2563EB")
    # if wip_qty:
    #     fig_gauge.add_hline(y=min(current_fg + wip_qty, max_val), line_color="#7C3AED",
    #                         line_width=2, line_dash="dot",
    #                         annotation_text=f"After WIP arrival: {current_fg + wip_qty:,}",
    #                         annotation_position="right", annotation_font_color="#7C3AED")
    # if backorder_now:
    #     fig_gauge.add_hline(y=max(0, current_fg - backorder_now), line_color="#DC2626",
    #                         line_width=2, line_dash="dash",
    #                         annotation_text=f"After backorder: {max(0, current_fg - backorder_now):,}",
    #                         annotation_position="right", annotation_font_color="#DC2626")
    # fig_gauge.update_layout(
    #     height=220, barmode="stack", yaxis_range=[0, max_val],
    #     yaxis_title="Cartons", paper_bgcolor="rgba(0,0,0,0)",
    #     plot_bgcolor="rgba(0,0,0,0)", margin=dict(l=0, r=200, t=20, b=0),
    #     legend=dict(orientation="h", y=-0.35),
    # )
    # st.plotly_chart(fig_gauge, use_container_width=True, key="plotly_1")

    # Backorder alert
    if backorder_now > 0:
        st.error(
            f"⚠️ **Backorder: {backorder_now:,} cartons** unfulfilled. "
            f"Will be cleared when next WO arrives ({wip_qty:,} units in production)."
            if wip_qty else
            f"⚠️ **Backorder: {backorder_now:,} cartons** unfulfilled. No WO in production — issue one immediately."
        )

    st.divider()

    # ── WIP — Work Orders In-Flight ──────────────────────────
    st.markdown("#### 🔧 WIP — Confirmed Work Orders")
    wip_rows  = []   # in-flight (not yet arrived)
    hist_rows = []   # already arrived (historical)
    for date_str, info in issued_wos.items():
        if not info.get("issued"):
            continue
        try:
            rel_date = datetime.date.fromisoformat(date_str)
        except:
            continue
        actual_lt_d = info.get("actual_lt_override", raw_data["prod_lt"])
        so_wo_d     = info.get("so_to_wo_days", raw_data.get("so_to_wo_days", 1))

        # Use pre-computed arrival date if available
        if info.get("wo_arrival_date"):
            arr_date = datetime.date.fromisoformat(info["wo_arrival_date"])
        elif info.get("wo_release_date"):
            _wrel    = datetime.date.fromisoformat(info["wo_release_date"])
            arr_date = _wrel + datetime.timedelta(days=int(actual_lt_d))
        else:
            arr_date = rel_date + datetime.timedelta(days=int(so_wo_d) + int(actual_lt_d))
        days_remaining = (arr_date - today).days
        if days_remaining > 0:
            status = f"🔵 In Production — {days_remaining}d to go"
        elif days_remaining == 0:
            status = "🟢 Arriving Today"
        else:
            status = f"✅ Arrived {abs(days_remaining)}d ago"

        row_data = {
            "WO Release":     _fmt_date(rel_date),
            "Gross Qty":      f"{info.get('gross_qty', 0):,}",
            "Net Yield":      f"{info.get('actual_produced', info.get('wo_qty', 0)):,}",
            "Arrival Date":   _fmt_date(arr_date),
            "Days Remaining": max(0, days_remaining),
            "Backorder Cleared": f"{min(backorder_now, info.get('actual_produced', info.get('wo_qty', 0))):,}" if backorder_now > 0 and days_remaining >= 0 else "—",
            "Status":         status,
        }
        if arr_date > today:
            wip_rows.append(row_data)
        else:
            hist_rows.append(row_data)

    if wip_rows:
        wip_df = pd.DataFrame(wip_rows)
        st.dataframe(
            wip_df.style.map(
                lambda v: "color:#16A34A;font-weight:600;" if "Today" in str(v) or "Arrived" in str(v)
                          else "color:#2563EB;font-weight:600;" if "Production" in str(v) else "",
                subset=["Status"],
            ),
            use_container_width=True, hide_index=True,
        )
        st.markdown("#### 📅 Arrival Timeline")
        for w in sorted(wip_rows, key=lambda x: x["Arrival Date"]):
            dr   = w["Days Remaining"]
            col  = "#16A34A" if dr == 0 else "#2563EB" if dr > 0 else "#6B7280"
            bo   = f" → Clears <b>{w['Backorder Cleared']}</b> backorder" if w["Backorder Cleared"] != "—" else ""
            st.markdown(
                f"<div style='border-left:4px solid {col};padding:8px 14px;"
                f"margin-bottom:6px;background:{col}10;border-radius:4px;'>"
                f"<b>{w['Arrival Date']}</b> &nbsp;|&nbsp; Net: <b>{w['Net Yield']}</b>"
                f"{bo} &nbsp;|&nbsp; Issued: {w['WO Release']} &nbsp;|&nbsp; {w['Status']}"
                f"</div>",
                unsafe_allow_html=True,
            )
    else:
        st.info("No active Work Orders in production.")

    if hist_rows:
        with st.expander(f"📋 Historical Work Orders ({len(hist_rows)}) — already reflected in current FG", expanded=False):
            hist_df = pd.DataFrame(hist_rows).drop(columns=["Days Remaining","Backorder Cleared"], errors="ignore")
            st.dataframe(
                hist_df.style.map(
                    lambda v: "color:#6B7280;" if "Arrived" in str(v) else "",
                    subset=["Status"],
                ),
                use_container_width=True, hide_index=True,
            )
            st.caption(
                "These WOs have already arrived. Their produced quantities are included in the "
                "current FG stock for the forward simulation."
            )

    st.divider()

    # ── Forward FG projection (next 30 days) ────────────────
    st.markdown("#### 📈 FG Projection — Next 60 Days")
    if not sim_df_fg.empty:
        next30 = sim_df_fg[sim_df_fg["date"] <= today + datetime.timedelta(days=60)].copy()
        if not next30.empty:
            fig30 = go.Figure()
            fig30.add_trace(go.Scatter(
                x=next30["date"], y=next30["closing_fg"],
                mode="lines+markers", name="Projected FG",
                line=dict(color="#2563EB", width=2),
                fill="tozeroy", fillcolor="rgba(37,99,235,0.08)",
            ))
            fig30.add_hline(y=r.get("max_fg", ss + target_fg), line_color="#1D4ED8", line_dash="longdash",
                            annotation_text=f"Max FG ({r.get('max_fg', ss + target_fg):,})")
            fig30.add_hline(y=target_fg, line_color="#10B981", line_dash="dot",
                            annotation_text=f"MTS ({target_fg:,})")
            fig30.add_hline(y=ss, line_color="#F59E0B", line_dash="dash",
                            annotation_text=f"SS ({ss:,})")
            fig30.add_hline(y=rop, line_color="#EF4444", line_dash="dash",
                            annotation_text=f"ROP ({rop:,})")
            # Arrival events
            arrivals_30 = next30[next30["arrival_qty"] > 0]
            if not arrivals_30.empty:
                fig30.add_trace(go.Scatter(
                    x=arrivals_30["date"], y=arrivals_30["opening_fg"],
                    mode="markers", name="WO Arrival",
                    marker=dict(symbol="triangle-up", size=14, color="#16A34A"),
                ))
            fig30.update_layout(
                height=300, margin=dict(l=0, r=160, t=20, b=0),
                yaxis_title="Cartons",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                legend=dict(orientation="h", y=-0.2),
            )
            st.plotly_chart(fig30, use_container_width=True, key="plotly_2")

    # ── Demand events in next 30 days ───────────────────────
    st.markdown("#### 📦 Demand Events — Next 60 Days")
    upcoming = [ev for ev in get_demand_events()
                if today <= ev["date"] <= today + datetime.timedelta(days=60)]
    if upcoming:
        upcoming_df = pd.DataFrame([{
            "Date":    _fmt_date(ev["date"]),
            "Qty":     f"{ev['qty']:,}",
            "Label":   ev.get("label", ""),
            "FG After": f"{max(0, next((row['closing_fg'] for _, row in sim_df_fg.iterrows() if row['date'] == ev['date']), 0)):,}",
        } for ev in sorted(upcoming, key=lambda x: x["date"])])
        st.dataframe(upcoming_df, hide_index=True, use_container_width=True)
    else:
        st.info("No demand events in the next 60 days.")


# ══════════════════════════════════════════════════════════════
# TAB 3 — MPS / FG PLANNING
# ══════════════════════════════════════════════════════════════
with tab_mps:
    st.subheader("🏭 MPS / FG Planning — 6-Month Daily View")

    # ── In-Transit WO Banner ──────────────────────────────────
    # Source 1: WO Log in-transit entries
    _in_transit_wos_log = [wo for wo in raw_data.get("wo_log", []) if wo.get("in_transit")]
    # Source 2: Dashboard-confirmed WOs whose arrival is still in the future
    _today_mps = datetime.date.today()
    _in_transit_wos_dashboard = []
    for _ds_it, _info_it in get_issued_wos().items():
        if not _info_it.get("issued") or _info_it.get("from_wo_log"):
            continue
        try:
            # Use stored arrival date if available
            if _info_it.get("wo_arrival_date"):
                _arr_it = datetime.date.fromisoformat(_info_it["wo_arrival_date"])
                _rel_it = datetime.date.fromisoformat(_info_it.get("wo_release_date", _ds_it))
            else:
                _rel_it = datetime.date.fromisoformat(_ds_it)
                _lt_it  = int(_info_it.get("actual_lt_override", raw_data.get("prod_lt", 10)))
                _so_it  = int(_info_it.get("so_to_wo_days", raw_data.get("so_to_wo_days", 1)))
                _rel_it = _rel_it + datetime.timedelta(days=_so_it)
                _arr_it = _rel_it + datetime.timedelta(days=_lt_it)
            if _arr_it > _today_mps:
                _in_transit_wos_dashboard.append({
                    "wo_num":     _info_it.get("wo_num", f"WO {_ds_it}"),
                    "so_num":     _info_it.get("so_num", "—"),
                    "release":    _rel_it,
                    "arrival":    _arr_it,
                    "target_qty": _info_it.get("wo_qty", 0),
                    "source":     "dashboard",
                })
        except Exception:
            continue

    # Display banners
    for _itwo in _in_transit_wos_log:
        try:
            _rel = pd.to_datetime(_itwo.get("release", ""), dayfirst=True, errors="coerce")
            _est = pd.to_datetime(_itwo.get("est_complete", ""), dayfirst=True, errors="coerce")
            if isinstance(_itwo.get("release"), (datetime.date, datetime.datetime)):
                _rel = pd.Timestamp(_itwo["release"])
            if isinstance(_itwo.get("est_complete"), (datetime.date, datetime.datetime)):
                _est = pd.Timestamp(_itwo["est_complete"])
            _rel_str = _rel.strftime("%d %b %Y") if pd.notna(_rel) else "—"
            _est_str = _est.strftime("%d %b %Y") if pd.notna(_est) else "—"
            _tgt     = int(_itwo.get("target_qty", 0))
        except Exception:
            _rel_str = _est_str = "—"; _tgt = 0
        st.warning(
            f"🚛 **Work Order In Production** (WO# {_itwo.get('wo_num','—')} · SO# {_itwo.get('so_num','—')})  \n"
            f"Released: **{_rel_str}** &nbsp;|&nbsp; "
            f"Estimated Completion: **{_est_str}** &nbsp;|&nbsp; "
            f"Target Qty: **{_tgt:,} cartons**  \n"
            f"This WO is currently in production. FG and MPS table include this arrival."
        )

    for _itwo in _in_transit_wos_dashboard:
        st.info(
            f"🏭 **Work Order Confirmed & In Production**  \n"
            f"Released: **{_fmt_date(_itwo['release'])}** &nbsp;|&nbsp; "
            f"Expected Arrival: **{_fmt_date(_itwo['arrival'])}** &nbsp;|&nbsp; "
            f"Net Yield: **{_itwo['target_qty']:,} cartons**  \n"
            f"This WO is reflected in the FG projection chart and MPS table."
        )

    st.info(
        f"**Inventory Model:**  \n"
        f"- Max FG = MTS + SS = {raw_data['target_fg']:,} + {r['safety_stock']:,} = **{r['max_fg']:,}** (replenish ceiling)  \n"
        f"- ROP = MTS + SS − MOQ = {raw_data['target_fg']:,} + {r['safety_stock']:,} − {raw_data['prod_moq']:,} = **{r['rop']:,}**  \n"
        f"- WO Qty = Max FG − FG at trigger, rounded up to MOQ ({raw_data['prod_moq']:,}) multiple  \n"
        f"- SS = **{r['safety_stock']:,}** — hard floor, never consumed  \n"
        f"- Prod LT Variability σ: **{r.get('lt_var_used', raw_data['lt_variability'])} days** "
        f"*(source: {r.get('lt_var_source', 'Excel')})*"
    )

    # ── Demand Event Input ────────────────────────────────────
    st.markdown("#### ➕ Enter Demand Events")

    # Available FG = current stock minus all previously entered demand events
    current_fg_stk  = raw_data.get("current_fg", 0)

    # Use the demand date selected in the input to find projected FG on that day
    _inp_date_for_cap = st.session_state.get("inp_date",
                         datetime.date.today() + datetime.timedelta(days=7))

    # Projected FG on the demand date from the simulation (includes WO arrivals and prior demand events)
    _sim = r["sim_df"]
    if not _sim.empty:
        _sim_row = _sim[_sim["date"] == pd.Timestamp(_inp_date_for_cap)]
        projected_fg = int(_sim_row["closing_fg"].values[0]) if not _sim_row.empty else current_fg_stk
    else:
        projected_fg = current_fg_stk

    # Subtract demand events already committed ON OR AFTER the demand date
    # (they also draw from this projected pool)
    committed_on_or_after = sum(
        ev["qty"] for ev in get_demand_events()
        if ev["date"] >= _inp_date_for_cap
    )
    available_fg = max(0, projected_fg - committed_on_or_after)

    st.caption(
        f"Current FG: **{current_fg_stk:,}**  |  "
        f"Projected on {_inp_date_for_cap.strftime('%d %b')}: **{projected_fg:,}**  |  "
        f"Already committed on/after that date: **{committed_on_or_after:,}**  |  "
        f"Indicative available: **{available_fg:,}** *(demand exceeding this will show as backorder)*"
    )

    with st.container():
        col_d, col_q, col_lbl, col_btn = st.columns([2, 2, 3, 1])
        with col_d:
            inp_date = st.date_input(
                "Delivery / Consumption Date",
                value=datetime.date.today() + datetime.timedelta(days=7),
                min_value=datetime.date.today(),
                max_value=datetime.date.today() + datetime.timedelta(days=183),
                key="inp_date",
            )
        with col_q:
            inp_qty = st.number_input(
                "Quantity (cartons)",
                min_value=1,
                max_value=2_000_000,
                value=5000,
                step=500, key="inp_qty",
            )
        with col_lbl:
            inp_label = st.text_input(
                "Label / PO reference (optional)", value="", key="inp_label",
            )
        with col_btn:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Add ➕", use_container_width=True, type="primary"):
                get_demand_events().append({
                    "date":  inp_date,
                    "qty":   int(inp_qty),
                    "label": inp_label.strip(),
                })
                st.rerun()

    if get_demand_events():
        st.markdown("**Entered demand events:**")
        ev_display = []
        for i, ev in enumerate(get_demand_events()):
            ev_display.append({
                "#":             i + 1,
                "Date":          ev["date"].strftime("%d %b %Y"),
                "Qty (cartons)": f"{ev['qty']:,}",
                "Label":         ev["label"],
            })
        st.dataframe(pd.DataFrame(ev_display), hide_index=True, use_container_width=True)

        col_del1, col_del2, col_clr = st.columns([1, 1, 4])
        with col_del1:
            del_idx = st.number_input("Remove row #", 1,
                                      len(get_demand_events()), 1, key="del_idx")
        with col_del2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑 Remove", use_container_width=True):
                get_demand_events().pop(int(del_idx) - 1)
                st.rerun()
        with col_clr:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑 Clear all demand events", use_container_width=True):
                st.session_state[_k_demand] = []
                st.rerun()
    else:
        st.info("No demand events yet. Add your first order above.")

    st.divider()
    st.plotly_chart(fg_chart(r, get_demand_events()), use_container_width=True, key="fg_chart_mps")
    st.divider()

    # ── ROP Alert / WO Confirmation UI ───────────────────────
    if not r["sim_df"].empty:
        rop_rows = r["sim_df"][r["sim_df"]["rop_hit"] & ~r["sim_df"]["wo_issued"]]

        if not rop_rows.empty:
            st.warning(
                f"⚠️ **{len(rop_rows)} ROP alert(s)** — FG drops to/below ROP ({r['rop']:,}).  "
                "Select an alert below to review and confirm WO issuance."
            )

            alert_options = {}
            for _, row in rop_rows.iterrows():
                ds    = _safe_date_str(row["date"])
                d_fmt = _fmt_date(row["date"])
                label = (
                    f"{d_fmt}  |  FG: {row['closing_fg']:,}  →  ROP: {r['rop']:,}  |  "
                    f"Rec. Net: {row['wo_net_qty']:,} cartons (MOQ: {raw_data['prod_moq']:,})"
                )
                alert_options[label] = (ds, row)

            selected_alert_label = st.selectbox(
                "🔽 Select alert to action",
                options=list(alert_options.keys()),
                key="alert_selector",
            )

            ds, row = alert_options[selected_alert_label]
            d_fmt   = _fmt_date(row["date"])
            arr_dt  = row["wo_arrival_date"]
            arr_fmt = _fmt_date(arr_dt)

            with st.container():
                st.markdown(
                    f"<div style='border:1px solid #FCA5A5;border-left:4px solid #EF4444;"
                    f"border-radius:8px;padding:12px 16px;background:#FEF2F2;margin-bottom:12px;"
                    f"color:black;'>"
                    f"<b>🔴 Alert — {d_fmt}</b>  |  FG will be {row['closing_fg']:,} "
                    f"(ROP = MTS+SS−MOQ = {r['rop']:,} | Replenish to Max FG = {r['max_fg']:,})</div>",
                    unsafe_allow_html=True,
                )
                ca, cb, cc = st.columns(3)
                with ca:
                    st.metric("Gross WO Qty (to issue)", f"{row['wo_gross_qty']:,}")
                    st.metric("Net FG Yield (after waste)", f"{row['wo_net_qty']:,}")
                    st.metric("MOQ Multiples", f"×{int(row['wo_net_qty'] / raw_data['prod_moq'])}")
                with cb:
                    st.metric("Default Arrival Date", arr_fmt)
                    fg_post = int(row["closing_fg"]) + int(row["wo_net_qty"])
                    st.metric("Projected FG After Arrival", f"{fg_post:,}")
                with cc:
                    actual_p = st.number_input(
                        "Actual qty to produce (cartons)",
                        min_value=0, max_value=500000,
                        value=int(row["wo_net_qty"]),
                        key=f"ap_{ds}",
                    )
                    actual_lt_override = st.number_input(
                        "Actual Production LT (days)",
                        min_value=1, max_value=120,
                        value=raw_data["prod_lt"],
                        help="Override lead time for this WO (excl. SO→WO days)",
                        key=f"lt_{ds}",
                    )
                    so_to_wo_input = st.number_input(
                        "SO → WO days",
                        min_value=0, max_value=30,
                        value=int(raw_data.get("so_to_wo_days", 1)),
                        help="Days between SO placement and WO release. Default read from SKU Master col O.",
                        key=f"so_wo_{ds}",
                    )
                    # Total LT = SO→WO + production LT
                    total_lt_input = so_to_wo_input + actual_lt_override
                    _std_total     = raw_data.get("so_to_wo_days", 1) + raw_data["prod_lt"]
                    if total_lt_input != _std_total:
                        st.caption(
                            f"⚠️ Total LT: **{total_lt_input}d** "
                            f"(SO→WO: {so_to_wo_input}d + Prod: {actual_lt_override}d) "
                            f"vs standard {_std_total}d"
                        )
                    # WO release date = alert date + SO→WO days
                    try:
                        _alert_date     = datetime.date.fromisoformat(ds)
                        wo_release_date = _alert_date + datetime.timedelta(days=so_to_wo_input)
                        _arr_date_calc  = wo_release_date + datetime.timedelta(days=actual_lt_override)
                    except Exception:
                        wo_release_date = datetime.date.today()
                        _arr_date_calc  = wo_release_date + datetime.timedelta(days=actual_lt_override)

                    if so_to_wo_input > 0:
                        st.caption(
                            f"📋 SO: {_alert_date.strftime('%d %b')} → "
                            f"WO release: **{wo_release_date.strftime('%d %b')}** → "
                            f"Arrival: **{_arr_date_calc.strftime('%d %b %Y')}**"
                        )

                    rm_check = check_rm_availability(
                        raw          = raw_data,
                        wo_qty       = int(actual_p),
                        wo_date      = wo_release_date,
                        reel_registry        = st.session_state.get("reel_registry", {}),
                        reel_pos_released    = st.session_state.get("reel_pos_released", {}),
                        already_issued_wos   = get_issued_wos(),
                        reel_actual_arrivals = st.session_state.get("reel_actual_arrivals", {}),
                        all_issued_wos_map   = all_issued_wos_map,
                        all_raw_map          = all_results_raw,
                    )

                    # Show RM status table
                    rm_rows = []
                    for rc in rm_check["components"]:
                        _rid  = (rc.get("rm_material_id") or "").strip()
                        _desc = (rc.get("rm_material_desc") or rc["reel_name"]).strip()
                        _rm_display = f"{_rid} — {_desc}" if _rid else _desc
                        rm_rows.append({
                            "RM":             _rm_display[:60],
                            "Required (KG)":  f"{rc['required_kg']:,}",
                            "Available (KG)": f"{rc['available_kg']:,}",
                            "Shortfall (KG)": f"{rc['shortfall_kg']:,}" if rc["shortfall_kg"] else "—",
                            "Status":         rc["status"],
                        })
                    rm_df = pd.DataFrame(rm_rows)

                    if rm_check["stock_ok"]:
                        st.success("✅ Sufficient RM available — WO can be confirmed today.")
                    elif rm_check["po_covered"]:
                        st.warning(
                            f"🚚 RM not yet available — PO in production. "
                            f"Earliest WO date: **{rm_check['earliest_ok_date'].strftime('%d %b %Y')}**"
                        )
                    else:
                        st.error("🔴 Insufficient RM — place RM purchase order before confirming this WO.")

                    st.dataframe(
                        rm_df.style.map(
                            lambda v: "color:#16A34A;font-weight:600;" if "✅" in str(v) else
                                      "color:#D97706;font-weight:600;" if "🚚" in str(v) else
                                      "color:#DC2626;font-weight:600;" if "🔴" in str(v) else "",
                            subset=["Status"],
                        ),
                        hide_index=True, use_container_width=True,
                    )

                    # Block / allow confirm button
                    can_confirm = rm_check["stock_ok"]
                    po_covered  = rm_check["po_covered"]
                    earliest    = rm_check["earliest_ok_date"]

                    if can_confirm:
                        btn_label    = f"✅ Confirm WO — {d_fmt}"
                        btn_disabled = False
                        force_override = False
                    elif po_covered and earliest:
                        st.info(f"🚚 RM in production — earliest ready: **{earliest.strftime('%d %b %Y')}**")
                        force_override = st.checkbox(
                            "⚠️ Force-confirm now (RM PO in production — confirm at your own risk)",
                            key=f"force_{ds}",
                        )
                        btn_label    = f"⚠️ Force Confirm WO — {d_fmt}" if force_override else f"🔒 Waiting for RM ({earliest.strftime('%d %b')})"
                        btn_disabled = not force_override
                    else:
                        st.warning("🔴 Insufficient RM stock and no PO in production. Consider placing an RM order first.")
                        force_override = st.checkbox(
                            "⚠️ Force-confirm anyway (I accept the RM shortage risk)",
                            key=f"force_{ds}",
                        )
                        btn_label    = f"⚠️ Force Confirm WO — {d_fmt}" if force_override else "🔒 RM short — tick box above to override"
                        btn_disabled = not force_override

                    if st.button(
                        btn_label,
                        key=f"issue_{ds}",
                        use_container_width=True,
                        type="primary",
                        disabled=btn_disabled,
                    ):
                        get_issued_wos()[ds] = {
                            "issued":              True,
                            "wo_qty":              int(row["wo_net_qty"]),
                            "gross_qty":           int(row["wo_gross_qty"]),
                            "actual_produced":     int(actual_p),
                            "actual_lt_override":  int(actual_lt_override),
                            "so_to_wo_days":       int(so_to_wo_input),
                            "wo_release_date":     wo_release_date.isoformat(),
                            "wo_arrival_date":     _arr_date_calc.isoformat(),
                            "rm_override":         not can_confirm,
                        }
                        st.rerun()
        else:
            st.success("✅ No open ROP alerts — all recommended WOs have been confirmed.")

        # ── Confirmed WOs panel ──────────────────────────────
        confirmed_wos = r["sim_df"][r["sim_df"]["wo_issued"]]
        if not confirmed_wos.empty:
            with st.expander("✅ Confirmed Work Orders", expanded=False):
                for _, crow in confirmed_wos.iterrows():
                    cds     = _safe_date_str(crow["date"])
                    c_d_fmt = _fmt_date(crow["date"])
                    c_arr   = _fmt_date(crow["wo_arrival_date"])
                    gq      = int(crow["wo_gross_qty"]) if crow["wo_gross_qty"] else 0
                    nq      = int(crow["wo_net_qty"])   if crow["wo_net_qty"]   else 0
                    ap      = int(crow["actual_produced"]) if crow["actual_produced"] else 0
                    # Actual LT from session state
                    wo_info  = get_issued_wos().get(cds, {})
                    act_lt   = wo_info.get("actual_lt_override", raw_data["prod_lt"])
                    so_wo    = wo_info.get("so_to_wo_days", raw_data.get("so_to_wo_days", 1))
                    total_lt = act_lt + so_wo
                    std_lt   = raw_data["prod_lt"] + raw_data.get("so_to_wo_days", 1)
                    lt_var   = total_lt - std_lt
                    lt_badge = (f"Prod LT: {act_lt}d + SO→WO: {so_wo}d = {total_lt}d total"
                                + (f" ({lt_var:+d}d vs standard)" if lt_var != 0 else " (on standard)"))
                    # Use saved arrival date
                    c_arr = _fmt_date(
                        datetime.date.fromisoformat(wo_info["wo_arrival_date"])
                        if wo_info.get("wo_arrival_date") else crow["wo_arrival_date"]
                    )

                    col_info, col_rm = st.columns([5, 1])
                    with col_info:
                        st.markdown(
                            f"**{c_d_fmt}** &nbsp;|&nbsp; "
                            f"Gross: `{gq:,}` &nbsp;|&nbsp; "
                            f"Net: `{nq:,}` &nbsp;|&nbsp; "
                            f"Actual: `{ap:,}` &nbsp;|&nbsp; "
                            f"Arrival: `{c_arr}` &nbsp;|&nbsp; "
                            f"`{lt_badge}`"
                        )
                    with col_rm:
                        if st.button("🗑 Remove", key=f"rm_wo_{cds}", use_container_width=True):
                            if cds in get_issued_wos():
                                del st.session_state[_k_issued][cds]
                            st.rerun()

    st.divider()

    # ── Day-wise Detail Table ─────────────────────────────────
    st.subheader("📅 Day-wise Detail — Full Timeline (Historical + 12 Months)")
    st.markdown(
        "🔴 Red = ROP alert &nbsp;|&nbsp; 🟠 Orange = SS breach &nbsp;|&nbsp; "
        "🟢 Green = Month-end &nbsp;|&nbsp; 🔵 Blue = WO arrival &nbsp;|&nbsp; "
        "🟡 Yellow = WO confirmed &nbsp;|&nbsp; ⬛ Grey = Historical period"
    )

    if r["sim_df"].empty:
        st.info("Add demand events above to generate the daily plan.")
    else:
        df_daily = r["sim_df"].copy()
        df_daily["Date"]            = df_daily["date"].apply(_fmt_date)
        df_daily["Month"]           = df_daily["month"]
        df_daily["Period"]          = df_daily["is_historical"].map({True: "📋 Historical", False: "🔮 Forecast"})
        df_daily["Opening FG"]      = df_daily["opening_fg"].map("{:,}".format)
        df_daily["WO Arrived"]      = df_daily["arrival_qty"].apply(
                                          lambda v: f"{int(v):,}" if v > 0 else "—")
        df_daily["Demand Consumed"] = df_daily["demand"].apply(
                                          lambda v: f"{int(v):,}" if v > 0 else "—")
        df_daily["Closing FG"]      = df_daily["closing_fg"].map("{:,}".format)
        df_daily["Backorder"]       = df_daily["backorder"].apply(
                                          lambda v: f"⚠️ {int(v):,}" if v > 0 else "—")
        df_daily["Below SS?"]       = df_daily["below_ss"].map({True: "⚠️ YES", False: ""})
        df_daily["ROP Hit?"]        = df_daily["rop_hit"].map({True: "🔴 YES", False: ""})
        df_daily["Month End"]       = df_daily["is_month_end"].map({True: "📅 EOM", False: ""})
        df_daily["WO Gross Qty"]    = df_daily["wo_gross_qty"].apply(
                                          lambda v: f"{int(v):,}" if v and v > 0 else "—")
        df_daily["WO Net Yield"]    = df_daily["wo_net_qty"].apply(
                                          lambda v: f"{int(v):,}" if v and v > 0 else "—")
        df_daily["WO Arrival Date"] = df_daily["wo_arrival_date"].apply(_fmt_date)
        df_daily["WO Confirmed?"]   = df_daily["wo_issued"].map(
                                          {True: "✅ Confirmed", False: ""})

        display_cols = [
            "Date", "Month", "Period", "Opening FG", "WO Arrived", "Demand Consumed",
            "Closing FG", "Backorder", "Below SS?", "ROP Hit?", "Month End",
             "WO Gross Qty", "WO Net Yield", "WO Arrival Date", "WO Confirmed?",
        ]
        out_df = df_daily[display_cols]

        def row_style(row):
            if row["Period"] == "📋 Historical":
                if row["WO Arrived"] != "—":
                    return ["background-color:#ECFDF5;color:#374151;"] * len(row)
                if row["Demand Consumed"] != "—":
                    return ["background-color:#FFFBEB;color:#374151;"] * len(row)
                return ["background-color:#F9FAFB;color:#6B7280;"] * len(row)
            if row["Backorder"] != "—":
                return ["background-color:#FEE2E2;color:#991B1B;font-weight:600;"] * len(row)
            if row["Below SS?"] == "⚠️ YES":
                return ["background-color:#FFF7ED;color:black;font-weight:600;"] * len(row)
            if row["ROP Hit?"] == "🔴 YES":
                return ["background-color:#FEF2F2;color:black;font-weight:600;"] * len(row)
            if row["Month End"] == "📅 EOM":
                return ["background-color:#F0FDF4;color:black;"] * len(row)
            if row["WO Arrived"] != "—":
                return ["background-color:#EFF6FF;color:black;"] * len(row)
            if row["WO Confirmed?"] == "✅ Confirmed":
                return ["background-color:#FFFBEB;color:black;"] * len(row)
            return [""] * len(row)

        st.dataframe(
            out_df.style.apply(row_style, axis=1),
            use_container_width=True,
            hide_index=True,
            height=600,
        )

        # ── Month-End Summary ─────────────────────────────────
        st.divider()
        st.subheader("📊 Month-End Summary")

        month_rows = []
        for lbl, ms, me in r["months_info"]:
            month_sim = r["sim_df"][
                (r["sim_df"]["date"] >= ms) & (r["sim_df"]["date"] <= me)
            ]
            dem_in_month = sum(
                ev["qty"] for ev in get_demand_events()
                if ms <= ev["date"] <= me
            )
            proj_eom  = int(month_sim["projected_fg_eom"].iloc[-1]) \
                        if not month_sim.empty else 0
            shortfall = max(0, raw_data["target_fg"] - proj_eom)
            ss_breaches_m = int(month_sim["below_ss"].sum()) if not month_sim.empty else 0

            all_wos_this_month = []
            if not month_sim.empty:
                issued_rows = month_sim[month_sim["wo_issued"]]
                for _, irow in issued_rows.iterrows():
                    gq = irow["wo_gross_qty"]
                    nq = irow["wo_net_qty"]
                    wo_info = get_issued_wos().get(_safe_date_str(irow["date"]), {})
                    act_lt  = wo_info.get("actual_lt_override", raw_data["prod_lt"])
                    all_wos_this_month.append({
                        "release":  _fmt_date(irow["date"]),
                        "gross":    f"{int(gq):,}" if gq and gq > 0 else "—",
                        "net":      f"{int(nq):,}" if nq and nq > 0 else "—",
                        "arrival":  _fmt_date(irow["wo_arrival_date"]),
                        "act_lt":   f"{act_lt}d",
                        "status":   "✅ Issued",
                    })
                rop_rows_m = month_sim[month_sim["rop_hit"] & ~month_sim["wo_issued"]]
                for _, rrow in rop_rows_m.iterrows():
                    gq = rrow["wo_gross_qty"]
                    nq = rrow["wo_net_qty"]
                    all_wos_this_month.append({
                        "release":  _fmt_date(rrow["date"]),
                        "gross":    f"{int(gq):,} (rec.)" if gq and gq > 0 else "—",
                        "net":      f"{int(nq):,} (rec.)" if nq and nq > 0 else "—",
                        "arrival":  _fmt_date(rrow["wo_arrival_date"]),
                        "act_lt":   f"{raw_data['prod_lt']}d (std)",
                        "status":   "🔴 Needed",
                    })

            n_issued = sum(1 for w in all_wos_this_month if w["status"] == "✅ Issued")
            n_needed = sum(1 for w in all_wos_this_month if w["status"] == "🔴 Needed")
            if n_issued and n_needed:
                wo_status = f"✅ {n_issued} / 🔴 {n_needed}"
            elif n_issued:
                wo_status = f"✅ {n_issued} Issued"
            elif n_needed:
                wo_status = f"🔴 {n_needed} Needed"
            else:
                wo_status = "—"

            wo_rel = (me - datetime.timedelta(days=raw_data["prod_lt"])).strftime("%d %b %Y")

            month_rows.append({
                "Month":              lbl,
                "Total Demand":       f"{dem_in_month:,}",
                "Target FG (EOM)":   f"{raw_data['target_fg']:,}",
                "Projected FG (EOM)":f"{proj_eom:,}",
                "Shortfall":          f"{shortfall:,}" if shortfall else "—",
                "SS Breach Days":     str(ss_breaches_m) if ss_breaches_m else "—",
                "# WOs":              str(len(all_wos_this_month)) if all_wos_this_month else "—",
                "WO Release By":      wo_rel,
                "WO Status":          wo_status,
                "_wos":               all_wos_this_month,
            })

        summary_display = pd.DataFrame(month_rows).drop(columns=["_wos"])
        st.dataframe(
            summary_display.style.map(
                lambda v: "color:#DC2626;font-weight:600;" if "Needed" in str(v) else
                          "color:#16A34A;font-weight:600;" if "Issued" in str(v) else "",
                subset=["WO Status"]
            ),
            hide_index=True, use_container_width=True,
        )

        st.markdown("#### 📋 Work Order Detail by Month")
        for mrow in month_rows:
            if mrow["_wos"]:
                with st.expander(
                    f"**{mrow['Month']}** — {len(mrow['_wos'])} WO(s)  |  {mrow['WO Status']}",
                    expanded=False,
                ):
                    wo_detail_df = pd.DataFrame(mrow["_wos"])
                    wo_detail_df.columns = [
                        "WO Release Date", "Gross Qty", "Net Yield",
                        "Arrival Date", "Prod LT", "Status"
                    ]
                    st.dataframe(
                        wo_detail_df.style.map(
                            lambda v: "color:#DC2626;font-weight:600;" if v == "🔴 Needed" else
                                      "color:#16A34A;font-weight:600;" if v == "✅ Issued" else "",
                            subset=["Status"],
                        ),
                        hide_index=True, use_container_width=True,
                    )
            else:
                st.markdown(f"**{mrow['Month']}** — No WOs planned")


# ══════════════════════════════════════════════════════════════
# TAB 4 — MRP
# ══════════════════════════════════════════════════════════════
with tab_mrp:
    st.subheader("📦 Raw Material Planning — Weekly View")
    st.markdown(
        "RM requirements are derived from confirmed Work Orders and pending ROP-triggered WOs.  "
        "**📅 Order By** = the latest date by which a PO must be placed so stock arrives before that week's requirement.  "
        "🔴 Reorder status means opening stock for that week is at or below ROP."
    )



    # ── RM Safety Stock & ROP Summary (filtered to current SKU) ──
    st.markdown("#### 📋 Raw Materials for this SKU — Global Safety Stock & ROP")
    st.caption(
        "SS and ROP account for **all SKUs sharing the same RM** (by Raw Material ID). "
        "D_combined = sum of daily KG across all sharing SKUs.  \n"
        "Formula: **SS = Z × √(LT × σ_demand² + D_combined² × σ_LT²)** | "
        "**ROP = D_combined × LT + SS**"
    )

    # Get the rm_material_ids used by the current SKU
    _cur_sku_rm_ids = set()
    for _c in raw_data.get("bom_components", []):
        _rid = (_c.get("rm_material_id") or _c.get("reel_name", "")).strip()
        if _rid:
            _cur_sku_rm_ids.add(_rid)

    _global_ss_rows = []
    for rm_id in sorted(_cur_sku_rm_ids):
        info = global_rm_ss.get(rm_id)
        if not info:
            continue
        _reel_nm   = info["reel_name"]
        _stk_raw   = st.session_state.get("reel_registry", {}).get(rm_id, st.session_state.get("reel_registry", {}).get(info["reel_name"], {})).get("stock_kg", 0)
        _fut_cons  = 0
        for _sl2, _wos2 in all_issued_wos_map.items():
            _raw2 = all_results_raw.get(_sl2, {})
            for _ds2, _inf2 in _wos2.items():
                if not _inf2.get("issued"):
                    continue
                try:
                    _rel2 = datetime.date.fromisoformat(_ds2)
                    _lt2  = _inf2.get("actual_lt_override", _raw2.get("prod_lt", 10))
                    _so2  = _inf2.get("so_to_wo_days", _raw2.get("so_to_wo_days", 1))
                    if _inf2.get("wo_arrival_date"):
                        _arr2 = datetime.date.fromisoformat(_inf2["wo_arrival_date"])
                    elif _inf2.get("wo_release_date"):
                        _arr2 = datetime.date.fromisoformat(_inf2["wo_release_date"]) + datetime.timedelta(days=int(_lt2))
                    else:
                        _arr2 = _rel2 + datetime.timedelta(days=int(_so2) + int(_lt2))
                except Exception:
                    continue
                if _arr2 <= datetime.date.today():
                    continue
                _gq2 = _inf2.get("gross_qty", _inf2.get("wo_qty", 0))
                for _c2 in _raw2.get("bom_components", []):
                    if (_c2.get("rm_material_id") or _c2.get("reel_name", "")) != rm_id:
                        continue
                    _sh2 = math.ceil(_gq2 / _c2["ups"]) if _c2.get("ups") else 0
                    _kg2 = math.ceil(_sh2 * _c2["width"] * _c2["length"] * _c2["gsm"] / 1_000_000_000) if _sh2 else 0
                    _fut_cons += _kg2
        _stk_net = max(0, _stk_raw - _fut_cons)
        _ss      = info["ss_kg"]
        _rop     = info["rop_kg"]
        _status  = "🔴 Below ROP" if _stk_net <= _rop else ("🟡 OK" if _stk_net < _rop * 1.5 else "✅ Sufficient")
        # Show sharing partners (exclude current SKU)
        _sharing = [s for s in info["sku_list"] if s != sku_label]
        _sharing_str = ", ".join(_sharing) if _sharing else "Only this SKU"
        _global_ss_rows.append({
            "RM Material ID":      rm_id,
            "Description":         (info["reel_name"] or info["rm_desc"] or rm_id)[:40],
            "Also used by":        _sharing_str,
            "LT (days)":           info["lt_days"],
            "σ LT":                info["lt_var"],
            "D combined (KG/d)":   f"{info['daily_kg']:.1f}",
            "MOQ (KG)":            f"{info['moq_kg']:,}",
            "Registry Stock (KG)": f"{int(_stk_raw):,}",
            "WO Committed (KG)":   f"{int(_fut_cons):,}",
            "Net Available (KG)":  f"{int(_stk_net):,}",
            "Safety Stock (KG)":   f"{int(_ss):,}",
            "ROP (KG)":            f"{int(_rop):,}",
            "Status":              _status,
        })

    if _global_ss_rows:
        _gss_df = pd.DataFrame(_global_ss_rows)
        st.dataframe(
            _gss_df.style.map(
                lambda v: "color:#DC2626;font-weight:600;" if "Below ROP" in str(v) else
                          "color:#16A34A;font-weight:600;" if "Sufficient" in str(v) else
                          "color:#D97706;font-weight:600;" if str(v).startswith("🟡") else "",
                subset=["Status"],
            ),
            hide_index=True, use_container_width=True,
        )
    st.divider()

    # ── Prominent Order By summary across both RMs ────────────
    today_mrp = datetime.date.today()
    reel_weekly_preview = build_weekly_mrp(raw_data, r, get_issued_wos(),
                                           rm_type="reel",
                                           po_released=get_po_released().get("reel", {}))
    pouch_weekly_preview = build_weekly_mrp(raw_data, r, get_issued_wos(),
                                            rm_type="pouch",
                                            po_released=get_po_released().get("pouch", {}))

    def _next_order_by(df):
        # Only consider rows with actual shortfall (Order By is not "—")
        actionable = df[df["📅 Order By"] != "—"]
        overdue  = actionable[actionable["📅 Order By"].str.contains("overdue", case=False, na=False)]
        urgent   = actionable[~actionable["📅 Order By"].str.contains("overdue", case=False, na=False)]
        if not overdue.empty:
            return "⚠️ OVERDUE", "#DC2626"
        if not urgent.empty:
            return urgent.iloc[0]["📅 Order By"], "#D97706"
        return "✅ No reorder needed", "#16A34A"

    reel_ob, reel_ob_color   = _next_order_by(reel_weekly_preview)
    pouch_ob, pouch_ob_color = _next_order_by(pouch_weekly_preview)

    # ── Order By cards — one per unique rm_material_id in the BOM ──
    # Using rm_material_id ensures Lid/Strip/Wall (same reel_name) each get a card
    _bom_rm_ids_seen: set = set()
    _ob_cards = []
    for _c in raw_data.get("bom_components", []):
        _rm_id_c = (_c.get("rm_material_id") or _c["reel_name"]).strip()
        if _rm_id_c in _bom_rm_ids_seen:
            continue
        _bom_rm_ids_seen.add(_rm_id_c)
        _rn = _c["reel_name"]
        # Order By: triggered by (a) WO-driven gross req, or (b) stock already ≤ ROP
        _reel_weekly = bom_req.get("reel_req", {}).get(_rn, {})
        _lt = _c.get("total_lt", 22)
        _cur_stk = st.session_state.get("reel_registry", {}).get(_rm_id_c, st.session_state.get("reel_registry", {}).get(_rn, {})).get("stock_kg", 0)
        _rm_id_ob = _rm_id_c
        _global_rop_ob = global_rm_ss.get(_rm_id_ob, {}).get("rop_kg", 0)
        _ob_date = None
        # Trigger if current stock is already at or below ROP
        if _cur_stk <= _global_rop_ob:
            _ob_date = _today  # order immediately
        else:
            # Trigger when the first week with WO requirements would breach ROP
            for _ws, _we in sorted(_all_weeks, key=lambda x: x[0]):
                _ws_iso = _ws.isoformat()
                if _reel_weekly.get(_ws_iso, 0) > 0:
                    _ob_date = _ws - datetime.timedelta(days=_lt)
                    break
        # Check if any PO is in production for this RM (keyed by rm_material_id)
        _pos_for_reel = st.session_state.get("reel_pos_released", {}).get(_rm_id_c, {})
        _actual_arr_map_ob = st.session_state.get("reel_actual_arrivals", {}).get(_rm_id_c, {})
        _in_transit_pos = []
        for _po_date_str, _po_qty in _pos_for_reel.items():
            try:
                _po_d = datetime.date.fromisoformat(_po_date_str)
                # Use confirmed arrival if set, else default LT
                _po_arr = _actual_arr_map_ob.get(_po_date_str) or \
                          _po_d + datetime.timedelta(days=_c.get("total_lt", 22))
                if _po_d <= _today <= _po_arr:
                    _in_transit_pos.append((_po_d, _po_arr, _po_qty))
            except Exception:
                pass
        _ob_str = _fmt_date(_ob_date) if _ob_date else "✅ No reorder needed"
        _is_urgent = _ob_date and _ob_date <= _today + datetime.timedelta(days=7)
        _is_overdue = _ob_date and _ob_date < _today
        _ob_color = "#DC2626" if _is_overdue else ("#F59E0B" if _is_urgent else "#2563EB")
        _transit_html = ""
        for _pd, _pa, _pq in _in_transit_pos:
            _transit_html += (
                f"<div style='margin-top:5px;font-size:.72rem;color:#16A34A;font-weight:600;'>"
                f"🚚 {int(_pq):,} KG in production — arriving {_pa.strftime('%d %b %Y')}</div>"
            )
        _ob_cards.append((_rm_id_c, _rn, _c, _ob_str, _ob_color, _transit_html))

    # Render cards in rows of 2
    for _i in range(0, len(_ob_cards), 2):
        _batch = _ob_cards[_i:_i+2]
        _cols  = st.columns(len(_batch))
        for _j, (_rm_id_c, _rn, _c, _ob_str, _ob_color, _transit_html) in enumerate(_batch):
            with _cols[_j]:
                # Card title: show rm_material_id if available, else reel_name
                # Component name shown in subtitle
                _card_title = _rm_id_c if _rm_id_c != _rn else _rn
                _card_sub   = f"{_c.get('component', '')} — {_rn[:30]}" if _c.get("component") else _rn[:40]
                st.markdown(
                    f"<div style='border:1px solid {_ob_color}44;border-left:5px solid {_ob_color};"
                    f"border-radius:8px;padding:14px 18px;background:{_ob_color}10;'>"
                    f"<div style='font-size:.72rem;color:{_ob_color};font-weight:700;letter-spacing:.06em;"
                    f"text-transform:uppercase'>📜 {_card_title[:40]} — Order By</div>"
                    f"<div style='font-size:1.3rem;font-weight:700;color:{_ob_color};margin-top:6px'>"
                    f"{_ob_str}</div>"
                    f"<div style='font-size:.72rem;color:#6B7280;margin-top:4px'>"
                    f"{_card_sub} | RM LT: {_c.get('total_lt', 22)}d</div>"
                    f"{_transit_html}"
                    f"</div>",
                    unsafe_allow_html=True,
                )
    st.markdown("<br>", unsafe_allow_html=True)

    def render_rm_section(rm_type: str, rm_label: str, unit: str,
                           kpi_stock, kpi_moq, kpi_rop,
                           detail_table_fn, summary_md: str,
                           raw_override: dict = None,
                           section_key: str = "",
                           other_weekly_kg: dict = None,
                           stock_offset: int = 0):

        # Use component-specific raw if provided, else fall back to raw_data
        _raw = raw_override if raw_override is not None else raw_data

        # Bind the correct per-RM PO dict once, used throughout this function
        # Key by reel name so each component gets its own PO tracking
        if rm_type == "reel":
            _reel_key = _raw.get("reel_name", "reel")          # registry key (shared pool)
            _rm_key   = _raw.get("rm_material_id") or _reel_key # PO key (per unique RM ID)
            rm_po_released = st.session_state.get("reel_pos_released", {}).get(_rm_key, {})
            rm_lt = _raw["reel_total_lt"]
        else:
            _rm_key = _raw.get("pouch_name", "pouch")
            rm_po_released = st.session_state.get("pouch_pos_released", {}).get(_rm_key, {})
            rm_lt = int(_raw["pouch_lt"])

        # Unique prefix for all widget keys in this section
        _wkey = f"{section_key}_{rm_type}_{_rm_key[:15]}"

        k1, k2, k3, k4 = st.columns(4)
        with k1: kpi_card("Current Stock", kpi_stock,
                           color="#3B82F6" if rm_type == "reel" else "#8B5CF6")
        with k2: kpi_card("MOQ",            kpi_moq,  color="#6B7280")
        with k3:
            # Parse numeric ROP from kpi_rop string for color check
            try:
                _rop_num = int(kpi_rop.replace(",","").split()[0])
                _stk_num = int(kpi_stock.replace(",","").replace(" KG","").replace(" units",""))
            except:
                _rop_num = 0; _stk_num = 9999
            _rop_color = "#DC2626" if _stk_num <= _rop_num else "#10B981"
            kpi_card("ROP", kpi_rop, color=_rop_color)
        with k4:
            # Show global SS for this RM
            _rm_id_kpi = _raw.get("rm_material_id") or _raw.get("reel_name", "")
            _ss_kpi = global_rm_ss.get(_rm_id_kpi, {}).get("ss_kg", 0)
            kpi_card("Safety Stock", f"{_ss_kpi:,} KG", color="#F59E0B")

        weekly_df = build_weekly_mrp(
            _raw, r,
            get_issued_wos(),
            rm_type=rm_type,
            po_released=rm_po_released,
            use_raw_stock=False,           # always use registry stock, not raw["reel_stock"]
            other_weekly_kg=other_weekly_kg,
            stock_offset=stock_offset,
            actual_arrivals=st.session_state.get("reel_actual_arrivals", {}).get(_rm_key, {}),
        )

        this_comp_col  = f"This Comp Req ({unit})"
        total_pool_col = f"Total Pool Req ({unit})"
        also_used_col  = f"Also Used ({unit})"
        opening_col    = f"Opening Stock ({unit})"
        net_col        = f"Net Requirement ({unit})"
        closing_col    = f"Closing Stock ({unit})"
        in_transit_col = f"🏭 In Production ({unit})"

        # Format dict — only include cols that exist
        fmt_dict = {
            opening_col:   "{:,}",
            net_col:       "{:,}",
            closing_col:   "{:,}",
        }
        if this_comp_col in weekly_df.columns:
            fmt_dict[this_comp_col] = "{:,}"
        if total_pool_col in weekly_df.columns:
            fmt_dict[total_pool_col] = "{:,}"
        if also_used_col in weekly_df.columns:
            fmt_dict[also_used_col] = lambda v: f"{int(v):,}" if v > 0 else "—"
        if in_transit_col in weekly_df.columns:
            fmt_dict[in_transit_col] = lambda v: f"{int(v):,}" if v > 0 else "—"

        def mrp_row_style(row):
            if row.get("Reorder?") == "🔴 YES":
                return ["background-color:#FEF2F2;color:#DC2626;font-weight:600;color:black;"] * len(row)
            if row.get("Reorder?") == "🚚 PO in production":
                return ["background-color:#EFF6FF;color:black;"] * len(row)
            if "Overdue" in str(row.get("📅 Order By", "")):
                return ["background-color:#FFF7ED;color:black;"] * len(row)
            if row.get(in_transit_col, 0) > 0:
                return ["background-color:#EFF6FF;color:black;"] * len(row)
            return [""] * len(row)

        st.dataframe(
            weekly_df.style.apply(mrp_row_style, axis=1).format(fmt_dict),
            use_container_width=True,
            hide_index=True,
            height=430,
        )

        # ── In-Transit summary ────────────────────────────────
        # Find POs released before today whose stock hasn't arrived yet
        today_check = datetime.date.today()
        in_transit_pos = []
        for rel_iso, qty in rm_po_released.items():
            try:
                rel_d = datetime.date.fromisoformat(rel_iso)
                arr_d = rel_d + datetime.timedelta(days=rm_lt)
                if rel_d <= today_check < arr_d:   # released, not yet arrived
                    in_transit_pos.append((arr_d, qty))
            except:
                pass

        if in_transit_pos:
            total_in_transit = sum(q for _, q in in_transit_pos)
            arriving_strs = [f"{_fmt_date(arr)} ({qty:,} {unit})"
                             for arr, qty in sorted(in_transit_pos)]
            st.info(
                f"🚚 **{total_in_transit:,} {unit} currently in production**  \n"
                + "  \n".join(f"• Arriving {s}" for s in arriving_strs)
            )

        # ── Order By summary ─────────────────────────────────
        overdue_weeks = weekly_df[weekly_df["📅 Order By"].str.contains("overdue", case=False, na=False)]
        urgent_weeks  = weekly_df[
            (weekly_df["Reorder?"] == "🔴 YES") &
            (weekly_df["📅 Order By"] != "—") &
            (~weekly_df["📅 Order By"].str.contains("overdue", case=False, na=False))
        ]
        if not overdue_weeks.empty:
            st.error(
                f"⚠️ **{len(overdue_weeks)} week(s)** where PO should already have been placed! "
                "Consider expediting."
            )
        if not urgent_weeks.empty:
            next_order = urgent_weeks.iloc[0]["📅 Order By"]
            st.warning(f"📅 **Next PO must be placed by: {next_order}** for {rm_label}")

        with st.expander(f"📋 Release a Purchase Order — {rm_label}", expanded=False):
            f1, f2, f3 = st.columns(3)
            with f1:
                po_date = st.date_input(
                    "PO Release Date",
                    value=datetime.date.today(),
                    key=f"po_date_{_wkey}",
                )
            with f2:
                default_moq = int(_raw["reel_moq"] if rm_type == "reel" else _raw["pouch_moq"])
                po_qty = st.number_input(
                    f"Order Qty ({unit})",
                    min_value=1, max_value=10_000_000,
                    value=default_moq, step=100,
                    key=f"po_qty_{_wkey}",
                )
            with f3:
                receiving = po_date + datetime.timedelta(days=rm_lt)
                st.markdown(f"<br>**Receiving Date:** {receiving.strftime('%d %b %Y')}",
                            unsafe_allow_html=True)
                if st.button(f"✅ Release PO", key=f"po_btn_{_wkey}",
                             use_container_width=True, type="primary"):
                    iso = po_date.isoformat()
                    if rm_type == "reel":
                        st.session_state["reel_pos_released"].setdefault(_rm_key, {})
                        st.session_state["reel_pos_released"][_rm_key][iso] = (
                            st.session_state["reel_pos_released"][_rm_key].get(iso, 0) + int(po_qty)
                        )
                    else:
                        st.session_state["pouch_pos_released"].setdefault(_rm_key, {})
                        st.session_state["pouch_pos_released"][_rm_key][iso] = (
                            st.session_state["pouch_pos_released"][_rm_key].get(iso, 0) + int(po_qty)
                        )
                    st.rerun()

        released = rm_po_released
        if released:
            with st.expander(f"📬 Released POs — {rm_label}", expanded=False):
                # Store actual receiving date overrides in session state
                _arr_override_key = f"arr_override_{_wkey}"
                if _arr_override_key not in st.session_state:
                    st.session_state[_arr_override_key] = {}

                for iso, qty in sorted(released.items()):
                    try:
                        rel_d = datetime.date.fromisoformat(iso)
                        default_arr = rel_d + datetime.timedelta(days=rm_lt)
                    except:
                        rel_d = default_arr = datetime.date.today()

                    # Current confirmed arrival (may have been overridden before)
                    _confirmed_arr = st.session_state[_arr_override_key].get(iso, default_arr)

                    c1, c2, c3, c4, c5 = st.columns([2, 2, 2, 2, 1])
                    with c1:
                        st.markdown(f"**Release:** {_fmt_date(rel_d)}")
                    with c2:
                        st.markdown(f"**Qty:** {qty:,} {unit}")
                    with c3:
                        new_arr = st.date_input(
                            "Actual Receiving Date",
                            value=_confirmed_arr,
                            key=f"arr_date_{_wkey}_{iso}",
                            label_visibility="collapsed",
                        )
                    with c4:
                        if _confirmed_arr != default_arr:
                            st.caption(f"📦 Expected: {_fmt_date(default_arr)} → **Confirmed: {_fmt_date(_confirmed_arr)}**")
                        else:
                            st.caption(f"📦 Default arrival: {_fmt_date(default_arr)}")
                        if new_arr != _confirmed_arr:
                            if st.button(f"✅ Confirm {_fmt_date(new_arr)}", key=f"arr_confirm_{_wkey}_{iso}",
                                         type="primary", use_container_width=True):
                                st.session_state[_arr_override_key][iso] = new_arr
                                st.session_state["reel_actual_arrivals"].setdefault(_rm_key, {})
                                st.session_state["reel_actual_arrivals"][_rm_key][iso] = new_arr
                                st.rerun()
                    with c5:
                        if st.button("🗑", key=f"del_po_{_wkey}_{iso}", help="Delete this PO"):
                            if rm_type == "reel":
                                del st.session_state["reel_pos_released"][_rm_key][iso]
                                st.session_state[_arr_override_key].pop(iso, None)
                                st.session_state.get("reel_actual_arrivals", {}).get(_rm_key, {}).pop(iso, None)
                            else:
                                del st.session_state["pouch_pos_released"][_rm_key][iso]
                            st.rerun()

                st.divider()
                if st.button(f"🗑 Clear all {rm_label} POs", key=f"clear_po_{_wkey}"):
                    if rm_type == "reel":
                        st.session_state["reel_pos_released"][_rm_key] = {}
                    else:
                        st.session_state["pouch_pos_released"][_rm_key] = {}
                    st.rerun()

        col_chart, col_detail = st.columns([2, 1])
        with col_chart:
            detail_table_fn()
        with col_detail:
            st.markdown(summary_md)

    # ── Per-component Reel sections start below ──────────────
    bom_comps = raw_data.get("bom_components", [])

    # Pre-compute per-component weekly KG for each reel so we can build "Also Used" totals
    # comp_weekly_kg[comp_idx] = {ws_iso → kg}
    issued_wos_snap = get_issued_wos()
    _today_mrp = datetime.date.today()
    _end_mrp   = _today_mrp + datetime.timedelta(days=183)
    _ws_mrp    = _today_mrp - datetime.timedelta(days=_today_mrp.weekday())
    _weeks_mrp = []
    _w = _ws_mrp
    while _w <= _end_mrp:
        _we = _w + datetime.timedelta(days=6)
        _weeks_mrp.append((_w, min(_we, _end_mrp)))
        _w += datetime.timedelta(days=7)

    comp_weekly_kg: list[dict] = []   # index = comp_idx, value = {ws_iso → kg}
    for comp in bom_comps:
        wk_kg: dict[str, int] = {}
        for _ws, _we in _weeks_mrp:
            _ws_iso = _ws.isoformat()
            _wo_qty = sum(
                info.get("gross_qty", info.get("actual_produced", info.get("wo_qty", 0)))
                for ds, info in issued_wos_snap.items()
                if info.get("issued") and _ws <= datetime.date.fromisoformat(ds) <= _we
            )
            if _wo_qty > 0:
                _sheets = math.ceil(_wo_qty / comp["ups"]) if comp["ups"] else 0
                _kg     = math.ceil(
                    _sheets * comp["width"] * comp["length"]
                    * comp["gsm"] / 1_000_000_000 * (comp.get("multiplier", 1) or 1)
                ) if _sheets else 0
                wk_kg[_ws_iso] = _kg
        comp_weekly_kg.append(wk_kg)

    for comp_idx, comp in enumerate(bom_comps):
        comp_reel_nm  = comp["reel_name"]
        comp_rm_id    = comp.get("rm_material_id") or comp_reel_nm   # RM Material ID for labels
        comp_total_lt = comp["total_lt"]
        comp_moq      = comp["moq_kg"]
        comp_reg_stk_direct = st.session_state.get("reel_registry", {}).get(_reel_reg_key(comp), st.session_state.get("reel_registry", {}).get(comp_reel_nm, {})).get("stock_kg", 0)
        # If this rm_id has 0 stock but other rm_ids share the same reel_name with stock,
        # use the highest-stock entry for that reel_name (same physical paper, different placeholder IDs)
        if comp_reg_stk_direct == 0:
            _same_reel_entries = [
                info.get("stock_kg", 0)
                for key, info in st.session_state.get("reel_registry", {}).items()
                if info.get("reel_name") == comp_reel_nm and info.get("stock_kg", 0) > 0
            ]
            comp_reg_stk = max(_same_reel_entries) if _same_reel_entries else 0
        else:
            comp_reg_stk = comp_reg_stk_direct

        # "Also Used" = sum of KG from all OTHER components sharing the same reel each week
        other_weekly: dict[str, int] = {}
        for other_idx, other_comp in enumerate(bom_comps):
            if other_idx == comp_idx:
                continue
            # Match by rm_material_id so Lid/Strip/Wall (same rm_id) share correctly
            if (other_comp.get("rm_material_id") or other_comp["reel_name"]) !=                (comp.get("rm_material_id") or comp_reel_nm):
                continue
            for ws_iso, kg in comp_weekly_kg[other_idx].items():
                other_weekly[ws_iso] = other_weekly.get(ws_iso, 0) + kg

        # Actual current stock = registry stock minus KG consumed by future WOs
        # that will use THIS specific RM (identified by rm_material_id)
        # Use only current SKU's WOs for the per-component KPI (other SKUs are
        # already reflected in the registry stock uploaded from the RM file)
        _this_comp_rm_id = _reel_reg_key(comp)
        total_consumed_this_reel = sum(
            bom_req.get("reel", {}).get(_this_comp_rm_id,
                bom_req.get("reel", {}).get(comp_reel_nm, {})
            ).get("weekly", {}).values()
        )
        actual_current_stk = max(0, comp_reg_stk - total_consumed_this_reel)

        # For the MRP table opening stock: use actual available (net of all SKU consumption)
        comp_raw = {**raw_data,
                    "rm_material_id": comp.get("rm_material_id", ""),
                    "reel_name":     comp_reel_nm,
                    "reel_stock":    comp_reg_stk,
                    "reel_moq":      comp_moq,
                    "reel_total_lt": comp_total_lt,
                    "reel_ups":      comp["ups"],
                    "reel_gsm":      comp["gsm"],
                    "reel_waste":    comp["wastage"],
                    "reel_width":    comp["width"],
                    "reel_length":   comp["length"],
                    "reel_lt":       comp["lt"],
                    "reel_inbound":  comp["inbound"],
                    "reel_coating":  comp["coating"],
                    "multiplier":    comp.get("multiplier", 1) or 1,
                    }

        st.markdown("---")
        # Show shared-reel note
        # Share = same rm_material_id (not reel_name — multiple specs can share an ID)
        shared_comps = [bom_comps[i]["component"] for i in range(len(bom_comps))
                        if i != comp_idx and
                        (bom_comps[i].get("rm_material_id") or bom_comps[i]["reel_name"]) ==
                        (comp.get("rm_material_id") or comp_reel_nm)]
        shared_note = f" *(shared with: {', '.join(shared_comps)})*" if shared_comps else ""
        st.markdown(f"### 📜 {comp['component']} — {comp_rm_id}{shared_note}")

        comp_reel_summary_md = f"""
| Parameter | Value |
|-----------|-------|
| Component | {comp['component']} |
| GSM / Size | {comp['gsm']} GSM, {int(comp['length'])}×{int(comp['width'])} mm |
| Ups | {comp['ups']} |
| Wastage % | {comp['wastage']}% |
| Supplier LT | {comp['lt']} days |
| Inbound LT | {comp['inbound']} days |
| Coating | {comp['coating']} days |
| **Total RM LT** | **{comp_total_lt} days** |
| MOQ | {int(comp_moq):,} KG |
| Registry Stock (gross) | {int(comp_reg_stk):,} KG |
| WO Consumed (all comps) | {int(total_consumed_this_reel):,} KG |
| **Current Available** | **{int(actual_current_stk):,} KG** |
| Shared With | {', '.join(shared_comps) if shared_comps else 'None'} |
"""

        # KG consumed by OTHER SKUs' WOs on this reel (to offset opening stock for 12018092 etc.)
        other_sku_consumed = max(0, total_consumed_this_reel - sum(
            sum(comp_weekly_kg[i].values())
            for i in range(len(bom_comps))
            if bom_comps[i]["reel_name"] == comp_reel_nm
        ))

        render_rm_section(
            rm_type         = "reel",
            rm_label        = f"{comp['component']} — {comp_rm_id}",
            unit            = "KG",
            kpi_stock       = f"{int(actual_current_stk):,} KG",
            kpi_moq         = f"{int(comp_moq):,} KG",
            kpi_rop         = f"{global_rm_ss.get(comp.get('rm_material_id') or comp_reel_nm, {}).get('rop_kg', r.get('reel_rop_map', {}).get(comp_reel_nm, r['reel_rop_kg'])):,} KG",
            detail_table_fn = lambda c=comp_raw, ci=comp_idx: st.plotly_chart(mrp_reel_chart({**r, "raw": c}), use_container_width=True, key=f"mrp_reel_chart_{ci}"),
            summary_md      = comp_reel_summary_md,
            raw_override    = comp_raw,
            section_key     = f"comp{comp_idx}",
            other_weekly_kg = other_weekly if other_weekly else None,
            stock_offset    = other_sku_consumed,
        )

    _mrp_pouch_stk_raw = st.session_state.get("pouch_registry", {}).get(raw_data["pouch_name"], {}).get("stock_units", raw_data["pouch_stock"])
    # Subtract pouches consumed by confirmed WOs
    _total_wo_qty_pouch = sum(
        info.get("gross_qty", info.get("actual_produced", info.get("wo_qty", 0)))
        for info in get_issued_wos().values() if info.get("issued")
    )
    _pouches_consumed = round(_total_wo_qty_pouch * raw_data.get("pouch_per_1m", 20000) / 1_000_000)
    _mrp_pouch_stk    = max(0, _mrp_pouch_stk_raw - _pouches_consumed)

    # ── POUCH summary (no MRP table) ─────────────────────────
    st.markdown("---")
    st.markdown("### 🛍 Pouch Requirements")
    st.markdown(
        "Pouches are SKU-specific (not planned via MRP). "
        "Required quantity is based on confirmed Work Orders."
    )
    total_wo_qty_pouch = sum(
        info.get("gross_qty", info.get("actual_produced", info.get("wo_qty", 0)))
        for info in get_issued_wos().values() if info.get("issued")
    )
    pouch_nm    = raw_data.get("pouch_name", "—")
    pouch_moq   = raw_data.get("pouch_moq", 20000)
    upc         = raw_data.get("pouch_per_1m", 20000) / 1_000_000
    req_pouches = round(total_wo_qty_pouch * upc) if total_wo_qty_pouch else 0
    pouch_stk   = st.session_state.get("pouch_registry", {}).get(pouch_nm, {}).get("stock_units", 0)
    short       = max(0, req_pouches - pouch_stk)

    pouch_data = {
        "Pouch Name":         pouch_nm,
        "Required Qty":       f"{req_pouches:,}",
        "Current Stock":      f"{int(pouch_stk):,}",
        "MOQ":                f"{int(pouch_moq):,}",
        "Shortfall":          f"🔴 {short:,}" if short > 0 else "✅ Covered",
    }
    c_p1, c_p2, c_p3, c_p4, c_p5 = st.columns(5)
    for col, (lbl, val) in zip([c_p1,c_p2,c_p3,c_p4,c_p5], pouch_data.items()):
        with col:
            kpi_card(lbl, val,
                     color="#DC2626" if "🔴" in str(val) else
                           "#16A34A" if "✅" in str(val) else "#8B5CF6")

    # ── Per-SKU BOM view ─────────────────────────────────────
    st.divider()
    st.subheader("🧩 BOM Explosion — This SKU")
    st.markdown("Component-level RM requirements based on confirmed Work Orders.")

    bom = raw_data.get("bom_components", [])
    total_wo_qty = sum(
        info.get("gross_qty", info.get("actual_produced", info.get("wo_qty", 0)))
        for info in get_issued_wos().values() if info.get("issued")
    )
    if bom and total_wo_qty > 0:
        bom_rows = []
        reel_kg_consumed: dict[str, int] = {}
        for comp in bom:
            sheets = math.ceil(total_wo_qty / comp["ups"]) if comp["ups"] else 0
            kg     = math.ceil(
                sheets * comp["width"] * comp["length"]
                * comp["gsm"] / 1_000_000_000
            ) if sheets else 0
            reg_stk = st.session_state["reel_registry"].get(
                comp["reel_name"], {}).get("stock_kg", 0)
            prior   = reel_kg_consumed.get(comp["reel_name"], 0)
            avail   = max(0, reg_stk - prior)
            reel_kg_consumed[comp["reel_name"]] = prior + kg
            bom_rows.append({
                "Component":       comp["component"],
                "Reel":            comp["reel_name"],
                "Width × Length":  f"{int(comp['width'])} × {int(comp['length'])} mm",
                "GSM":             comp["gsm"],
                "Ups":             comp["ups"],
                "Wastage %":       f"{comp['wastage']}%",
                "Sheets Needed":   f"{sheets:,}",
                "KG Required":     f"{kg:,}",
                "Available Stock": f"{int(avail):,} KG",
                "Net Short?":      "🔴 YES" if avail < kg else "✅ OK",
            })
        bom_df = pd.DataFrame(bom_rows)
        st.dataframe(
            bom_df.style.map(
                lambda v: "color:#DC2626;font-weight:600;" if v == "🔴 YES" else
                          "color:#16A34A;font-weight:600;" if v == "✅ OK" else "",
                subset=["Net Short?"],
            ),
            hide_index=True, use_container_width=True,
        )
    else:
        st.info("Confirm a Work Order in MPS tab to see BOM explosion.")


# ══════════════════════════════════════════════════════════════
# TAB 5 — RM REGISTRY  (consolidated across all SKUs)
# ══════════════════════════════════════════════════════════════
with tab_registry:
    st.subheader("🗄️ RM Registry — Consolidated Across All SKUs")
    st.markdown(
        "This tab shows **shared stock** for all reels and pouches across every loaded SKU.  \n"
        "Edit current stock here — it is used by MRP planning for all SKUs.  \n"
        "POs released here reduce net requirements across all SKUs sharing that RM."
    )

    reg_tab_reel, reg_tab_pouch = st.tabs(["📜 Reels", "🛍 Pouches"])

    # ── REELS ────────────────────────────────────────────────
    with reg_tab_reel:
        st.markdown("#### Current Reel Stock")

        reel_reg = st.session_state["reel_registry"]
        if not reel_reg:
            st.info("No reels registered yet. Upload SKU Excel files to auto-populate.")
        else:
            # reel_registry is now keyed by rm_material_id; each entry has a 'display' field

            # ── Quick PO Entry ────────────────────────────────────
            st.markdown("#### 📦 Log an RM Purchase Order")
            st.caption("Enter a PO you've placed with a supplier — it will appear in the MRP table as inbound stock.")
            _po_c1, _po_c2, _po_c3, _po_c4 = st.columns([3, 1, 1, 1])
            with _po_c1:
                _po_reel_sel = st.selectbox(
                    "Select RM / Reel",
                    options=list(reel_reg.keys()),
                    format_func=lambda n: reel_reg.get(n, {}).get("display", n),
                    key="quick_po_reel_sel",
                )
            with _po_c2:
                _po_qty = st.number_input(
                    "Qty (KG)", min_value=1, max_value=10_000_000,
                    value=int(reel_reg[_po_reel_sel].get("moq", 3000)) if _po_reel_sel else 3000,
                    step=100, key="quick_po_qty",
                )
            with _po_c3:
                _po_order_date = st.date_input("Order Date", value=_today, key="quick_po_order_date")
                _po_lt = reel_reg[_po_reel_sel].get("total_lt", 22) if _po_reel_sel else 22
                _po_arrival = _po_order_date + datetime.timedelta(days=_po_lt)
            with _po_c4:
                st.markdown(f"<br>**Arriving:** {_po_arrival.strftime('%d %b %Y')}", unsafe_allow_html=True)
                if st.button("✅ Log PO", key="quick_po_btn", type="primary", use_container_width=True):
                    _iso = _po_order_date.isoformat()
                    st.session_state["reel_pos_released"][_po_reel_sel][_iso] = (
                        st.session_state["reel_pos_released"][_po_reel_sel].get(_iso, 0) + _po_qty
                    )
                    st.rerun()

            # ── Active PO summary across all reels ───────────────
            # Build complete reverse map: rm_material_id → reel_name from ALL BOM components
            _rm_id_to_reel: dict[str, str] = {}
            for _sl_r, _rd_r in all_results_raw.items():
                for _c_r in _rd_r.get("bom_components", []):
                    _rid_r = (_c_r.get("rm_material_id") or "").strip()
                    _rn_r  = (_c_r.get("reel_name") or "").strip()
                    if _rid_r and _rn_r:
                        _rm_id_to_reel[_rid_r] = _rn_r
            _all_pos_rows = []
            _seen_po_sigs: set = set()
            for _nm, _pos in st.session_state["reel_pos_released"].items():
                if not _pos:
                    continue
                # Resolve rm_material_id → reel_name for display
                # _nm may be rm_material_id — resolve to reel_name via reverse map
                _reel_nm_disp = _nm if _nm in reel_reg else _rm_id_to_reel.get(_nm, _nm)
                _rm_id_disp   = reel_reg.get(_reel_nm_disp, {}).get("rm_material_id", _nm if _nm not in reel_reg else "—")
                _info = reel_reg.get(_reel_nm_disp, {})
                for _ri, _qty in sorted(_pos.items()):
                    try:
                        _rd = datetime.date.fromisoformat(_ri)
                        # Use confirmed arrival override if set
                        _actual_arr_disp = st.session_state.get("reel_actual_arrivals", {}).get(_nm, {}).get(_ri)
                        _ad = _actual_arr_disp if _actual_arr_disp else _rd + datetime.timedelta(days=_info.get("total_lt", 22))
                    except Exception:
                        continue
                    _sig = (_nm, _ri, _qty)   # use raw key (_nm = rm_material_id) to keep distinct rm_ids separate
                    if _sig in _seen_po_sigs:
                        continue
                    _seen_po_sigs.add(_sig)
                    _status = "🏭 In Production" if _rd <= _today <= _ad else ("✅ Arrived" if _today > _ad else "📋 Pending")
                    _all_pos_rows.append({
                        "RM":              f"{_reel_nm_disp[:35]} [{_rm_id_disp}]",
                        "Order Date":      _fmt_date(_rd),
                        "Qty (KG)":        f"{int(_qty):,}",
                        "Expected Arrival":_fmt_date(_rd + datetime.timedelta(days=_info.get("total_lt", 22))),
                        "Confirmed Arrival":_fmt_date(_actual_arr_disp) if _actual_arr_disp else "—",
                        "Status":          _status,
                    })
            if _all_pos_rows:
                st.divider()
                st.markdown("#### 📋 All Active POs")
                _po_summary_df = pd.DataFrame(_all_pos_rows)
                st.dataframe(
                    _po_summary_df.style.map(
                        lambda v: "color:#2563EB;font-weight:600;" if v == "🏭 In Production" else
                                  "color:#16A34A;font-weight:600;" if v == "✅ Arrived" else "",
                        subset=["Status"],
                    ),
                    hide_index=True, use_container_width=True,
                )
                if st.button("🗑 Clear All POs", key="clear_all_pos"):
                    for _nm in st.session_state["reel_pos_released"]:
                        st.session_state["reel_pos_released"][_nm] = {}
                    st.rerun()

            st.divider()
            # ── Editable stock table ──────────────────────────────
            st.markdown("**Update current stock (KG) for each reel:**")
            cols_per_row = 2
            reel_names = list(reel_reg.keys())
            for i in range(0, len(reel_names), cols_per_row):
                batch = reel_names[i:i+cols_per_row]
                cols  = st.columns(cols_per_row)
                for j, nm in enumerate(batch):
                    with cols[j]:
                        _reg_display = st.session_state["reel_registry"].get(nm, {}).get("display", nm)
                        new_stock = st.number_input(
                            f"📜 {_reg_display[:50]}",
                            min_value=0, max_value=10_000_000,
                            value=int(reel_reg[nm].get("stock_kg", 0)),
                            step=100,
                            key=f"reg_reel_stock_{nm}",
                        )
                        st.session_state["reel_registry"][nm]["stock_kg"] = new_stock

            st.divider()
            st.markdown("#### Reel Registry Summary")
            # Ensure reverse map is available even if Active POs section was skipped
            if "_rm_id_to_reel" not in dir():
                _rm_id_to_reel = {}
                for _sl_r2, _rd_r2 in all_results_raw.items():
                    for _c_r2 in _rd_r2.get("bom_components", []):
                        _rid2 = (_c_r2.get("rm_material_id") or "").strip()
                        _rn2  = (_c_r2.get("reel_name") or "").strip()
                        if _rid2 and _rn2:
                            _rm_id_to_reel[_rid2] = _rn2
            summary_rows = []
            _summary_seen: set = set()
            for nm, info in reel_reg.items():
                req_data     = bom_req["reel"].get(nm, {})
                total_req    = sum(req_data.get("weekly", {}).values())
                wo_consumed  = total_req
                actual_avail = max(0, int(info.get("stock_kg", 0)) - wo_consumed)
                comps = ", ".join(
                    f"{c['component']} ({c['sku'].split(chr(8212))[0].strip()})"
                    for c in req_data.get("components", [])
                )
                _base = {
                    "Reel Name":           st.session_state["reel_registry"].get(nm, {}).get("display", nm),
                    "Registry Stock (KG)": f"{int(info.get('stock_kg', 0)):,}",
                    "WO Consumed (KG)":    f"{int(wo_consumed):,}" if wo_consumed else "—",
                    "Available (KG)":      f"{int(actual_avail):,}",
                    "MOQ (KG)":            f"{int(info.get('moq', 0)):,}",
                    "Total LT (days)":     info.get("total_lt", "—"),
                    "Used By":             comps if comps else "—",
                }
                # Find all rm_material_ids that map to this reel_name
                _rm_ids_for_reel = sorted({
                    rid for rid, rn in _rm_id_to_reel.items() if rn == nm
                })
                _added = False
                for _rid in _rm_ids_for_reel:
                    _pos = st.session_state["reel_pos_released"].get(_rid, {})
                    _in_t = 0
                    _po_labels = []
                    for _ri, _qty in _pos.items():
                        try:
                            _rd = datetime.date.fromisoformat(_ri)
                            _arr_ov = st.session_state.get("reel_actual_arrivals", {}).get(_rid, {}).get(_ri)
                            _ad = _arr_ov if _arr_ov else _rd + datetime.timedelta(days=info.get("total_lt", 22))
                            if _rd <= _today < _ad:
                                _in_t += _qty
                                _po_labels.append(f"{int(_qty):,}KG→{_fmt_date(_ad)}")
                        except Exception:
                            pass
                    if _pos or _in_t:
                        row = dict(_base)
                        row["Reel Name"] = f"{nm} [{_rid}]"
                        row["In Transit (KG)"] = (f"{int(_in_t):,} ({', '.join(_po_labels)})" if _po_labels else f"{int(_in_t):,}") if _in_t else "—"
                        if (nm, _rid) not in _summary_seen:
                            _summary_seen.add((nm, _rid))
                            summary_rows.append(row)
                        _added = True
                # Handle POs stored under reel_name key (legacy)
                _pos_rn = st.session_state["reel_pos_released"].get(nm, {})
                if _pos_rn and nm not in _summary_seen:
                    _in_t = sum(
                        qty for ri, qty in _pos_rn.items()
                        if (lambda rd: rd <= _today < (rd + datetime.timedelta(days=info.get("total_lt", 22))))(
                            datetime.date.fromisoformat(ri))
                    )
                    row = dict(_base)
                    row["In Transit (KG)"] = f"{int(_in_t):,}" if _in_t else "—"
                    _summary_seen.add(nm)
                    summary_rows.append(row)
                    _added = True
                if not _added:
                    row = dict(_base)
                    row["In Transit (KG)"] = "—"
                    if nm not in _summary_seen:
                        _summary_seen.add(nm)
                        summary_rows.append(row)
            st.dataframe(pd.DataFrame(summary_rows), hide_index=True, use_container_width=True)

            st.divider()
            st.markdown("#### Consolidated Weekly MRP — All Reels")

            for nm in reel_names:
                req_data = bom_req["reel"].get(nm, {})
                weekly   = req_data.get("weekly", {})
                if not weekly:
                    continue
                info     = reel_reg[nm]
                pos      = st.session_state["reel_pos_released"].get(nm, {})

                avail_stk = max(0, int(info.get('stock_kg', 0)) - sum(bom_req["reel"].get(nm, {}).get("weekly", {}).values()))
                with st.expander(
                    f"📜 **{nm}** — Available: {avail_stk:,} KG (Registry: {int(info.get('stock_kg',0)):,}) | "
                    f"LT: {info.get('total_lt','?')}d | MOQ: {int(info.get('moq',0)):,} KG",
                    expanded=False,
                ):
                    # In-transit banner
                    in_t = [(datetime.date.fromisoformat(ri),
                             ri, qty)
                            for ri, qty in pos.items()
                            if (lambda rd, ld=info.get("total_lt",22):
                                rd <= _today < rd + datetime.timedelta(days=ld))(
                                datetime.date.fromisoformat(ri)
                            )]
                    if in_t:
                        total_it = sum(q for _, _, q in in_t)
                        details  = ", ".join(
                            f"{_fmt_date(rd + datetime.timedelta(days=info.get('total_lt',22)))} ({q:,} KG)"
                            for rd, _, q in sorted(in_t)
                        )
                        st.info(f"🚚 **{total_it:,} KG in production** — arriving: {details}")

                    _po_key_bcm    = st.session_state.get("_reel_rm_id_map", {}).get(nm, nm)
                    _actual_arr_bcm = st.session_state.get("reel_actual_arrivals", {}).get(_po_key_bcm, {})
                    mrp_df = build_consolidated_mrp(nm, "reel", weekly, _all_weeks,
                                                     po_key=_po_key_bcm,
                                                     actual_arrivals=_actual_arr_bcm)
                    gross_c   = "Gross Req (KG)"
                    opening_c = "Opening Stock (KG)"
                    net_c     = "Net Req (KG)"
                    closing_c = "Closing Stock (KG)"
                    transit_c = "🏭 In Production (KG)"

                    def _reel_row_style(row):
                        if row.get("Reorder?") == "🔴 YES":
                            return ["background-color:#FEF2F2;color:#DC2626;font-weight:600;"] * len(row)
                        if row.get("Reorder?") == "🚚 PO in production":
                            return ["background-color:#EFF6FF;"] * len(row)
                        if "overdue" in str(row.get("📅 Order By", "")):
                            return ["background-color:#FFF7ED;"] * len(row)
                        if row.get(transit_c, 0) > 0:
                            return ["background-color:#EFF6FF;"] * len(row)
                        return [""] * len(row)

                    st.dataframe(
                        mrp_df.style.apply(_reel_row_style, axis=1).format({
                            gross_c:   "{:,}", opening_c: "{:,}",
                            net_c:     "{:,}", closing_c: "{:,}",
                            transit_c: lambda v: f"{int(v):,}" if v > 0 else "—",
                        }),
                        hide_index=True, use_container_width=True, height=380,
                    )

                    # PO Release form
                    st.divider()
                    st.markdown(f"**📋 Release PO — {nm[:50]}**")
                    pf1, pf2, pf3 = st.columns(3)
                    with pf1:
                        po_d = st.date_input("PO Date", value=_today,
                                              key=f"reg_reel_po_date_{nm}")
                    with pf2:
                        po_q = st.number_input(
                            "Qty (KG)", min_value=1, max_value=10_000_000,
                            value=int(info.get("moq", 3000)), step=100,
                            key=f"reg_reel_po_qty_{nm}",
                        )
                    with pf3:
                        lt_d = info.get("total_lt", 22)
                        arr  = po_d + datetime.timedelta(days=lt_d)
                        st.markdown(f"<br>**Receiving:** {arr.strftime('%d %b %Y')}",
                                    unsafe_allow_html=True)
                        if st.button("✅ Release PO", key=f"reg_reel_po_btn_{nm}",
                                     use_container_width=True, type="primary"):
                            iso = po_d.isoformat()
                            st.session_state["reel_pos_released"][nm][iso] = (
                                st.session_state["reel_pos_released"][nm].get(iso, 0) + po_q
                            )
                            st.rerun()

                    # Released POs
                    if pos:
                        po_tbl = []
                        for ri, qty in sorted(pos.items()):
                            try:
                                rd = datetime.date.fromisoformat(ri)
                                ad = rd + datetime.timedelta(days=info.get("total_lt", 22))
                            except:
                                rd = ad = None
                            po_tbl.append({
                                "Release Date": _fmt_date(rd),
                                "Qty (KG)":     f"{qty:,}",
                                "Arriving":     _fmt_date(ad),
                            })
                        st.dataframe(pd.DataFrame(po_tbl), hide_index=True,
                                     use_container_width=True)
                        if st.button(f"🗑 Clear POs — {nm[:30]}", key=f"reg_reel_clr_{nm}"):
                            st.session_state["reel_pos_released"][nm] = {}
                            st.rerun()

    # ── POUCHES ──────────────────────────────────────────────
    with reg_tab_pouch:
        st.markdown("#### 🛍 Pouch Summary — All SKUs")
        st.markdown(
            "Pouches are SKU-specific and not planned via MRP. "
            "Update current stock below; required qty is based on confirmed WOs per SKU."
        )

        pouch_reg = st.session_state["pouch_registry"]
        if not pouch_reg:
            st.info("No pouches registered yet. Upload SKU Master to auto-populate.")
        else:
            # Editable stock
            st.markdown("**Update current pouch stock:**")
            p_cols = st.columns(2)
            for pi, (nm, info) in enumerate(pouch_reg.items()):
                with p_cols[pi % 2]:
                    new_stk = st.number_input(
                        f"🛍 {nm[:45]}",
                        min_value=0, max_value=50_000_000,
                        value=int(info.get("stock_units", 0)),
                        step=1000, key=f"reg_pouch_stock_{nm}",
                    )
                    st.session_state["pouch_registry"][nm]["stock_units"] = new_stk

            st.divider()
            st.markdown("#### Summary Table")
            p_summary = []
            for sku_id, raw_s in all_results_raw.items():
                pnm = raw_s.get("pouch_name", "—")
                if not pnm or pnm == "-":
                    continue
                upc_s   = raw_s.get("pouch_per_1m", 20000) / 1_000_000
                wo_qty_s = sum(
                    info.get("gross_qty", info.get("actual_produced", info.get("wo_qty", 0)))
                    for info in st.session_state.get(_sku_key(sku_id, "issued_wos"), {}).values()
                    if info.get("issued")
                )
                req_s   = round(wo_qty_s * upc_s)
                stk_s   = int(st.session_state.get("pouch_registry", {}).get(pnm, {}).get("stock_units", 0))
                short_s = max(0, req_s - stk_s)
                p_summary.append({
                    "SKU":          sku_id,
                    "Pouch Name":   pnm,
                    "WO Qty":       f"{wo_qty_s:,}",
                    "Required":     f"{req_s:,}",
                    "In Stock":     f"{stk_s:,}",
                    "MOQ":          f"{int(raw_s.get('pouch_moq', 20000)):,}",
                    "Status":       f"🔴 Short by {short_s:,}" if short_s > 0 else "✅ OK",
                })
            if p_summary:
                p_df = pd.DataFrame(p_summary)
                st.dataframe(
                    p_df.style.map(
                        lambda v: "color:#DC2626;font-weight:600;" if "🔴" in str(v) else
                                  "color:#16A34A;font-weight:600;" if "✅" in str(v) else "",
                        subset=["Status"],
                    ),
                    hide_index=True, use_container_width=True,
                )


# ══════════════════════════════════════════════════════════════
# TAB 6 — CAPACITY  (per-shift, actual demand)
# ══════════════════════════════════════════════════════════════
with tab_capacity:
    st.subheader("⚙️ Process Capacity Analysis — Monthly View")
    st.markdown(
        "Capacity is evaluated **per shift** against actual demand entered in MPS. "
        "Daily capacity = Shift capacity × Number of shifts."
    )

    if not get_demand_events():
        st.info("Add demand events in MPS tab to see capacity utilisation by month.")
    else:
        # Process summary table
        st.markdown("#### Process Configuration")
        proc_tbl = []
        for proc in raw_data["processes"]:
            proc_tbl.append({
                "Process":          proc["name"],
                "Unit":             proc["cap_unit"],
                "Shifts/Day":       proc["shifts"],
                "Capacity/Shift":   f"{proc['shift_cap']:,}",
                "Daily Capacity":   f"{proc['capacity']:,}",
                "Prod Duration":    f"{proc['days']}d",
            })
        st.dataframe(pd.DataFrame(proc_tbl), hide_index=True, use_container_width=True)

        st.divider()
        st.markdown("#### Monthly Utilisation by Process")
        st.plotly_chart(
            capacity_chart_monthly(cap_res, raw_data["processes"]),
            use_container_width=True
        , key="plotly_4")

        # Monthly detail table
        st.markdown("#### Monthly Capacity Details")
        cap_month_rows = []
        for row in cap_res["cap_rows"]:
            entry = {
                "Month":        row["month"],
                "FG Demand":    f"{row['month_dem']:,}",
                "Avg Daily FG": f"{row['daily_dem']:,}",
                "Sheets/Day":   f"{row['sheets_daily']:,}",
            }
            for proc in raw_data["processes"]:
                pu = row["proc_utils"][proc["name"]]
                entry[f"{proc['name']} (util%)"] = f"{pu['util_pct']:.1f}%"
                entry[f"{proc['name']} demand"]  = f"{pu['daily_dem']:,}"
            cap_month_rows.append(entry)

        st.dataframe(pd.DataFrame(cap_month_rows), hide_index=True, use_container_width=True)

        st.divider()
        # Per-process breakdown with shift-level detail
        st.markdown("#### Per-Process Shift Analysis")
        for ps in cap_res["proc_summary"]:
            color = "#EF4444" if ps["max_util"] > 90 else "#F59E0B" if ps["max_util"] > 60 else "#10B981"
            with st.expander(
                f"**{ps['name']}** — Max util: {ps['max_util']:.1f}% | Avg: {ps['avg_util']:.1f}% | "
                f"{ps['shifts']} shifts × {ps['shift_cap']:,} {ps['cap_unit']}/shift",
                expanded=False,
            ):
                c1, c2, c3, c4 = st.columns(4)
                with c1: kpi_card("Shifts/Day", str(ps["shifts"]), color=color)
                with c2: kpi_card(f"Cap/Shift ({ps['cap_unit']})", f"{ps['shift_cap']:,}", color="#2563EB")
                with c3: kpi_card("Peak Utilisation", f"{ps['max_util']:.1f}%", color=color)
                with c4: kpi_card("Avg Utilisation", f"{ps['avg_util']:.1f}%", color="#6B7280")

                # Month-by-month for this process
                month_utils = [
                    {
                        "Month": row["month"],
                        "FG Demand": f"{row['month_dem']:,}",
                        f"Daily Demand ({ps['cap_unit']})": f"{row['proc_utils'][ps['name']]['daily_dem']:,}",
                        f"Shift Cap ({ps['cap_unit']})": f"{row['proc_utils'][ps['name']]['shift_cap']:,}",
                        f"Daily Cap ({ps['cap_unit']})": f"{row['proc_utils'][ps['name']]['daily_cap']:,}",
                        "Utilisation %": f"{row['proc_utils'][ps['name']]['util_pct']:.2f}%",
                        "Status": "🔴 Over" if row['proc_utils'][ps['name']]['util_pct'] > 90
                                  else "🟡 Caution" if row['proc_utils'][ps['name']]['util_pct'] > 60
                                  else "🟢 OK",
                    }
                    for row in cap_res["cap_rows"]
                ]
                st.dataframe(pd.DataFrame(month_utils), hide_index=True, use_container_width=True)

        st.divider()
        st.subheader("Production Lead Time Breakdown")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Process sequence:**")
            for p in raw_data["processes"]:
                bar = "█" * p["days"]
                st.markdown(f"- {p['name']}: **{p['days']}d** {bar}")
            st.markdown(f"- **Total: {raw_data['prod_lt']} days**")
        with col2:
            st.markdown("**Bottleneck (peak utilisation):**")
            st.markdown(f"→ **{cap_res['bottleneck']}**")


# ══════════════════════════════════════════════════════════════
# TAB 6 — WO LOG
# ══════════════════════════════════════════════════════════════
with tab_wo:
    st.subheader("📋 Work Order History")

    if r["wo_df"].empty:
        st.info("No WO history in this Excel file. Fill the WO Log sheet.")
    else:
        wo = r["wo_df"]
        actual_lts = wo["Actual LT (days)"].dropna().astype(float)
        avg_lt  = round(actual_lts.mean(), 1) if len(actual_lts) else "—"
        on_time = (wo["On Time?"] == "✅ Yes").sum()
        total   = len(wo)

        c1, c2, c3, c4 = st.columns(4)
        with c1: kpi_card("Total WOs", str(total), color="#2563EB")
        with c2: kpi_card("Avg Lead Time", f"{avg_lt}d",
                          help_text=f"Target: {raw_data['prod_lt']}d", color="#F59E0B")
        with c3: kpi_card("On-Time WOs", f"{on_time}/{total}", color="#10B981")
        with c4:
            avg_yield = wo["Yield %"].mean() if "Yield %" in wo.columns else 0
            kpi_card("Avg Yield %", f"{avg_yield:.1f}%",
                     help_text=f"Waste target: {raw_data['wastage_pct']}%", color="#8B5CF6")

        st.divider()
        st.dataframe(
            wo.style.map(
                lambda v: "color:#DC2626;font-weight:600;" if v == "❌ No" else
                          "color:#16A34A;font-weight:600;" if v == "✅ Yes" else "",
                subset=["On Time?"]
            ).format({"Target": "{:,}", "Produced": "{:,}",
                      "Waste %": "{:.1f}%", "Yield %": "{:.1f}%"}),
            use_container_width=True, hide_index=True,
        )

        if len(actual_lts) > 0:
            st.subheader("Lead Time vs. Target")
            fig_lt = go.Figure()
            fig_lt.add_trace(go.Scatter(
                x=wo["WO #"], y=actual_lts,
                mode="markers+lines", name="Actual LT",
                marker=dict(size=10, color="#3B82F6"),
            ))
            fig_lt.add_hline(y=raw_data["prod_lt"], line_dash="dash",
                             line_color="#EF4444",
                             annotation_text=f"Target LT ({raw_data['prod_lt']}d)")
            fig_lt.update_layout(
                height=280, yaxis_title="Days",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig_lt, use_container_width=True, key="plotly_5")


def run_forecasts(monthly_series: pd.Series, n_months: int = 6) -> dict:
    """
    Run 4 forecasting models on a monthly series.
    Returns dict of model_name → list of (period_label, forecast_qty)
    """
    import numpy as np
    results = {}
    vals = monthly_series.values.astype(float)
    n    = len(vals)

    # ── 1. Moving Average (3-month) ──────────────────────────
    window = min(3, n)
    ma_base = float(np.mean(vals[-window:]))
    results["Moving Average (3M)"] = [ma_base] * n_months

    # ── 2. Exponential Smoothing (α=0.3) ─────────────────────
    alpha = 0.3
    es_val = vals[0]
    for v in vals[1:]:
        es_val = alpha * v + (1 - alpha) * es_val
    results["Exponential Smoothing"] = [es_val] * n_months

    # ── 3. Holt-Winters (Double Exponential — trend) ─────────
    if n >= 3:
        alpha_h, beta_h = 0.3, 0.1
        level = vals[0]
        trend = (vals[-1] - vals[0]) / max(n - 1, 1)
        for v in vals[1:]:
            prev_level = level
            level = alpha_h * v + (1 - alpha_h) * (level + trend)
            trend = beta_h * (level - prev_level) + (1 - beta_h) * trend
        hw = [max(0, level + (i + 1) * trend) for i in range(n_months)]
        results["Holt-Winters (Trend)"] = hw
    else:
        results["Holt-Winters (Trend)"] = [ma_base] * n_months

    # ── 4. Linear Regression ─────────────────────────────────
    if n >= 2:
        x = np.arange(n)
        m_coef, b_coef = np.polyfit(x, vals, 1)
        lr = [max(0, m_coef * (n + i) + b_coef) for i in range(n_months)]
        results["Linear Regression"] = lr
    else:
        results["Linear Regression"] = [ma_base] * n_months

    return results


# ══════════════════════════════════════════════════════════════
# TAB 8 — DEMAND FORECAST
# ══════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════
# TAB 8 — DEMAND VARIABILITY
# ══════════════════════════════════════════════════════════════
with tab_variability:
    st.subheader("📉 Demand Variability Analysis")
    st.markdown(
        "Monthly demand variability per SKU from PO history. "
        "Used to compute statistical Safety Stock once ≥10 months of data is available. "
        "**Formula:** SS = Z × √(LT × σ_demand² + D² × σ_LT²)"
    )

    if po_df_global.empty:
        st.info("Upload the Billing Excel in the sidebar (②) to analyse demand variability.")
    else:
        known_skus = set(all_results_raw.keys())
        var_df     = po_df_global[po_df_global["mat_id"].isin(known_skus)].copy()

        if var_df.empty:
            st.warning("No matching SKUs between PO history and SKU Master.")
        else:
            var_df["month"] = var_df["date"].dt.to_period("M")
            # Build complete month list covering full date range — no gaps
            # (some SKUs may have 0 orders in a month but month still exists)
            _min_month = var_df["month"].min()
            _max_month = var_df["month"].max()
            all_months_list = []
            _m = _min_month
            while _m <= _max_month:
                all_months_list.append(_m)
                _m = _m + 1
            first_po_date   = var_df["date"].min().date()
            months_of_data  = (datetime.date.today() - first_po_date).days / 30
            data_sufficient = months_of_data >= 10

            st.info(
                f"📅 Data from **{first_po_date.strftime('%b %Y')}** to "
                f"**{var_df['date'].max().strftime('%b %Y')}** — "
                f"**{months_of_data:.1f} months**. "
                + ("✅ Sufficient for demand variance SS formula." if data_sufficient
                   else f"⚠️ Need {10 - months_of_data:.1f} more months before demand variance is used in SS formula.")
            )

            # ── Per-SKU variability table ─────────────────────
            st.markdown("#### Per-SKU Monthly Demand Statistics")

            var_rows = []
            for mid in sorted(var_df["mat_id"].unique()):
                mts_target = all_results_raw.get(mid, {}).get("target_fg", 0)
                sku_monthly = (
                    var_df[var_df["mat_id"] == mid]
                    .groupby("month")["qty"].sum()
                    .reindex(all_months_list, fill_value=0)
                )
                vals      = sku_monthly.values.astype(float)
                n         = len(vals)
                mean_d    = float(np.mean(vals)) if n > 0 else 0

                # Raw demand stats
                var_d     = float(np.var(vals, ddof=1)) if n > 1 else 0
                sigma_d   = float(np.std(vals, ddof=1)) if n > 1 else 0
                cv        = (sigma_d / mean_d * 100) if mean_d > 0 else 0

                # Deviation from MTS stats — std dev of |actual − MTS|
                abs_deviations = np.abs(vals - mts_target)
                sigma_dev  = float(np.std(abs_deviations, ddof=1)) if n > 1 else 0
                mean_abs_dev = float(np.mean(abs_deviations)) if n > 0 else 0
                max_dev    = float(np.max(abs_deviations)) if n > 0 else 0
                mean_dev   = float(np.mean(vals - mts_target)) if n > 0 else 0

                month_vals = {p.strftime("%b %Y"): int(v) for p, v in zip(all_months_list, vals)}

                row = {
                    "SKU":                    mid,
                    "Description":            all_results_raw.get(mid, {}).get("mat_desc", "")[:35],
                    "MTS Target":             mts_target,
                    "Months":                 n,
                    "Mean Demand":            f"{mean_d:,.0f}",
                    "Mean vs MTS":            f"{mean_dev:+,.0f}",
                    "σ (raw demand)":         f"{sigma_d:,.0f}",
                    "Mean |dev from MTS|":    f"{mean_abs_dev:,.0f}",
                    "σ |dev from MTS|":       f"{sigma_dev:,.0f}",
                    "Max |deviation|":        f"{max_dev:,.0f}",
                    "CV (σ/mean %)":          f"{cv:.1f}%",
                    "SS Formula":             "Demand σ active" if data_sufficient else f"LT σ only (need {max(0,10-months_of_data):.1f} more months)",
                }
                row.update(month_vals)
                var_rows.append(row)

            var_summary_df = pd.DataFrame(var_rows)

            # Style: highlight high CV
            def _var_style(row):
                try:
                    cv_val = float(str(row.get("CV (σ/mean %)", "0")).replace("%",""))
                    if cv_val > 50:
                        return ["background-color:#FEE2E2;color:black;"] * len(row)
                    elif cv_val > 25:
                        return ["background-color:#FEF9C3;color:black;"] * len(row)
                except:
                    pass
                return [""] * len(row)

            summary_cols = ["SKU", "Description", "MTS Target", "Months",
                            "Mean Demand", "Mean vs MTS",
                            "σ (raw demand)", "Mean |dev from MTS|",
                            "σ |dev from MTS|", "Max |deviation|",
                            "CV (σ/mean %)", "SS Formula"]
            st.dataframe(
                var_summary_df[summary_cols].style.apply(_var_style, axis=1),
                use_container_width=True, hide_index=True,
            )

            # ── Monthly breakdown per SKU ─────────────────────
            st.divider()
            st.markdown("#### Monthly Demand Breakdown — All SKUs")

            month_cols  = [p.strftime("%b %Y") for p in all_months_list]
            display_cols = ["SKU", "Description", "MTS Target"] + month_cols + ["Mean Demand", "σ |dev from MTS|", "CV (σ/mean %)"]
            available   = [c for c in display_cols if c in var_summary_df.columns]
            st.dataframe(var_summary_df[available], use_container_width=True, hide_index=True)

            # ── Per-SKU detail chart ──────────────────────────
            st.divider()
            st.markdown("#### Variability Chart — Select SKU")
            sel_sku = st.selectbox(
                "SKU", sorted(var_df["mat_id"].unique()),
                format_func=lambda x: f"{x} — {all_results_raw.get(x,{}).get('mat_desc','')[:40]}",
                key="var_sku_select",
            )
            sel_monthly = (
                var_df[var_df["mat_id"] == sel_sku]
                .groupby("month")["qty"].sum()
                .reindex(all_months_list, fill_value=0)
            )
            sel_vals  = sel_monthly.values.astype(float)
            sel_mean  = float(np.mean(sel_vals))
            sel_sigma = float(np.std(sel_vals, ddof=1)) if len(sel_vals) > 1 else 0
            sel_labels = [p.strftime("%b %Y") for p in all_months_list]

            fig_var = go.Figure()
            fig_var.add_trace(go.Bar(
                x=sel_labels, y=sel_vals.tolist(),
                name="Monthly Demand", marker_color="#3B82F6", opacity=0.8,
            ))
            fig_var.add_hline(y=sel_mean, line_color="#10B981", line_width=2,
                              annotation_text=f"Mean: {sel_mean:,.0f}",
                              annotation_font_color="#10B981")
            fig_var.add_hline(y=sel_mean + sel_sigma, line_color="#F59E0B",
                              line_dash="dot", line_width=1.5,
                              annotation_text=f"Mean+σ: {sel_mean+sel_sigma:,.0f}",
                              annotation_font_color="#F59E0B")
            fig_var.add_hline(y=max(0, sel_mean - sel_sigma), line_color="#F59E0B",
                              line_dash="dot", line_width=1.5,
                              annotation_text=f"Mean-σ: {max(0,sel_mean-sel_sigma):,.0f}",
                              annotation_font_color="#F59E0B")
            mts = all_results_raw.get(sel_sku, {}).get("target_fg", 0)
            if mts:
                fig_var.add_hline(y=mts, line_color="#6B7280", line_dash="dash",
                                  line_width=1.5,
                                  annotation_text=f"MTS: {mts:,}",
                                  annotation_font_color="#6B7280")
            fig_var.update_layout(
                title=f"Monthly Demand — {sel_sku}",
                xaxis_title="Month", yaxis_title="Cartons",
                plot_bgcolor="#0F172A", paper_bgcolor="#0F172A",
                font=dict(color="#E2E8F0"), height=380,
            )
            fig_var.update_xaxes(gridcolor="#1E293B")
            fig_var.update_yaxes(gridcolor="#1E293B")
            st.plotly_chart(fig_var, use_container_width=True, key="plotly_6")

            # ── SS formula preview ────────────────────────────
            st.divider()
            st.markdown("#### Safety Stock Formula Preview")
            st.markdown(
                "**SS = Z × √(LT × σ_demand² + D² × σ_LT²)**  \n"
                "where D = daily avg KG from MTS, "
                "**σ_demand = std dev of |actual monthly demand − MTS|** converted to daily KG, "
                "σ_LT = √(LT variability days), LT = total RM lead time.  \n"
                "*σ_demand measures how consistently demand hits the MTS target. "
                "If demand is always exactly MTS, σ_demand = 0 → SS = Z×D×σ_LT only.*"
            )

            ss_rows = []
            for mid in sorted(var_df["mat_id"].unique()):
                raw_s    = all_results_raw.get(mid, {})
                z_score  = raw_s.get("z_score", 1.65)
                lt_var   = raw_s.get("lt_variability", 5)
                sigma_lt = lt_var

                sku_monthly_s = (
                    var_df[var_df["mat_id"] == mid]
                    .groupby("month")["qty"].sum()
                    .reindex(all_months_list, fill_value=0)
                )
                vals_s  = sku_monthly_s.values.astype(float)
                n_s     = len(vals_s)
                tgt_s   = raw_s.get("target_fg", 0)

                # σ of |actual − MTS|: std dev of absolute deviations from planning target
                abs_dev_s = np.abs(vals_s - tgt_s)
                sigma_d_cartons = float(np.std(abs_dev_s, ddof=1)) if n_s > 1 else 0

                # Use MTS-based daily KG as proxy D
                bom_s    = raw_s.get("bom_components", [])
                for comp_s in bom_s[:1]:  # first component as representative
                    rm_lt_s  = comp_s["total_lt"]
                    sheets_s = math.ceil(tgt_s / comp_s["ups"]) if comp_s["ups"] else 0
                    kg_s     = math.ceil(
                        sheets_s * comp_s["width"] * comp_s["length"]
                        * comp_s["gsm"] / ((1 - comp_s["wastage"] / 100) * 1_000_000_000)
                    ) if sheets_s else 0
                    daily_d  = kg_s / 30

                    # Convert demand σ from cartons to KG (proportional to MTS KG/carton ratio)
                    daily_sigma_kg = (sigma_d_cartons / 30) * (kg_s / max(tgt_s, 1)) if tgt_s else 0

                    if data_sufficient and daily_sigma_kg > 0:
                        ss_formula = math.ceil(z_score * math.sqrt(
                            rm_lt_s * daily_sigma_kg**2 + daily_d**2 * sigma_lt**2
                        ))
                        formula_used = "Z×√(LT×σ_d²+D²×σ_LT²)"
                    else:
                        ss_formula = math.ceil(z_score * daily_d * sigma_lt)
                        formula_used = "Z×D×σ_LT (demand σ pending)"

                    ss_rows.append({
                        "SKU":          mid,
                        "Reel":         comp_s["reel_name"][:30],
                        "D (daily KG)": f"{daily_d:.1f}",
                        "σ_demand (KG/day)": f"{daily_sigma_kg:.2f}" if data_sufficient else "—",
                        "σ_LT (days)":  f"{sigma_lt:.2f}",
                        "LT (days)":    rm_lt_s,
                        "Z-score":      z_score,
                        "SS (KG)":      f"{ss_formula:,}",
                        "Formula":      formula_used,
                    })
                    break

            if ss_rows:
                st.dataframe(pd.DataFrame(ss_rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# TAB 9 — DEMAND FORECAST
# ══════════════════════════════════════════════════════════════
with tab_forecast:
    st.subheader("📈 Demand Forecasting")
    st.markdown(
        "Upload PO history to generate 6-month demand forecasts per SKU using "
        "4 models: Moving Average, Exponential Smoothing, Holt-Winters, and Linear Regression."
    )

    if not po_hist_file:
        st.info("Upload the Billing Excel in the sidebar (②) to generate forecasts.")
    else:
        try:
            po_df = po_df_global.copy() if not po_df_global.empty else pd.DataFrame()
        except Exception as e:
            st.error(f"Error reading PO history: {e}")
            po_df = pd.DataFrame()

        if po_df.empty:
            st.warning("No valid data found in the uploaded file.")
        else:
            # Filter to SKUs in the master
            known_skus = set(all_results_raw.keys())
            po_df = po_df[po_df["mat_id"].isin(known_skus)]

            if po_df.empty:
                st.warning("No matching SKUs found between PO history and SKU Master.")
            else:
                # Build monthly pivot
                po_df["month"] = po_df["date"].dt.to_period("M")
                monthly_pivot  = (
                    po_df.groupby(["mat_id","month"])["qty"]
                    .sum()
                    .reset_index()
                )
                monthly_pivot["month_str"] = monthly_pivot["month"].astype(str)

                # Build complete month range — no gaps between first and last month
                _all_min = monthly_pivot["month"].min()
                _all_max = monthly_pivot["month"].max()
                all_months = []
                _am = _all_min
                while _am <= _all_max:
                    all_months.append(_am)
                    _am = _am + 1
                last_month = all_months[-1] if all_months else pd.Period("2026-04","M")
                forecast_periods = [last_month + i for i in range(1, 7)]
                forecast_labels  = [p.strftime("%b %Y") for p in forecast_periods]
                hist_labels      = [p.strftime("%b %Y") for p in all_months]

                st.success(
                    f"✅ Loaded {len(po_df):,} PO lines | "
                    f"{po_df['mat_id'].nunique()} SKUs | "
                    f"{all_months[0].strftime('%b %Y')} → {last_month.strftime('%b %Y')}"
                )

                # SKU selector
                fc_sku = st.selectbox(
                    "Select SKU to forecast",
                    options=sorted(po_df["mat_id"].unique()),
                    format_func=lambda x: f"{x} — {all_results_raw.get(x,{}).get('mat_desc','')[:40]}",
                    key="fc_sku_select",
                )

                sku_monthly = (
                    monthly_pivot[monthly_pivot["mat_id"] == fc_sku]
                    .set_index("month")["qty"]
                    .reindex(all_months, fill_value=0)
                )

                # Aggregate demand events for selected SKU across all SKUs (monthly)
                demand_monthly: dict[str, float] = {}
                for sl, rd in all_results_raw.items():
                    if sl != fc_sku:
                        continue
                    for ev in st.session_state.get(_sku_key(sl, "demand_events"), []):
                        try:
                            ev_period = pd.Period(ev["date"].strftime("%Y-%m"), "M")
                            ev_label  = ev_period.strftime("%b %Y")
                            demand_monthly[ev_label] = demand_monthly.get(ev_label, 0) + ev["qty"]
                        except:
                            pass

                n_hist = len(sku_monthly)
                if n_hist < 2:
                    st.warning(f"Only {n_hist} month(s) of data for this SKU — forecasts may be unreliable.")

                forecasts = run_forecasts(sku_monthly, n_months=6)

                # ── Chart ─────────────────────────────────────────────
                import plotly.graph_objects as go

                fig = go.Figure()

                # Historical actuals (PO history)
                fig.add_trace(go.Bar(
                    x=hist_labels,
                    y=sku_monthly.values.tolist(),
                    name="Historical Actuals (PO History)",
                    marker_color="#3B82F6",
                    opacity=0.8,
                ))

                # Demand events overlay (from MPS tab)
                if demand_monthly:
                    dem_x = list(demand_monthly.keys())
                    dem_y = [demand_monthly[k] for k in dem_x]
                    fig.add_trace(go.Bar(
                        x=dem_x,
                        y=dem_y,
                        name="Demand Events (MPS)",
                        marker_color="#F97316",
                        opacity=0.85,
                    ))

                # Forecast lines
                colors = {
                    "Moving Average (3M)":     "#F59E0B",
                    "Exponential Smoothing":   "#10B981",
                    "Holt-Winters (Trend)":    "#EF4444",
                    "Linear Regression":       "#8B5CF6",
                }
                for model_name, fc_vals in forecasts.items():
                    fig.add_trace(go.Scatter(
                        x=forecast_labels,
                        y=[round(v) for v in fc_vals],
                        mode="lines+markers",
                        name=model_name,
                        line=dict(color=colors[model_name], width=2, dash="dot"),
                        marker=dict(size=7),
                    ))

                # MTS reference line
                mts = all_results_raw.get(fc_sku, {}).get("monthly_demand", 0)
                if mts:
                    all_x = hist_labels + forecast_labels
                    fig.add_trace(go.Scatter(
                        x=all_x,
                        y=[mts] * len(all_x),
                        mode="lines",
                        name=f"MTS Target ({mts:,})",
                        line=dict(color="#6B7280", width=1.5, dash="dash"),
                    ))

                fig.update_layout(
                    title=f"Demand Forecast — {fc_sku}",
                    xaxis_title="Month",
                    yaxis_title="Quantity (Cartons)",
                    barmode="overlay",
                    plot_bgcolor="#0F172A",
                    paper_bgcolor="#0F172A",
                    font=dict(color="#E2E8F0"),
                    legend=dict(
                        orientation="h", yanchor="bottom", y=1.02,
                        xanchor="right", x=1,
                    ),
                    height=450,
                    bargap=0.3,
                )
                fig.update_xaxes(gridcolor="#1E293B", showgrid=True)
                fig.update_yaxes(gridcolor="#1E293B", showgrid=True)
                st.plotly_chart(fig, use_container_width=True, key="plotly_7")

                # ── Forecast table ────────────────────────────────────
                st.markdown("#### Forecast Summary (Next 6 Months)")
                fc_table_rows = []
                for i, (period, label) in enumerate(zip(forecast_periods, forecast_labels)):
                    row = {"Month": label}
                    for model_name, fc_vals in forecasts.items():
                        row[model_name] = int(round(fc_vals[i]))
                    fc_table_rows.append(row)

                fc_df = pd.DataFrame(fc_table_rows)
                st.dataframe(fc_df.set_index("Month"), use_container_width=True)

                # ── Historical actuals table ──────────────────────────
                with st.expander("📋 Historical Monthly Actuals", expanded=False):
                    hist_df = pd.DataFrame({
                        "Month":    hist_labels,
                        "Quantity": sku_monthly.values.tolist(),
                    })
                    st.dataframe(hist_df.set_index("Month"), use_container_width=True)

                # ── All-SKU summary ───────────────────────────────────
                st.divider()
                st.markdown("#### All SKUs — 6-Month Moving Average Forecast")
                all_fc_rows = []
                for mid in sorted(po_df["mat_id"].unique()):
                    s = (
                        monthly_pivot[monthly_pivot["mat_id"] == mid]
                        .set_index("month")["qty"]
                        .reindex(all_months, fill_value=0)
                    )
                    if len(s) == 0:
                        continue
                    fc_vals = run_forecasts(s, 6)["Moving Average (3M)"]
                    row_d = {
                        "SKU":         mid,
                        "Description": all_results_raw.get(mid, {}).get("mat_desc","")[:40],
                        "MTS Target":  all_results_raw.get(mid, {}).get("monthly_demand", 0),
                    }
                    for i, lbl in enumerate(forecast_labels):
                        row_d[lbl] = int(round(fc_vals[i]))
                    all_fc_rows.append(row_d)

                if all_fc_rows:
                    all_fc_df = pd.DataFrame(all_fc_rows).set_index("SKU")
                    st.dataframe(all_fc_df, use_container_width=True)


# ══════════════════════════════════════════════════════════════
# TAB 10 — DELIVERY
# ══════════════════════════════════════════════════════════════
with tab_delivery:
    st.subheader("🚚 Delivery Tracker")

    _dtk_df: pd.DataFrame = st.session_state.get("_del_track_df", pd.DataFrame())

    if _dtk_df.empty:
        st.info("Upload the Delivery Tracking Excel from the **sidebar (⑤)** to see pending shipments.")
    else:
        def _col(df, *kws):
            for kw in kws:
                m = next((c for c in df.columns if kw.lower() in c.lower()), None)
                if m: return m
            return None

        C_BILL    = _col(_dtk_df, "bill.no", "bill no")
        C_GST     = _col(_dtk_df, "gst invoice")
        C_BILLDT  = _col(_dtk_df, "bill.date", "bill date")
        C_MAT     = _col(_dtk_df, "material code")
        C_DESC    = _col(_dtk_df, "material description")
        C_QTY     = _col(_dtk_df, "billed quantity")
        C_CUSTPO  = _col(_dtk_df, "customer po no")
        C_TRANS   = _col(_dtk_df, "transporter name")
        C_LR      = _col(_dtk_df, "lr number")
        C_DEST    = _col(_dtk_df, "ship to destination")
        C_PLANT   = _col(_dtk_df, "prod.plant", "plant")
        C_SOLD    = _col(_dtk_df, "sold-to-party name")
        C_DAYS    = _col(_dtk_df, "no. of days", "days to tci")
        C_FP      = _col(_dtk_df, "factory pickup")
        C_TCI     = _col(_dtk_df, "actual tci delivery")
        C_SCHED   = _col(_dtk_df, "scheduled hyperpure", "sched")
        C_HP      = _col(_dtk_df, "actual hyperpure delivery", "actual hyperpure")

        # Classify every row
        def _classify(row):
            tci = str(row[C_TCI] if C_TCI else "").strip().lower()
            hp  = str(row[C_HP]  if C_HP  else "").strip().lower()
            if hp not in ("", "in transit", "not arrived"):
                return "arrived"
            if tci == "in transit":
                return "tci_transit"
            return "tci_done"   # reached TCI DC, not yet at HP DC

        _dtk_df = _dtk_df.copy()
        _dtk_df["_status"] = _dtk_df.apply(_classify, axis=1)

        _df_tci    = _dtk_df[_dtk_df["_status"] == "tci_transit"]
        _df_hp     = _dtk_df[_dtk_df["_status"] == "tci_done"]
        _df_arr    = _dtk_df[_dtk_df["_status"] == "arrived"]

        # KPIs
        _k1,_k2,_k3,_k4 = st.columns(4)
        with _k1: kpi_card("Total Lines",              f"{len(_dtk_df):,}",   color="#3B82F6")
        with _k2: kpi_card("✅ Arrived at HP DC",       f"{len(_df_arr):,}",   color="#10B981")
        with _k3: kpi_card("🚛 In Transit to TCI DC",  f"{len(_df_tci):,}",   color="#F59E0B")
        with _k4: kpi_card("📦 Sitting at TCI DC",     f"{len(_df_hp):,}",    color="#DC2626")

        st.divider()

        # Display helpers
        _SHOW = [c for c in [C_BILL,C_GST,C_BILLDT,C_MAT,C_DESC,C_QTY,C_CUSTPO,
                               C_TRANS,C_LR,C_DEST,C_PLANT,C_DAYS,C_FP,
                               C_TCI,C_SCHED,C_HP] if c]
        _REN  = {C_BILL:"Bill No.",C_GST:"GST Invoice",C_BILLDT:"Bill Date",
                 C_MAT:"Material",C_DESC:"Description",C_QTY:"Qty",
                 C_CUSTPO:"Customer PO",C_TRANS:"Transporter",C_LR:"LR Number",
                 C_DEST:"Destination",C_PLANT:"Plant",C_DAYS:"Days to TCI",
                 C_FP:"Factory Pickup",C_TCI:"Actual TCI Delivery",
                 C_SCHED:"Scheduled HP Delivery",C_HP:"Actual HP Delivery"}
        _DATE_DISP = {"Bill Date","Factory Pickup","Actual TCI Delivery",
                      "Scheduled HP Delivery","Actual HP Delivery"}

        def _parse_date_disp(v):
            """Format a date value as dd-mm-yyyy; preserve text like 'in transit'."""
            # Handle datetime/Timestamp objects directly — no string conversion needed
            if isinstance(v, (datetime.date, datetime.datetime)):
                return v.strftime("%d-%m-%Y")
            if hasattr(v, 'strftime'):
                return v.strftime("%d-%m-%Y")
            s = str(v).strip()
            if s.lower() in ("", "in transit", "not arrived", "nat", "nan", "none"):
                return s
            # Already dd-mm-yyyy format
            if len(s) == 10 and s[2] == '-' and s[5] == '-':
                return s
            # Try explicit dd-mm-yyyy parse first, then others
            for _fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
                try:
                    return datetime.datetime.strptime(s[:10], _fmt).strftime("%d-%m-%Y")
                except Exception:
                    pass
            # Last resort: pandas with dayfirst (only for ambiguous strings)
            try:
                _parsed = pd.to_datetime(s, dayfirst=True, errors="coerce")
                if pd.notna(_parsed):
                    return _parsed.strftime("%d-%m-%Y")
            except Exception:
                pass
            return s

        def _fmt_df(df):
            d = df[[c for c in _SHOW if c in df.columns]].copy()
            d = d.rename(columns={k:v for k,v in _REN.items() if k})
            for dc in _DATE_DISP:
                if dc in d.columns:
                    d[dc] = d[dc].apply(_parse_date_disp)
            return d.reset_index(drop=True)

        # Updates stored as {row_idx: {"action": "tci"|"hp", "date": "dd-mm-yyyy"}}
        if "del_arrival_updates" not in st.session_state:
            st.session_state["del_arrival_updates"] = {}

        def _apply_updates_to_df(df):
            """Apply pending updates to a working copy so UI reflects them immediately."""
            df = df.copy()
            for _idx, _upd in st.session_state["del_arrival_updates"].items():
                if not isinstance(_upd, dict) or _idx >= len(df): continue
                if _upd.get("action") == "tci" and C_TCI:
                    df.at[_idx, C_TCI] = _upd["date"]
                    # HP column should become "in transit" once goods reach TCI DC
                    if C_HP:
                        _hp_cur = str(df.at[_idx, C_HP]).strip().lower()
                        if _hp_cur in ("", "not arrived"):
                            df.at[_idx, C_HP] = "in transit"
                elif _upd.get("action") == "hp" and C_HP:
                    df.at[_idx, C_HP] = _upd["date"]
                    if C_TCI and str(df.at[_idx, C_TCI]).strip().lower() == "in transit":
                        df.at[_idx, C_TCI] = _upd["date"]
            df["_status"] = df.apply(_classify, axis=1)
            return df

        # Re-apply any pending updates so sections show current state
        _dtk_working = _apply_updates_to_df(_dtk_df)
        _df_tci  = _dtk_working[_dtk_working["_status"] == "tci_transit"]
        _df_hp   = _dtk_working[_dtk_working["_status"] == "tci_done"]
        _df_arr  = _dtk_working[_dtk_working["_status"] == "arrived"]

        # ── Helper: group by LR → nested expanders ────────────
        def _lr_section(df_section, action, date_label, btn_label, section_key):
            if df_section.empty:
                st.caption("None")
                return
            _lrs = df_section[C_LR].unique() if C_LR else ["—"]
            for _lr in _lrs:
                _lr_rows = df_section[df_section[C_LR] == _lr] if C_LR else df_section
                _dest    = str(_lr_rows[C_DEST].iloc[0]) if C_DEST else ""
                _origin  = str(_lr_rows[C_PLANT].iloc[0]) if C_PLANT else ""
                _idxs    = list(_lr_rows.index)
                _n       = len(_lr_rows)
                # Scheduled TCI delivery = Factory Pickup date + No. of Days to TCI
                _sched_tci_str = ""
                try:
                    _fp_val   = _lr_rows[C_FP].iloc[0] if C_FP else None
                    _days_val = _lr_rows[C_DAYS].iloc[0] if C_DAYS else None
                    if _fp_val is not None and _days_val is not None:
                        _fp_d  = (_fp_val if isinstance(_fp_val, (datetime.date, datetime.datetime))
                                  else pd.to_datetime(str(_fp_val), dayfirst=True, errors="coerce"))
                        _days_n = int(float(str(_days_val))) if str(_days_val).strip() not in ("","nan") else 0
                        if pd.notna(_fp_d) and _days_n:
                            _sched_tci = ((_fp_d.date() if isinstance(_fp_d, datetime.datetime) else _fp_d)
                                          + datetime.timedelta(days=_days_n))
                            _sched_tci_str = f" &nbsp;|&nbsp; 📅 Sched. TCI: **{_sched_tci.strftime('%d %b %Y')}**"
                except Exception:
                    pass
                st.markdown(
                    f"**LR {_lr}** &nbsp;|&nbsp; {_origin} → {_dest}"
                    f"{_sched_tci_str} &nbsp;|&nbsp; {_n} line(s)"
                )
                st.dataframe(_fmt_df(_lr_rows), use_container_width=True, hide_index=True)
                _c1, _c2 = st.columns([2, 1])
                with _c1:
                    _chosen_date = st.date_input(
                        date_label,
                        value=datetime.date.today(),
                        key=f"{section_key}_date_{_lr}",
                    )
                with _c2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button(btn_label, key=f"{section_key}_btn_{_lr}",
                                 type="primary", use_container_width=True):
                        for _i in _idxs:
                            st.session_state["del_arrival_updates"][_i] = {
                                "action": action,
                                "date":   _chosen_date.strftime("%d-%m-%Y"),
                            }
                        st.rerun()
                st.divider()

        # ── Section 1: In Transit to TCI DC ──────────────────
        with st.expander(f"🚛 In Transit to TCI DC — {len(_df_tci)} line(s)  ▼", expanded=False):
            _lr_section(_df_tci, "tci", "📅 Date arrived at TCI DC",
                        "✅ Mark arrived at TCI DC", "sec_tci")

        # ── Section 2: Sitting at TCI DC ─────────────────────
        with st.expander(f"📦 Sitting at TCI DC (not yet at HP DC) — {len(_df_hp)} line(s)  ▼", expanded=False):
            _lr_section(_df_hp, "hp", "📅 Date arrived at HP DC",
                        "✅ Mark arrived at HP DC", "sec_hp")

        # ── Arrived (reference) ───────────────────────────────
        with st.expander(f"✅ Arrived at HP DC — {len(_df_arr)} line(s)  ▼", expanded=False):
            if not _df_arr.empty:
                st.dataframe(_fmt_df(_df_arr), use_container_width=True, hide_index=True)
            else:
                st.caption("None yet.")

        # ── Save & Download ───────────────────────────────────
        _changes = {k: v for k, v in st.session_state.get("del_arrival_updates", {}).items()
                    if isinstance(v, dict)}
        if _changes:
            st.divider()
            st.info(f"{len(_changes)} line(s) with pending updates.")
            if st.button("💾 Generate Updated Tracking Excel",
                         type="primary", use_container_width=True):
                _out = _dtk_df.drop(columns=["_status"], errors="ignore").copy()
                for _idx, _upd in _changes.items():
                    if _idx >= len(_out): continue
                    if _upd.get("action") == "tci" and C_TCI:
                        _out.at[_idx, C_TCI] = _upd["date"]
                        if C_HP:
                            _hp_cur = str(_out.at[_idx, C_HP]).strip().lower()
                            if _hp_cur in ("", "not arrived"):
                                _out.at[_idx, C_HP] = "in transit"
                    elif _upd.get("action") == "hp" and C_HP:
                        _out.at[_idx, C_HP] = _upd["date"]
                        if C_TCI and str(_out.at[_idx, C_TCI]).strip().lower() == "in transit":
                            _out.at[_idx, C_TCI] = _upd["date"]
                from io import BytesIO as _BIO
                _buf = _BIO()
                with pd.ExcelWriter(_buf, engine="openpyxl") as _w:
                    _out.to_excel(_w, index=False, sheet_name="Delivery Tracking")
                _ts = datetime.date.today().strftime("%d-%m-%Y")
                st.session_state["_del_excel_bytes"]    = _buf.getvalue()
                st.session_state["_del_excel_filename"] = f"delivery_tracking_{_ts}.xlsx"
                st.session_state["_del_track_df"]       = _out
                st.session_state["del_arrival_updates"] = {}

        # Persistent download button — survives reruns
        if st.session_state.get("_del_excel_bytes"):
            st.download_button(
                "⬇️ Download Updated Tracking File",
                data=st.session_state["_del_excel_bytes"],
                file_name=st.session_state.get("_del_excel_filename", "delivery_tracking.xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="del_download_btn",
            )
            if st.button("✅ Done — clear download", use_container_width=True, key="del_clear_btn"):
                st.session_state.pop("_del_excel_bytes", None)
                st.rerun()


with tab_outsource:
    st.subheader("🏭 Outsourced SKUs")
    st.caption(
        "These SKUs are procured from external vendors (not produced in-house). "
        "Stock is read from the RM inventory file using the SKU Material ID as the RM Material ID. "
        "Safety Stock and ROP are calculated from the Outsource sheet in SKU Master."
    )

    _os_skus = raw_data.get("outsource_skus", [])
    if not _os_skus:
        st.info("No outsourced SKUs found. Add them to the 'Outsource' sheet in SKU Master.")
    else:
        # Read current stock from RM inventory — SKU mat_id = RM material ID
        _rm_inv_cache = st.session_state.get("_cached_rm_inv", {})

        # Build demand history from billing file for each outsource SKU
        _billing_df = st.session_state.get("_cached_demand_df", pd.DataFrame())

        # KPI summary row
        _os_rows = []
        for _osku in _os_skus:
            _mid  = _osku["mat_id"]
            _stock = _rm_inv_cache.get(_mid, 0)
            _ss    = _osku["safety_stock"]
            _rop   = _osku["rop"]
            _mts   = _osku["monthly_demand"]
            _status = ("🔴 Below ROP"  if _stock <= _rop  else
                       "🟡 Below MTS"  if _stock < _mts   else "✅ OK")

            # Recent demand from billing
            _hist_qty = 0
            if not _billing_df.empty and "mat_id" in _billing_df.columns:
                _hist_rows = _billing_df[_billing_df["mat_id"] == _mid]
                _hist_qty  = int(_hist_rows["qty"].sum()) if not _hist_rows.empty else 0

            _os_rows.append({
                "Material ID":        _mid,
                "Description":        _osku["mat_desc"][:55],
                "Factory":            _osku["factory"],
                "Vendor Location":    _osku["place"],
                "Monthly Demand":     f"{_mts:,}",
                "Inbound LT (days)":  _osku["inbound_lt"],
                "LT Variability σ":   _osku["lt_var"],
                "Safety Stock":       f"{_ss:,}",
                "ROP":                f"{_rop:,}",
                "Current Stock":      f"{int(_stock):,}" if _stock else "0",
                "Historical Demand":  f"{_hist_qty:,}" if _hist_qty else "—",
                "Status":             _status,
            })

        _os_df = pd.DataFrame(_os_rows)

        # KPI cards
        _os_below_rop = sum(1 for r in _os_rows if "Below ROP" in r["Status"])
        _os_ok        = sum(1 for r in _os_rows if r["Status"] == "✅ OK")
        _k1, _k2, _k3 = st.columns(3)
        with _k1: kpi_card("Total Outsourced SKUs", str(len(_os_skus)), color="#3B82F6")
        with _k2: kpi_card("Below ROP",             str(_os_below_rop), color="#DC2626" if _os_below_rop else "#10B981")
        with _k3: kpi_card("Sufficient Stock",       str(_os_ok),        color="#10B981")

        st.divider()
        st.markdown("#### SKU Details & Stock Status")

        def _os_style(v):
            if "Below ROP"  in str(v): return "color:#DC2626;font-weight:700"
            if "Below MTS"  in str(v): return "color:#D97706;font-weight:600"
            if "✅ OK"       in str(v): return "color:#16A34A;font-weight:600"
            return ""

        st.dataframe(
            _os_df.style.map(_os_style, subset=["Status"]),
            hide_index=True, use_container_width=True,
        )

        # ── Per-SKU demand history ────────────────────────────
        if not _billing_df.empty and "mat_id" in _billing_df.columns:
            _os_mat_ids_in_billing = [
                o["mat_id"] for o in _os_skus
                if o["mat_id"] in _billing_df["mat_id"].values
            ]
            if _os_mat_ids_in_billing:
                st.divider()
                st.markdown("#### 📦 Demand History (from Billing File)")
                _sel_os = st.selectbox(
                    "Select SKU",
                    options=[o["mat_id"] for o in _os_skus],
                    format_func=lambda m: f"{m} — {next((o['mat_desc'][:45] for o in _os_skus if o['mat_id']==m), '')}",
                    key="outsource_sku_sel",
                )
                _os_billing = _billing_df[_billing_df["mat_id"] == _sel_os].copy()
                if _os_billing.empty:
                    st.info("No billing records found for this SKU.")
                else:
                    _os_billing_disp = _os_billing[[
                        c for c in ["date", "invoice_no", "customer_po", "qty",
                                    "transporter", "destination", "bill_date", "lr_number"]
                        if c in _os_billing.columns
                    ]].rename(columns={
                        "date": "PO Date", "invoice_no": "Invoice", "customer_po": "Customer PO",
                        "qty": "Qty (Cartons)", "transporter": "Delivery Partner",
                        "destination": "Destination", "bill_date": "Bill Date",
                        "lr_number": "LR Number",
                    })
                    for _dc in ["PO Date", "Bill Date"]:
                        if _dc in _os_billing_disp.columns:
                            _os_billing_disp[_dc] = _os_billing_disp[_dc].apply(
                                lambda v: v.strftime("%d %b %Y") if pd.notna(v) and hasattr(v, "strftime") else "—"
                            )
                    st.dataframe(_os_billing_disp.reset_index(drop=True), use_container_width=True, hide_index=True)
                    st.caption(f"Total ordered: **{int(_os_billing['qty'].sum()):,} cartons** across {len(_os_billing)} invoices")
